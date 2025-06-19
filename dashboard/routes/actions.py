# dashboard/routes/actions.py

import random
import string
import requests
import logging
import threading
import datetime
import pytz
import re
import io
import json
import os # For SMTP env vars
import itertools
from jinja2 import Template
import smtplib # For SMTP
from email.mime.text import MIMEText # For email body
from email.mime.multipart import MIMEMultipart # For email structure
from flask import (
    Blueprint, request, redirect, url_for, flash, current_app, jsonify, render_template, send_file
)
from pymongo.errors import PyMongoError, BulkWriteError
import pymongo
from pymongo import ASCENDING, UpdateOne
from collections import defaultdict
from typing import List, Dict, Optional, Tuple, Any, Union, Set
# Import extensions, utils, tasks from dashboard package
from ..extensions import (
    scheduler,
    db,
    interim_alloc_collection, # For saving review data
    alloc_collection,         # For toggle_power and teardown_tag
    trainer_email_collection,
    host_collection,          # For getting host list in build_review
    course_config_collection,  # For getting memory in build_review
    build_rules_collection
)
from ..email_utils import send_allocation_email
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Color
from openpyxl.utils import get_column_letter
import math
from ..utils import (build_args_from_form, parse_command_line, 
                     update_power_state_in_db, get_hosts_available_memory_parallel)
from ..tasks import run_labbuild_task

# Import top-level utils if needed (e.g., delete_from_database)
from config_utils import get_host_by_name
from db_utils import delete_from_database
from bson.errors import InvalidId
from bson import ObjectId

try:
    from zoneinfo import ZoneInfo, ZoneInfoNotFoundError
except ImportError:
    # Fallback for Python < 3.9 or if zoneinfo data isn't available
    from datetime import timezone as ZoneInfo # Use basic UTC offset if zoneinfo missing

from constants import SUBSEQUENT_POD_MEMORY_FACTOR

# Define Blueprint
bp = Blueprint('actions', __name__)
logger = logging.getLogger('dashboard.routes.actions')

# --- Email Generation Logic (Moved to a Helper) ---
def _generate_email_previews(all_review_items: List[Dict]) -> List[Dict[str, Any]]:
    """
    Server-side logic to generate DATA for trainer email previews.
    It consolidates data, creating multi-line strings for split-host allocations,
    and ensures all numeric values are correctly calculated and passed.
    """
    email_data_list = []
    
    student_items = [
        item for item in all_review_items 
        if item.get("type") == "Student Build" and item.get("sf_trainer_name") and item.get("assignments")
    ]
    if not student_items:
        return []

    try:
        all_hosts_set = {asgn.get("host") for item in student_items for asgn in item.get("assignments", []) if asgn.get("host")}
        hosts_info_map = {h["host_name"]: h for h in host_collection.find({"host_name": {"$in": list(all_hosts_set)}})}
    except PyMongoError as e:
        logger.error(f"Failed to fetch host details for email previews: {e}")
        hosts_info_map = {}

    emails_grouped = defaultdict(list)
    for item in student_items:
        key = f"{item.get('sf_trainer_name', 'N/A')}|{item.get('original_sf_course_code', 'N/A')}"
        emails_grouped[key].append(item)

    for key, items in emails_grouped.items():
        trainer_name, sf_code = key.split('|')
        first_item = items[0]
        
        # --- Group assignments by host ---
        assignments_by_host = defaultdict(list)
        for item in items:
            for asgn in item.get("assignments", []):
                host = asgn.get("host")
                if host:
                    assignments_by_host[host].append(asgn)

        # --- Generate multi-line strings for display ---
        host_lines = []
        vcenter_lines = []
        pod_range_lines = []
        total_pods_for_course = set() # Use a set to handle potential overlaps
        total_ram_for_course = 0.0

        # ** THIS IS THE KEY CORRECTION BLOCK **
        for host, host_assignments in sorted(assignments_by_host.items()):
            host_pod_numbers = {p for asgn in host_assignments for p in range(int(asgn.get('start_pod')), int(asgn.get('end_pod')) + 1) if asgn.get('start_pod') is not None}
            total_pods_for_course.update(host_pod_numbers)
            
            pod_numbers_sorted = sorted(list(host_pod_numbers))
            groups = [list(g) for k, g in itertools.groupby(enumerate(pod_numbers_sorted), lambda x: x[0]-x[1])]
            host_start_end_pod_str = ", ".join([f"{r[0][1]}-{r[-1][1]}" if len(r) > 1 else str(r[0][1]) for r in groups])

            host_info = hosts_info_map.get(host, {})
            num_pods_on_host = len(host_pod_numbers)
            
            # Use the memory from the first item as it's consistent for the course
            memory_per_pod = float(first_item.get('memory_gb_one_pod', 0.0))
            ram_for_this_host = memory_per_pod * num_pods_on_host
            
            # This was missing in the previous version, causing the error.
            # Now we correctly add this host's RAM to the grand total for the course.
            total_ram_for_course += ram_for_this_host 

            host_lines.append(f"{host} ({host_start_end_pod_str})")
            vcenter_lines.append(host_info.get("vcenter", "N/A"))
            pod_range_lines.append(host_start_end_pod_str)
        # ** END OF CORRECTION BLOCK **
            
        virtual_host_display = "\n".join(host_lines)
        vcenter_display = "\n".join(vcenter_lines)
        pod_range_display = "\n".join(pod_range_lines)

        start_date_str, end_date_str = first_item.get("start_date"), first_item.get("end_date")
        date_range_display, end_day_abbr = "N/A", "N/A"
        try:
            start_dt = datetime.datetime.fromisoformat(start_date_str.replace('Z', '+00:00'))
            end_dt = datetime.datetime.fromisoformat(end_date_str.replace('Z', '+00:00'))
            start_day, end_day = start_dt.strftime("%a"), end_dt.strftime("%a")
            date_range_display = f"{start_day}-{end_day}" if start_day != end_day else start_day
            end_day_abbr = end_day
        except (ValueError, TypeError): pass

        email_data_list.append({
            "key": key,
            "trainer_name": trainer_name,
            "sf_course_code": sf_code,
            "email_subject": f"Lab Allocation for {sf_code}",
            "payload_items": items,
            "template_data": {
                "original_sf_course_code": sf_code,
                "date_range_display": date_range_display,
                "end_day_abbr": end_day_abbr,
                "primary_location": "Virtual",
                "sf_course_type": first_item.get('sf_course_type', 'N/A'),
                "start_end_pod_str": pod_range_display,
                "username": first_item.get("apm_username", "N/A"),
                "password": first_item.get("apm_password", "UseProvidedPassword"),
                "effective_pods_req": len(total_pods_for_course), # Use the size of the final set
                "final_labbuild_course": first_item.get('labbuild_course', 'N/A'),
                "virtual_host_display": virtual_host_display,
                "primary_vcenter": vcenter_display,
                "total_ram_for_course": total_ram_for_course # Pass the calculated total RAM
            }
        })
    return email_data_list

def _get_memory_for_course_local(course_name: str, local_course_configs_map: Dict[str, Any]) -> float:
    if not course_name: return 0.0
    config = local_course_configs_map.get(course_name)
    if config:
        mem_gb_str = config.get('memory_gb_per_pod', config.get('memory', '0'))
        try:
            mem_gb = float(str(mem_gb_str).strip() or '0')
            return mem_gb if mem_gb > 0 else 0.0
        except (ValueError, TypeError):
            logger.warning(f"Could not parse memory '{mem_gb_str}' for course '{course_name}'.")
            return 0.0
    return 0.0

def _get_memory_for_course(course_name: str) -> float:
    """Fetches memory_gb_per_pod for a given course name."""
    if course_config_collection is None: return 0.0
    try:
        config = course_config_collection.find_one({"course_name": course_name})
        if config:
            # --- Adjust this key based on your actual course config structure ---
            mem_gb = float(config.get('memory', 0))
            return mem_gb if mem_gb > 0 else 0.0
            # --- Or implement component summing logic here ---
        return 0.0
    except Exception as e:
        logger.error(f"Error getting memory for course '{course_name}': {e}")
        return 0.0

# --- NEW Route to Generate Email Previews ---
@bp.route('/prepare-email-previews', methods=['POST'])
def prepare_email_previews():
    """
    Server-side route to generate all email previews.
    Takes the full review data and returns structured HTML for the modal.
    """
    all_review_items = request.json.get('all_review_items')
    if not isinstance(all_review_items, list):
        return jsonify({"error": "Invalid data format."}), 400
    
    email_previews = _generate_email_previews(all_review_items)
    
    return jsonify({"previews": email_previews})


# --- UPDATED Route to Send Emails ---
@bp.route('/send-trainer-email', methods=['POST'])
def send_trainer_email():
    data = request.json
    trainer_name = data.get('trainer_name')
    course_items_for_email = data.get('course_item_to_email')
    edited_subject = data.get('edited_subject')
    edited_html_body = data.get('edited_html_body')

    if not all([trainer_name, isinstance(course_items_for_email, list), course_items_for_email, edited_subject, edited_html_body]):
        return jsonify({"status": "error", "message": "Missing required data for sending email."}), 400

    trainer_email_doc = trainer_email_collection.find_one({"trainer_name": trainer_name, "active": True})
    if not trainer_email_doc or not trainer_email_doc.get("email_address"):
        return jsonify({"status": "error", "message": f"Active email address not found for trainer '{trainer_name}'."}), 404
    to_email_address = trainer_email_doc.get("email_address")
    
    # We pass course_items_for_email only for the plain-text fallback generation.
    success, message = send_allocation_email(
        to_address=to_email_address,
        trainer_name=trainer_name,
        subject=edited_subject,
        course_allocations_data=course_items_for_email, # For plain-text fallback
        html_body_override=edited_html_body, # The user's preview is now the source of truth
        is_test=False
    )

    if success:
        return jsonify({"status": "success", "message": message})
    else:
        return jsonify({"status": "error", "message": f"Email sending failed: {message}"}), 500


@bp.route('/send-test-email', methods=['POST'])
def send_test_email():
    data = request.json
    trainer_name = data.get('trainer_name')
    course_items = data.get('course_item_to_email')
    edited_subject = data.get('edited_subject')
    edited_html_body = data.get('edited_html_body')

    if not all([trainer_name, isinstance(course_items, list), course_items, edited_subject, edited_html_body]):
        return jsonify({"status": "error", "message": "Missing required data for sending test email."}), 400
    
    success, message = send_allocation_email(
        to_address="placeholder@example.com",
        trainer_name=trainer_name,
        subject=edited_subject,
        course_allocations_data=course_items, # For plain-text fallback
        html_body_override=edited_html_body,
        is_test=True
    )

    if success:
        return jsonify({"status": "success", "message": message})
    else:
        return jsonify({"status": "error", "message": f"Test email failed: {message}"}), 500

@bp.route('/run', methods=['POST'])
def run_now():
    """Handle immediate run request from the main form."""
    form_data = request.form.to_dict()
    args_list, error_msg = build_args_from_form(form_data) # Use form helper

    if error_msg:
        flash(f'Invalid form data: {error_msg}', 'danger')
        return redirect(url_for('main.index')) # Redirect to main blueprint index
    if not args_list:
         flash('Failed to build command arguments.', 'danger')
         return redirect(url_for('main.index'))

    try:
        # Use daemon thread for background execution
        thread = threading.Thread(target=run_labbuild_task, args=(args_list,), daemon=True)
        thread.start()
        flash(f"Submitted immediate run: {' '.join(args_list)}", 'info')
    except Exception as e:
        logger.error(f"Failed start thread run_now: {e}", exc_info=True)
        flash("Error starting background task.", 'danger')
    return redirect(url_for('main.index'))


@bp.route('/schedule', methods=['POST'])
def schedule_run():
    """Handle schedule run request from the main form."""
    form_data = request.form.to_dict()
    args_list, error_msg = build_args_from_form(form_data) # Use form helper

    if error_msg:
        flash(f'Invalid form data: {error_msg}', 'danger')
        return redirect(url_for('main.index'))
    if not args_list:
         flash('Failed to build command arguments.', 'danger')
         return redirect(url_for('main.index'))

    if not scheduler or not scheduler.running:
        flash("Scheduler not running. Cannot schedule job.", "danger")
        return redirect(url_for('main.index'))

    # Extract schedule details
    schedule_type = form_data.get('schedule_type', 'date')
    schedule_time_str = form_data.get('schedule_time')
    cron_expression = form_data.get('cron_expression')
    interval_value = form_data.get('interval_value')
    interval_unit = form_data.get('interval_unit', 'minutes')

    try:
        job_name = f"{form_data.get('command')}_{form_data.get('vendor')}_{form_data.get('course', 'N/A')}"
        trigger = None
        flash_msg = ""
        log_time_str = "N/A"

        # Import trigger types here
        from apscheduler.triggers.date import DateTrigger
        from apscheduler.triggers.cron import CronTrigger
        from apscheduler.triggers.interval import IntervalTrigger

        # Get scheduler's timezone
        scheduler_tz = scheduler.timezone

        if schedule_type == 'date' and schedule_time_str:
            naive_dt = datetime.datetime.fromisoformat(schedule_time_str)
            # *** MODIFIED TIMEZONE HANDLING ***
            try:
                # Make aware using the scheduler's timezone object
                aware_dt = naive_dt.replace(tzinfo=scheduler_tz)
            except TypeError:
                 # Fallback for older pytz versions if scheduler_tz is pytz
                 if hasattr(scheduler_tz, 'localize'):
                      aware_dt = scheduler_tz.localize(naive_dt)
                 else: # Could not make aware, proceed with naive (might be UTC if lucky)
                      aware_dt = naive_dt
                      logger.warning(f"Could not make datetime timezone-aware for scheduling: {naive_dt}. Assuming UTC or scheduler default.")

            run_date_utc = aware_dt.astimezone(datetime.timezone.utc) # Always convert to UTC for DateTrigger
            # *** END MODIFICATION ***
            trigger = DateTrigger(run_date=run_date_utc)
            flash_msg = f"for {aware_dt.strftime('%Y-%m-%d %H:%M:%S %Z%z')}" # Display in scheduler's TZ
            log_time_str = f"UTC: {run_date_utc.isoformat()}, SchedulerTZ: {aware_dt.isoformat()}"

        elif schedule_type == 'cron' and cron_expression:
            parts = cron_expression.split()
            if len(parts) != 5: raise ValueError("Invalid cron format (must have 5 parts).")
            trigger = CronTrigger.from_crontab(cron_expression, timezone=scheduler_tz)
            flash_msg = f"with cron: '{cron_expression}' ({scheduler_tz})"
            log_time_str = flash_msg

        elif schedule_type == 'interval' and interval_value:
             interval_val_int = int(interval_value)
             if interval_val_int < 1: raise ValueError("Interval must be >= 1.")
             kwargs = {interval_unit: interval_val_int}
             trigger = IntervalTrigger(**kwargs)
             flash_msg = f"every {interval_val_int} {interval_unit}"
             log_time_str = flash_msg
        else:
             flash('Invalid schedule details provided.', 'danger')
             return redirect(url_for('main.index'))

        # Add job
        job = scheduler.add_job(
            run_labbuild_task, trigger=trigger, args=[args_list],
            name=job_name, misfire_grace_time=3600, replace_existing=False
        )
        flash(f"Scheduled job '{job.id}' {flash_msg}", 'success')
        logger.info(f"Job '{job.id}' ({job_name}) scheduled. Trigger info: {log_time_str}")

    except ValueError as ve: logger.error(f"Invalid schedule input: {ve}", exc_info=True); flash(f"Invalid schedule input: {ve}", 'danger')
    except Exception as e: logger.error(f"Failed schedule job: {e}", exc_info=True); flash(f"Error scheduling job: {e}", 'danger')

    return redirect(url_for('main.index'))

@bp.route('/schedule-batch', methods=['POST'])
def schedule_batch():
    """Handle schedule batch run request from the main form."""
    # Logic moved from original app.py
    import io # Need io for TextIOWrapper
    import werkzeug.utils # For secure_filename

    # --- File Handling ---
    if 'batch_file' not in request.files: flash('No file part.', 'danger'); return redirect(url_for('main.index'))
    file = request.files['batch_file']
    if file.filename == '': flash('No file selected.', 'danger'); return redirect(url_for('main.index'))

    # --- Time/Delay Parsing ---
    start_time_str = request.form.get('start_time')
    delay_minutes_str = request.form.get('delay_minutes', '30')
    if not start_time_str: flash('Start time required for batch.', 'danger'); return redirect(url_for('main.index'))
    try: delay_minutes = int(delay_minutes_str); assert delay_minutes >= 1
    except (ValueError, AssertionError): flash('Invalid delay minutes (must be >= 1).', 'danger'); return redirect(url_for('main.index'))

    # --- Scheduler Check ---
    if not scheduler or not scheduler.running: flash("Scheduler not running.", "danger"); return redirect(url_for('main.index'))

    # --- Parse Start Time (using scheduler timezone) ---
    try:
        from apscheduler.triggers.date import DateTrigger # Import here
        scheduler_tz = scheduler.timezone
        naive_dt = datetime.datetime.fromisoformat(start_time_str)
        # *** MODIFIED TIMEZONE HANDLING ***
        try:
            # Make the naive datetime aware using the scheduler's timezone
            first_run_local = naive_dt.replace(tzinfo=scheduler_tz)
        except TypeError:
            # Fallback for older pytz versions if scheduler_tz is pytz
            if hasattr(scheduler_tz, 'localize'):
                 first_run_local = scheduler_tz.localize(naive_dt)
            else: # Could not make aware, proceed with naive (less ideal)
                 first_run_local = naive_dt
                 logger.warning(f"Could not make batch start time timezone-aware: {naive_dt}. Assuming UTC or scheduler default.")

        # Calculate subsequent run times in UTC for DateTrigger
        first_run_utc = first_run_local.astimezone(datetime.timezone.utc)
        # *** END MODIFICATION ***
        logger.info(f"Batch Schedule: First job UTC start: {first_run_utc} (from local input {first_run_local})")
    except Exception as e: logger.error(f"Err parse batch start: {e}", exc_info=True); flash("Error processing start time.", "danger"); return redirect(url_for('main.index'))

    # --- Process File ---
    scheduled_count, failed_lines = 0, 0
    current_run_time_utc = first_run_utc
    job_delay = datetime.timedelta(minutes=delay_minutes)
    filename = werkzeug.utils.secure_filename(file.filename)

    try:
        logger.info(f"Processing batch file: {filename}")
        stream = io.TextIOWrapper(file.stream, encoding='utf-8')
        lines = stream.readlines()

        for i, line in enumerate(lines):
            args_list = parse_command_line(line) # Use helper from utils
            if args_list:
                run_date_utc = current_run_time_utc
                job_name = f"batch_{filename}_{i+1}_{args_list[0]}" # Include command in name
                trigger = DateTrigger(run_date=run_date_utc)
                try:
                    job = scheduler.add_job( run_labbuild_task, trigger=trigger, args=[args_list], name=job_name, misfire_grace_time=3600, replace_existing=False )
                    run_date_local = run_date_utc.astimezone(scheduler_tz) # Format for logging
                    log_time_str = f"{run_date_local.strftime('%Y-%m-%d %H:%M:%S %Z%z')} / UTC: {run_date_utc.strftime('%Y-%m-%d %H:%M:%S %Z')}"
                    logger.info(f"Scheduled batch job {i+1}: ID={job.id}, Name='{job_name}', RunAt={log_time_str}")
                    scheduled_count += 1; current_run_time_utc += job_delay
                except Exception as e_sched: logger.error(f"Fail schedule line {i+1}: {e_sched}", exc_info=True); failed_lines += 1
            elif line.strip() and not line.strip().startswith('#'): logger.warning(f"Skip invalid line {i+1}: {line.strip()}"); failed_lines += 1

        flash(f"Scheduled {scheduled_count} jobs from '{filename}'. {failed_lines} lines failed/skipped.", 'success' if scheduled_count > 0 else 'warning')
    except Exception as e: logger.error(f"Error processing batch file '{filename}': {e}", exc_info=True); flash(f"Error processing batch file: {e}", 'danger')

    return redirect(url_for('main.index'))


@bp.route('/jobs/delete/<job_id>', methods=['POST'])
def delete_job(job_id):
    """Delete a single scheduled job."""
    # Logic moved from original app.py
    if not scheduler or not scheduler.running: flash("Scheduler not running.", "danger"); return redirect(url_for('main.index'))
    try: scheduler.remove_job(job_id); flash(f"Job {job_id} deleted.", 'success')
    except Exception as e: logger.error(f"Failed delete job {job_id}: {e}", exc_info=True); flash(f"Error deleting job: {e}", 'danger')
    return redirect(url_for('main.index'))

@bp.route('/jobs/delete-bulk', methods=['POST'])
def delete_bulk_jobs():
    """Deletes multiple scheduled jobs based on selected IDs."""
    # Logic moved from original app.py
    if not scheduler or not scheduler.running: flash("Scheduler not running.", "danger"); return redirect(url_for('main.index'))
    job_ids_to_delete = request.form.getlist('job_ids')
    if not job_ids_to_delete: flash("No jobs selected for deletion.", "warning"); return redirect(url_for('main.index'))
    deleted_count, failed_count = 0, 0
    for job_id in job_ids_to_delete:
        try: scheduler.remove_job(job_id); deleted_count += 1
        except Exception as e: logger.error(f"Failed delete job {job_id} bulk: {e}", exc_info=True); failed_count += 1
    flash(f"Deleted {deleted_count} job(s). Failed: {failed_count} job(s).", 'success' if failed_count == 0 else 'warning')
    return redirect(url_for('main.index'))


@bp.route('/toggle-power', methods=['POST'])
def toggle_power():
    """Handles power toggle actions from the allocations page."""
    # Logic moved from original app.py
    scope = request.form.get('scope'); tag = request.form.get('tag')
    item_type = request.form.get('item_type'); item_number_str = request.form.get('item_number')
    host = request.form.get('host'); vendor = request.form.get('vendor'); course = request.form.get('course')
    item_class_number_str = request.form.get('item_class_number')

    if not scope or not tag: flash("Missing scope or tag for power toggle.", "danger"); return redirect(url_for('main.view_allocations', **request.args))
    logger.info(f"Power toggle request: Scope='{scope}', Tag='{tag}'")
    tasks_to_run = [] # List of tuples: (args_list, description, db_update_info)

    try:
        if db is None or alloc_collection is None: raise ConnectionError("Database unavailable.")
        if scope == 'tag':
            tag_doc = alloc_collection.find_one({"tag": tag});
            if not tag_doc: raise ValueError(f"Tag '{tag}' not found.")
            courses_in_tag = tag_doc.get("courses", [])
            if not isinstance(courses_in_tag, list): courses_in_tag = []
            for course_item in courses_in_tag:
                if not isinstance(course_item, dict): continue
                c_name, c_vendor, pod_details = course_item.get("course_name"), course_item.get("vendor"), course_item.get("pod_details", [])
                if not c_name or not c_vendor or not isinstance(pod_details, list): continue
                for pd in pod_details:
                    if not isinstance(pd, dict): continue
                    pd_host, pd_pod, pd_class = pd.get("host", pd.get("pod_host")), pd.get("pod_number"), pd.get("class_number")
                    current_power_on = str(pd.get("poweron", False)).lower() == 'true'
                    action = 'stop' if current_power_on else 'start'; new_power_state_bool = not current_power_on
                    if not pd_host: continue
                    is_f5 = c_vendor.lower() == 'f5'; item_detail = None
                    if is_f5 and pd_class is not None:
                        args = ['manage', '-v', c_vendor, '-g', c_name, '--host', pd_host, '-t', tag, '-o', action, '-cn', str(pd_class)]
                        desc = f"Tag '{tag}' - Manage {action} Class {pd_class}"; item_detail = (c_name, pd_host, pd_class, "f5_class", new_power_state_bool)
                    elif pd_pod is not None:
                        args = ['manage', '-v', c_vendor, '-g', c_name, '--host', pd_host, '-t', tag, '-o', action, '-s', str(pd_pod), '-e', str(pd_pod)]
                        if is_f5 and pd_class is not None: args.extend(['-cn', str(pd_class)])
                        desc = f"Tag '{tag}' - Manage {action} Pod {pd_pod}"; item_detail = (c_name, pd_host, pd_pod, "pod", new_power_state_bool)
                    if item_detail: tasks_to_run.append((args, desc, item_detail))
        elif scope == 'item':
            if not all([item_type, item_number_str, host, vendor, course]): raise ValueError("Missing item details.")
            item_number = int(item_number_str)
            current_state_is_on = request.form.get('current_state') == 'on'
            action = 'stop' if current_state_is_on else 'start'; new_power_state_bool = not current_state_is_on
            args = ['manage', '-v', vendor, '-g', course, '--host', host, '-t', tag, '-o', action]
            item_detail = None; item_class_num = None
            if item_type == 'f5_class':
                args.extend(['-cn', str(item_number)]); desc = f"Item - Manage {action} Class {item_number}"
                item_detail = (course, host, item_number, "f5_class", new_power_state_bool)
            elif item_type == 'pod':
                args.extend(['-s', str(item_number), '-e', str(item_number)])
                if vendor.lower() == 'f5' and item_class_number_str: item_class_num = int(item_class_number_str); args.extend(['-cn', item_class_number_str])
                desc = f"Item - Manage {action} Pod {item_number}"; item_detail = (course, host, item_number, "pod", new_power_state_bool)
                if item_class_num is not None: item_detail = (course, host, item_number, "pod", new_power_state_bool, item_class_num)
            else: raise ValueError(f"Invalid item_type: {item_type}")
            if item_detail: tasks_to_run.append((args, desc, item_detail))
        else: raise ValueError(f"Invalid scope: {scope}")

        # --- Execute Tasks and Update DB ---
        if not tasks_to_run: flash(f"No valid manage tasks for Scope '{scope}', Tag '{tag}'.", "warning")
        else:
            logger.info(f"Submitting {len(tasks_to_run)} tasks for Scope '{scope}', Tag '{tag}'...")
            def run_sequential_manage_and_update(tasks):
                for i, (args, desc, db_info) in enumerate(tasks):
                    logger.info(f"Starting task {i+1}/{len(tasks)}: {desc}")
                    run_labbuild_task(args); logger.info(f"Finished task {i+1}/{len(tasks)}: {desc}")
                    if db_info:
                        try: update_power_state_in_db(tag, *db_info)
                        except Exception as db_err: logger.error(f"Failed DB update for {desc}: {db_err}")
                logger.info(f"All tasks submitted for Scope '{scope}', Tag '{tag}'.")
            thread = threading.Thread(target=run_sequential_manage_and_update, args=(tasks_to_run,), daemon=True)
            thread.start()
            flash(f"Submitted {len(tasks_to_run)} tasks for Scope '{scope}', Tag '{tag}'.", "info")

    except (ConnectionError, ValueError) as e: logger.error(f"Error during toggle for {tag}: {e}"); flash(f"Error: {e}", "danger")
    except Exception as e: logger.error(f"Unexpected error toggle for {tag}: {e}", exc_info=True); flash(f"Error submitting toggle: {e}", 'danger')

    query_params = {k: v for k, v in request.form.items() if k.startswith('filter_')}
    return redirect(url_for('main.view_allocations', **query_params)) # Redirect to main blueprint


@bp.route('/teardown-item', methods=['POST'])
def teardown_item():
    """Handles teardown/delete actions from the allocations page."""
    # Logic moved from original app.py
    try:
        tag = request.form.get('tag'); host = request.form.get('host'); vendor = request.form.get('vendor'); course = request.form.get('course')
        pod_num_str = request.form.get('pod_number'); class_num_str = request.form.get('class_number'); delete_level = request.form.get('delete_level')
        if not delete_level or not tag: flash("Missing delete level or tag.", "danger"); return redirect(url_for('main.view_allocations'))
        pod_num = int(pod_num_str) if pod_num_str else None; class_num = int(class_num_str) if class_num_str else None

        # --- Handle DB Deletion Levels ---
        if delete_level.endswith('_db'):
            logger.info(f"DB delete request: Level='{delete_level}', Tag='{tag}'")
            success, item_desc = False, "Unknown"
            try:
                if delete_level == 'tag_db': success = delete_from_database(tag=tag); item_desc = f"Tag '{tag}'"
                elif delete_level == 'course_db': success = delete_from_database(tag=tag, course_name=course); item_desc = f"Course '{course}' in Tag '{tag}'"
                elif delete_level == 'class_db': success = delete_from_database(tag=tag, course_name=course, class_number=class_num); item_desc = f"Class {class_num} in '{course}'/'{tag}'"
                elif delete_level == 'pod_db': success = delete_from_database(tag=tag, course_name=course, pod_number=pod_num, class_number=class_num); item_desc = f"Pod {pod_num}" + (f" (Class {class_num})" if class_num else "") + f" in '{course}'/'{tag}'"
                else: flash("Invalid DB delete level.", "warning"); return redirect(url_for('main.view_allocations'))
                flash(f"Removed DB entry for {item_desc}." if success else f"Failed DB removal for {item_desc}.", 'success' if success else 'danger')
            except Exception as e_db: flash(f"Error during DB delete: {e_db}", 'danger'); logger.error(f"DB delete error: {e_db}", exc_info=True)
            query_params = {k: v for k, v in request.form.items() if k.startswith('filter_')}
            return redirect(url_for('main.view_allocations', **query_params))
        # --- Handle Full Teardown ---
        elif delete_level in ['class', 'pod']:
            if not all([vendor, course, host]): flash("Vendor, Course, Host required for infrastructure teardown.", "danger"); return redirect(url_for('main.view_allocations'))
            args_list, item_desc = [], ""
            if delete_level == 'class' and vendor.lower() == 'f5' and class_num is not None: args_list = ['teardown', '-v', vendor, '-g', course, '--host', host, '-t', tag, '-cn', str(class_num)]; item_desc = f"F5 Class {class_num}"
            elif delete_level == 'pod' and pod_num is not None:
                args_list = ['teardown', '-v', vendor, '-g', course, '--host', host, '-t', tag, '-s', str(pod_num), '-e', str(pod_num)]
                item_desc = f"Pod {pod_num}"
                if vendor.lower() == 'f5' and class_num is not None: args_list.extend(['-cn', str(class_num)]); item_desc += f" (Class {class_num})"
            else: flash("Invalid teardown level or missing identifiers.", "danger"); return redirect(url_for('main.view_allocations'))
            if args_list:
                thread = threading.Thread(target=run_labbuild_task, args=(args_list,), daemon=True); thread.start()
                flash(f"Submitted infrastructure teardown for {item_desc}.", 'info')
            else: flash("Failed build teardown command.", "danger")
        else: flash(f"Unsupported teardown level: '{delete_level}'.", "warning")
    except Exception as e: logger.error(f"Error processing teardown/delete request: {e}", exc_info=True); flash(f"Error processing request: {e}", 'danger')
    query_params = {k: v for k, v in request.form.items() if k.startswith('filter_')}
    return redirect(url_for('main.view_allocations', **query_params))


@bp.route('/teardown-tag', methods=['POST'])
def teardown_tag():
    """Handles teardown of an entire tag from the allocations page."""
    # Logic moved from original app.py
    from collections import defaultdict # Needed here
    tag = request.form.get('tag')
    if not tag: flash("Tag identifier is missing.", "danger"); return redirect(url_for('main.view_allocations'))
    logger.info(f"Initiating FULL teardown for Tag: '{tag}'")
    tasks_to_run = [] # List of ([args_list], description) tuples

    try:
        if db is None or alloc_collection is None: raise Exception("DB unavailable.")
        tag_doc = alloc_collection.find_one({"tag": tag})
        if not tag_doc: flash(f"Tag '{tag}' not found.", "warning"); return redirect(url_for('main.view_allocations'))

        courses_in_tag = tag_doc.get("courses", [])
        if not isinstance(courses_in_tag, list): courses_in_tag = []

        for course in courses_in_tag:
            if not isinstance(course, dict): continue
            course_name, vendor, pod_details = course.get("course_name"), course.get("vendor"), course.get("pod_details", [])
            if not course_name or not vendor or not isinstance(pod_details, list): continue
            items_by_host = defaultdict(lambda: {"pods": set(), "classes": set()})
            for pd in pod_details:
                if not isinstance(pd, dict): continue
                host, pod_num, class_num = pd.get("host", pd.get("pod_host")), pd.get("pod_number"), pd.get("class_number")
                if not host: continue
                is_f5 = vendor.lower() == 'f5'
                if is_f5 and class_num is not None: items_by_host[host]["classes"].add(class_num)
                elif pod_num is not None: items_by_host[host]["pods"].add(pod_num)

            for host, items in items_by_host.items():
                for class_num in sorted(list(items["classes"])): args = ['teardown', '-v', vendor, '-g', course_name, '--host', host, '-t', tag, '-cn', str(class_num)]; desc = f"Tag '{tag}', Course '{course_name}', Class '{class_num}'"; tasks_to_run.append((args, desc))
                if items["pods"]:
                    for pod_num in sorted(list(items["pods"])): args = ['teardown', '-v', vendor, '-g', course_name, '--host', host, '-t', tag, '-s', str(pod_num), '-e', str(pod_num)]; desc = f"Tag '{tag}', Course '{course_name}', Pod '{pod_num}'"; tasks_to_run.append((args, desc))

        if not tasks_to_run: flash(f"No teardown tasks found for Tag '{tag}'.", "warning")
        else:
            logger.info(f"Submitting {len(tasks_to_run)} tasks for Tag '{tag}' sequentially...")
            def run_sequential_tasks(tasks):
                for i, (args, desc) in enumerate(tasks): logger.info(f"Starting task {i+1}/{len(tasks)}: {desc}"); run_labbuild_task(args); logger.info(f"Finished task {i+1}/{len(tasks)}: {desc}")
                logger.info(f"All teardown tasks submitted for Tag '{tag}'.")
            thread = threading.Thread(target=run_sequential_tasks, args=(tasks_to_run,), daemon=True); thread.start()
            flash(f"Submitted {len(tasks_to_run)} teardown tasks for Tag '{tag}'.", "info")
    except Exception as e: logger.error(f"Error processing teardown Tag '{tag}': {e}", exc_info=True); flash(f"Error submitting teardown Tag '{tag}': {e}", 'danger')

    query_params = {k: v for k, v in request.form.items() if k.startswith('filter_')}
    return redirect(url_for('main.view_allocations', **query_params))


@bp.route('/build-row', methods=['POST'])
def build_row():
    """Handles 'Build' action from the upcoming courses page."""
    # Logic moved from original app.py
    try:
        data = request.json
        if not data: return jsonify({"status": "error", "message": "No data received."}), 400
        labbuild_course, start_pod, end_pod, host, vendor = data.get('labbuild_course'), data.get('start_pod'), data.get('end_pod'), data.get('host'), data.get('vendor')
        sf_course_code = data.get('sf_course_code')
        if not all([labbuild_course, start_pod, end_pod, host, vendor]): return jsonify({"status": "error", "message": "Missing required fields."}), 400
        try: s_pod, e_pod = int(start_pod), int(end_pod); assert s_pod >= 0 and e_pod >= s_pod
        except (ValueError, AssertionError): return jsonify({"status": "error", "message": "Invalid Pod numbers."}), 400
        tag = f"uc_{sf_course_code}"[:50] if sf_course_code else "uc_dashboard"
        args_list = [ 'setup', '-v', vendor, '-g', labbuild_course, '--host', host, '-s', str(s_pod), '-e', str(e_pod), '-t', tag ]
        thread = threading.Thread(target=run_labbuild_task, args=(args_list,), daemon=True); thread.start()
        # Use flash and redirect for UI feedback, or stick to JSON for pure AJAX
        flash(f"Submitted build for {labbuild_course} (Pods {s_pod}-{e_pod}) on {host}.", "info") # Example Flash
        return jsonify({"status": "success", "message": "Build submitted."}), 200
    except Exception as e: logger.error(f"Error in /build-row: {e}", exc_info=True); return jsonify({"status": "error", "message": "Internal server error."}), 500

def _find_all_matching_rules(rules: list, vendor: str, code: str, type_str: str) -> list[dict]:
    """
    Finds all build rules matching the course criteria.
    Assumes input 'rules' list is already sorted by priority (ascending).
    """
    matching_rules_list = []
    # Normalize inputs for case-insensitive matching
    cc_lower = code.lower() if code else ""
    ct_lower = type_str.lower() if type_str else ""
    v_lower = vendor.lower() if vendor else ""

    for rule in rules: 
        conditions = rule.get("conditions", {})
        
        match = True # Assume match until a condition fails

        # 1. Vendor check (must match if specified in rule)
        rule_vendor_cond = conditions.get("vendor")
        if rule_vendor_cond and rule_vendor_cond.lower() != v_lower:
            match = False
        
        # 2. Course Code Contains check (only if match is still true)
        if match and "course_code_contains" in conditions:
            terms = conditions["course_code_contains"]
            terms = [terms] if not isinstance(terms, list) else terms # Ensure list
            if not any(str(t).lower() in cc_lower for t in terms):
                match = False
        
        # 3. Course Code NOT Contains check (only if match is still true)
        if match and "course_code_not_contains" in conditions:
            terms = conditions["course_code_not_contains"]
            terms = [terms] if not isinstance(terms, list) else terms # Ensure list
            if any(str(t).lower() in cc_lower for t in terms):
                match = False

        # 4. Course Type Contains check (only if match is still true)
        if match and "course_type_contains" in conditions:
            terms = conditions["course_type_contains"]
            terms = [terms] if not isinstance(terms, list) else terms # Ensure list
            if not any(str(t).lower() in ct_lower for t in terms):
                match = False
        
        if match: # If all specific conditions passed (or no specific conditions to fail), add the rule
            matching_rules_list.append(rule)
            
    return matching_rules_list


@bp.route('/build-rules/add', methods=['POST'])
def add_build_rule():
    """Adds a new build rule to the collection."""
    if build_rules_collection is None:
        flash("Build rules database collection is unavailable.", "danger")
        return redirect(url_for('settings.view_build_rules')) # Redirect to settings blueprint to view rules

    try:
        # Extract data from the form
        rule_name = request.form.get('rule_name', '').strip()
        priority_str = request.form.get('priority', '').strip()
        conditions_json = request.form.get('conditions', '{}').strip() # Default to empty JSON string
        actions_json = request.form.get('actions', '{}').strip()   # Default to empty JSON string

        # --- Basic Validation ---
        if not rule_name:
            flash("Rule Name is required.", "danger")
            return redirect(url_for('settings.view_build_rules'))
        
        try:
            priority = int(priority_str)
            if priority < 1: # Priority should be positive
                raise ValueError("Priority must be a positive integer.")
        except (ValueError, TypeError):
             flash("Priority must be a positive integer.", "danger")
             return redirect(url_for('settings.view_build_rules'))
        
        try:
            conditions = json.loads(conditions_json)
            if not isinstance(conditions, dict):
                raise ValueError("Conditions field must contain a valid JSON object.")
        except (json.JSONDecodeError, ValueError) as e:
             flash(f"Conditions field contains invalid JSON: {e}", "danger")
             return redirect(url_for('settings.view_build_rules'))
        
        try:
            actions = json.loads(actions_json)
            if not isinstance(actions, dict):
                raise ValueError("Actions field must contain a valid JSON object.")
        except (json.JSONDecodeError, ValueError) as e:
             flash(f"Actions field contains invalid JSON: {e}", "danger")
             return redirect(url_for('settings.view_build_rules'))

        # Prepare the document to be inserted
        new_rule = {
            "rule_name": rule_name,
            "priority": priority,
            "conditions": conditions,
            "actions": actions,
            "created_at": datetime.datetime.now(pytz.utc) # Add a creation timestamp in UTC
            # "updated_at": datetime.datetime.now(pytz.utc) # Optionally add updated_at too
        }

        # Insert into the database
        result = build_rules_collection.insert_one(new_rule)
        
        # Log success and flash message
        logger.info(f"Added new build rule '{rule_name}' (Priority: {priority}) with ID: {result.inserted_id}")
        flash(f"Successfully added build rule '{rule_name}'.", "success")

    except PyMongoError as e:
        logger.error(f"Database error occurred while adding build rule: {e}")
        flash("A database error occurred while adding the rule. Please try again.", "danger")
    except Exception as e:
        logger.error(f"An unexpected error occurred while adding build rule: {e}", exc_info=True)
        flash("An unexpected error occurred. Please check the logs.", "danger")

    return redirect(url_for('settings.view_build_rules')) # Redirect back to the rules

@bp.route('/build-rules/update', methods=['POST'])
def update_build_rule():
    """Updates an existing build rule."""
    if build_rules_collection is None:
        flash("Build rules database collection is unavailable.", "danger")
        return redirect(url_for('settings.view_build_rules'))

    rule_id_str = request.form.get('rule_id')
    if not rule_id_str:
        flash("Rule ID missing for update.", "danger")
        return redirect(url_for('settings.view_build_rules'))

    try:
        rule_oid = ObjectId(rule_id_str) # Convert string ID to ObjectId
    except InvalidId:
        flash("Invalid Rule ID format.", "danger")
        return redirect(url_for('settings.view_build_rules'))

    try:
        rule_name = request.form.get('rule_name', '').strip()
        priority_str = request.form.get('priority', '').strip()
        conditions_json = request.form.get('conditions', '{}').strip()
        actions_json = request.form.get('actions', '{}').strip()

        # Validation (similar to add)
        if not rule_name: raise ValueError("Rule Name is required.")
        try: priority = int(priority_str); assert priority >= 1
        except: raise ValueError("Priority must be a positive integer.")
        try: conditions = json.loads(conditions_json); assert isinstance(conditions, dict)
        except: raise ValueError("Conditions field contains invalid JSON.")
        try: actions = json.loads(actions_json); assert isinstance(actions, dict)
        except: raise ValueError("Actions field contains invalid JSON.")

        # Prepare update document
        update_doc = {
            "$set": {
                "rule_name": rule_name,
                "priority": priority,
                "conditions": conditions,
                "actions": actions,
                "updated_at": datetime.datetime.now(pytz.utc) # Add update timestamp
            }
        }

        # Perform update
        result = build_rules_collection.update_one({"_id": rule_oid}, update_doc)

        if result.matched_count == 0:
            flash(f"Rule with ID {rule_id_str} not found.", "warning")
        elif result.modified_count == 0:
             flash(f"Rule '{rule_name}' was not modified (no changes detected).", "info")
        else:
            logger.info(f"Updated build rule '{rule_name}' (ID: {rule_id_str})")
            flash(f"Successfully updated build rule '{rule_name}'.", "success")

    except ValueError as ve: # Catch validation errors
        flash(f"Invalid input: {ve}", "danger")
    except PyMongoError as e:
        logger.error(f"Database error updating build rule {rule_id_str}: {e}")
        flash("Database error updating rule.", "danger")
    except Exception as e:
        logger.error(f"Unexpected error updating build rule {rule_id_str}: {e}", exc_info=True)
        flash("An unexpected error occurred while updating.", "danger")

    return redirect(url_for('settings.view_build_rules'))


@bp.route('/build-rules/delete/<rule_id>', methods=['POST'])
def delete_build_rule(rule_id):
    """Deletes a build rule by its ID."""
    if build_rules_collection is None:
        flash("Build rules database collection is unavailable.", "danger")
        return redirect(url_for('settings.view_build_rules'))

    try:
        rule_oid = ObjectId(rule_id) # Convert string ID from URL
    except InvalidId:
        flash("Invalid Rule ID format for deletion.", "danger")
        return redirect(url_for('settings.view_build_rules'))

    try:
        result = build_rules_collection.delete_one({"_id": rule_oid})

        if result.deleted_count == 1:
            logger.info(f"Deleted build rule with ID: {rule_id}")
            flash("Successfully deleted build rule.", "success")
        else:
            flash(f"Rule with ID {rule_id} not found for deletion.", "warning")

    except PyMongoError as e:
        logger.error(f"Database error deleting build rule {rule_id}: {e}")
        flash("Database error deleting rule.", "danger")
    except Exception as e:
        logger.error(f"Unexpected error deleting build rule {rule_id}: {e}", exc_info=True)
        flash("An unexpected error occurred during deletion.", "danger")

    return redirect(url_for('settings.view_build_rules'))

@bp.route('/course-configs/add', methods=['POST'])
def add_course_config():
    """Adds a new course configuration."""
    if course_config_collection is None:
        flash("Course config collection unavailable.", "danger")
        return redirect(url_for('settings.view_course_configs')) # Redirect to settings blueprint

    config_json_str = request.form.get('config_json', '{}').strip()
    try:
        new_config_data = json.loads(config_json_str)
        if not isinstance(new_config_data, dict):
            raise ValueError("Configuration must be a valid JSON object.")

        # Validate required fields
        course_name = new_config_data.get('course_name')
        vendor_shortcode = new_config_data.get('vendor_shortcode')
        if not course_name or not isinstance(course_name, str) or not course_name.strip():
            flash("Valid 'course_name' (string) is required in the JSON.", "danger")
            return redirect(url_for('settings.view_course_configs'))
        if not vendor_shortcode or not isinstance(vendor_shortcode, str) or not vendor_shortcode.strip():
            flash("Valid 'vendor_shortcode' (string) is required in the JSON.", "danger")
            return redirect(url_for('settings.view_course_configs'))

        new_config_data['course_name'] = course_name.strip() # Ensure trimmed
        new_config_data['vendor_shortcode'] = vendor_shortcode.strip().lower() # Trim & lowercase vendor

        # Add created_at timestamp
        new_config_data['created_at'] = datetime.datetime.now(pytz.utc)

        # Attempt to insert
        # Consider adding a unique index on (course_name, vendor_shortcode) in MongoDB
        # to prevent exact duplicates if course_name itself isn't globally unique.
        # For now, let's assume course_name should be unique.
        if course_config_collection.count_documents({"course_name": new_config_data['course_name']}) > 0:
            flash(f"Course configuration with name '{new_config_data['course_name']}' already exists.", "warning")
            return redirect(url_for('settings.view_course_configs'))

        result = course_config_collection.insert_one(new_config_data)
        logger.info(f"Added new course config '{new_config_data['course_name']}' with ID: {result.inserted_id}")
        flash(f"Successfully added course configuration '{new_config_data['course_name']}'.", "success")

    except json.JSONDecodeError:
        flash("Invalid JSON format for configuration.", "danger")
    except ValueError as ve: # Catch our custom validation errors
        flash(str(ve), "danger")
    except PyMongoError as e:
        logger.error(f"Database error adding course config: {e}")
        if e.code == 11000: # Duplicate key error
             flash(f"A course configuration with that name or key combination already exists.", "danger")
        else:
             flash("Database error adding configuration.", "danger")
    except Exception as e:
        logger.error(f"Unexpected error adding course config: {e}", exc_info=True)
        flash("An unexpected error occurred.", "danger")

    return redirect(url_for('settings.view_course_configs'))


@bp.route('/course-configs/update', methods=['POST'])
def update_course_config():
    """Updates an existing course configuration."""
    if course_config_collection is None:
        flash("Course config collection unavailable.", "danger")
        return redirect(url_for('settings.view_course_configs'))

    config_id_str = request.form.get('config_id')
    config_json_str = request.form.get('config_json', '{}').strip()

    if not config_id_str:
        flash("Configuration ID missing for update.", "danger")
        return redirect(url_for('settings.view_course_configs'))
    try:
        config_oid = ObjectId(config_id_str)
    except InvalidId:
        flash("Invalid Configuration ID format.", "danger")
        return redirect(url_for('settings.view_course_configs'))

    try:
        updated_config_data = json.loads(config_json_str)
        if not isinstance(updated_config_data, dict):
            raise ValueError("Configuration must be a valid JSON object.")

        # Validate required fields are still present and valid
        course_name = updated_config_data.get('course_name')
        vendor_shortcode = updated_config_data.get('vendor_shortcode')
        if not course_name or not isinstance(course_name, str) or not course_name.strip():
            raise ValueError("Valid 'course_name' (string) is required.")
        if not vendor_shortcode or not isinstance(vendor_shortcode, str) or not vendor_shortcode.strip():
            raise ValueError("Valid 'vendor_shortcode' (string) is required.")

        updated_config_data['course_name'] = course_name.strip()
        updated_config_data['vendor_shortcode'] = vendor_shortcode.strip().lower()

        # Add updated_at timestamp
        updated_config_data['updated_at'] = datetime.datetime.now(pytz.utc)

        # Remove _id from update data if it was accidentally included from textarea
        updated_config_data.pop('_id', None)

        # Check for name collision if course_name is being changed to an existing one
        # (and it's not the current document being edited)
        existing_with_new_name = course_config_collection.find_one({
            "course_name": updated_config_data['course_name'],
            "_id": {"$ne": config_oid}
        })
        if existing_with_new_name:
             flash(f"Another course configuration with the name '{updated_config_data['course_name']}' already exists.", "danger")
             return redirect(url_for('settings.view_course_configs'))


        result = course_config_collection.update_one(
            {"_id": config_oid},
            {"$set": updated_config_data}
        )

        if result.matched_count == 0:
            flash(f"Course configuration with ID {config_id_str} not found.", "warning")
        elif result.modified_count == 0:
             flash(f"Course configuration '{updated_config_data['course_name']}' was not modified (no changes detected or attempt to change to existing name).", "info")
        else:
            logger.info(f"Updated course config '{updated_config_data['course_name']}' (ID: {config_id_str})")
            flash(f"Successfully updated course configuration '{updated_config_data['course_name']}'.", "success")

    except json.JSONDecodeError:
        flash("Invalid JSON format for configuration.", "danger")
    except ValueError as ve:
        flash(str(ve), "danger")
    except PyMongoError as e:
        logger.error(f"Database error updating course config {config_id_str}: {e}")
        if e.code == 11000: # Duplicate key error
             flash(f"Update failed: A course configuration with the new name or key combination may already exist.", "danger")
        else:
            flash("Database error updating configuration.", "danger")
    except Exception as e:
        logger.error(f"Unexpected error updating course config {config_id_str}: {e}", exc_info=True)
        flash("An unexpected error occurred.", "danger")

    return redirect(url_for('settings.view_course_configs'))


@bp.route('/course-configs/delete/<config_id>', methods=['POST'])
def delete_course_config(config_id):
    """Deletes a course configuration by its ID."""
    if course_config_collection is None:
        flash("Course config collection unavailable.", "danger")
        return redirect(url_for('settings.view_course_configs'))

    try:
        config_oid = ObjectId(config_id)
    except InvalidId:
        flash("Invalid Configuration ID format for deletion.", "danger")
        return redirect(url_for('settings.view_course_configs'))

    try:
        result = course_config_collection.delete_one({"_id": config_oid})
        if result.deleted_count == 1:
            logger.info(f"Deleted course configuration with ID: {config_id}")
            flash("Successfully deleted course configuration.", "success")
        else:
            flash(f"Course configuration with ID {config_id} not found for deletion.", "warning")
    except PyMongoError as e:
        logger.error(f"Database error deleting course config {config_id}: {e}")
        flash("Database error deleting configuration.", "danger")
    except Exception as e:
        logger.error(f"Unexpected error deleting course config {config_id}: {e}", exc_info=True)
        flash("An unexpected error occurred during deletion.", "danger")

    return redirect(url_for('settings.view_course_configs'))

@bp.route('/toggle-tag-extend', methods=['POST'])
def toggle_tag_extend():
    """
    Toggles the 'extend' field (string "true"/"false") for a given tag
    in the 'currentallocation' collection based on the current state provided
    in the form submission.

    Redirects back to the allocations page, preserving any active filters
    and pagination settings.
    """
    tag_name = request.form.get('tag_name')
    # Get the status *as sent from the form* (state *before* the click)
    # Default to 'false' string if the input is missing, for safety.
    current_status_str = request.form.get('current_extend_status', 'false')

    # --- Preserve Filters & Pagination for Redirect ---
    # Collect all relevant query parameters that might have been passed back
    # from the allocations page form submission.
    redirect_args = {}
    for key in ['filter_tag', 'filter_vendor', 'filter_course', 'filter_host', 'filter_number', 'page', 'per_page']:
        # Use request.form for POST data, fallback to request.args for GET params if needed
        # Form submission usually puts everything in request.form for POST
        value = request.form.get(key)
        if value is not None: # Only include params that are present
             # Clean up filter keys if they have the prefix from the form
             clean_key = key.replace('filter_', '')
             redirect_args[clean_key] = value

    logger.info(f"Received toggle request for tag: '{tag_name}', current extend status from form: '{current_status_str}'")
    logger.debug(f"Redirect args collected: {redirect_args}")

    # --- Input Validation ---
    if not tag_name:
        flash("Tag name missing for toggle operation.", "danger")
        return redirect(url_for('main.view_allocations', **redirect_args))

    # --- Database Check ---
    if alloc_collection is None:
        flash("Allocation database collection is unavailable.", "danger")
        logger.error("alloc_collection is None in toggle_tag_extend.")
        return redirect(url_for('main.view_allocations', **redirect_args))

    # --- Determine the NEW status (opposite of current) ---
    try:
        # Convert received string status to boolean for logical flipping
        current_status_bool = current_status_str.lower() == 'true'
        # The new state is the opposite boolean
        new_status_bool = not current_status_bool
        # Convert the NEW state back to the STRING format ("true" or "false")
        # This must match the data type expected by DB queries/other logic.
        new_status_str = "true" if new_status_bool else "false"
    except Exception as e:
        logger.error(f"Error determining new status for tag '{tag_name}': {e}")
        flash("Internal error processing status toggle.", "danger")
        return redirect(url_for('main.view_allocations', **redirect_args))

    logger.info(f"Attempting to update tag '{tag_name}' extend status to: '{new_status_str}'")

    # --- Perform Database Update ---
    try:
        result = alloc_collection.update_one(
            {"tag": tag_name}, # Filter: Find the document by tag name
            {"$set": {"extend": new_status_str}} # Update: Set the 'extend' field
        )

        # --- Check Update Result and Provide Feedback ---
        logger.debug(f"MongoDB update result for tag '{tag_name}': Matched={result.matched_count}, Modified={result.modified_count}")

        if result.matched_count == 0:
            flash(f"Tag '{tag_name}' not found in database.", "warning")
            logger.warning(f"Tag '{tag_name}' not found during extend toggle update.")
        elif result.modified_count == 0:
             # Occurs if the document was found but the value was already the target value.
             flash(f"Tag '{tag_name}' extend status was not changed (already set to '{new_status_str}'?).", "info")
             logger.warning(f"Tag '{tag_name}' matched but extend status not modified. Value might already be '{new_status_str}'.")
        else:
            # Success Case
            new_status_display = "NO REUSE (Locked)" if new_status_str == "true" else "ALLOW REUSE (Unlocked)"
            flash(f"Tag '{tag_name}' updated. Pod reuse policy set to: {new_status_display}.", "success")
            logger.info(f"Successfully toggled extend status for tag '{tag_name}' to '{new_status_str}'.")

    except PyMongoError as e:
        logger.error(f"Database error toggling extend status for tag '{tag_name}': {e}", exc_info=True)
        flash("Database error updating tag status.", "danger")
    except Exception as e:
        logger.error(f"Unexpected error toggling extend status for tag '{tag_name}': {e}", exc_info=True)
        flash("An unexpected error occurred.", "danger")

    # --- Redirect back to the allocations page, preserving filters & pagination ---
    # The **redirect_args unpacks the dictionary into keyword arguments for url_for
    return redirect(url_for('main.view_allocations', **redirect_args))

def _format_date_for_review(raw_date_str: Optional[str], context: str) -> str:
    # ... (Implementation from previous responses) ...
    if not raw_date_str or raw_date_str == "N/A": return "N/A"
    try: datetime.datetime.strptime(raw_date_str, "%Y-%m-%d"); return raw_date_str
    except ValueError:
        for fmt in ("%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%m-%d-%Y"):
            try: return datetime.datetime.strptime(raw_date_str, fmt).strftime("%Y-%m-%d")
            except ValueError: continue
        logger.warning(f"Could not parse date '{raw_date_str}' for {context}, keeping original.")
        return raw_date_str

@bp.route('/intermediate-build-review', methods=['POST'])
def intermediate_build_review():
    """
    Handles the intermediate build review process. Includes special logic for "Maestro"
    split-build courses, creating separate build items for each component type.
    """
    current_theme = request.cookies.get('theme', 'light')
    selected_courses_json = request.form.get('selected_courses')

    # --- 1. NESTED HELPER FUNCTIONS (UNCHANGED) ---

    def _fetch_and_prepare_initial_data() -> Optional[Dict[str, Any]]:
        # This helper function remains the same as the previous version.
        initial_data = {
            "build_rules": [], "all_available_host_names": [], "host_memory_start_of_batch": {},
            "db_locked_pods": defaultdict(set),
            "course_configs": [], "available_lab_courses": defaultdict(set),
            "f5_highest_class_number_used": 0
        }
        try:
            initial_data["build_rules"] = list(build_rules_collection.find().sort("priority", pymongo.ASCENDING))
            hosts_docs = list(host_collection.find({"include_for_build": "true"}))
            initial_data["all_available_host_names"] = sorted([h['host_name'] for h in hosts_docs if 'host_name' in h])
            initial_data["course_configs"] = list(course_config_collection.find({}, {"course_name": 1, "vendor_shortcode": 1, "memory_gb_per_pod": 1, "memory": 1, "_id": 0}))
            for cfg in initial_data["course_configs"]: initial_data["available_lab_courses"][cfg.get('vendor_shortcode', '').lower()].add(cfg.get('course_name'))
            db_allocs_cursor = alloc_collection.find({"extend": "true"})
            for tag_doc in db_allocs_cursor:
                for course in tag_doc.get("courses", []):
                    vendor = course.get("vendor", "").lower()
                    if not vendor: continue
                    for pd in course.get("pod_details", []):
                        try:
                            if vendor == 'f5':
                                if pd.get("class_number") is not None:
                                    class_num_int = int(pd["class_number"])
                                    initial_data["db_locked_pods"]['f5'].add(class_num_int)
                                    initial_data["f5_highest_class_number_used"] = max(initial_data["f5_highest_class_number_used"], class_num_int)
                                for nested in pd.get("pods", []):
                                    if nested.get("pod_number") is not None: initial_data["db_locked_pods"]['f5'].add(int(nested["pod_number"]))
                            elif pd.get("pod_number") is not None: initial_data["db_locked_pods"][vendor].add(int(pd["pod_number"]))
                        except (ValueError, TypeError): continue
            initial_data["db_locked_pods"]['f5'].add(17)
            logger.info("Reserved pod number 17 for vendor 'f5'. It will not be allocated.")
            raw_caps_gb = get_hosts_available_memory_parallel(hosts_docs) if hosts_docs else {}
            for h_doc in hosts_docs:
                h_name = h_doc.get("host_name")
                vc_cap = raw_caps_gb.get(h_name)
                if h_name and vc_cap is not None: initial_data["host_memory_start_of_batch"][h_name] = round(vc_cap, 2)
            return initial_data
        except Exception as e:
            logger.error(f"Error during initial data fetch: {e}", exc_info=True)
            flash("Error fetching configuration data. Cannot proceed.", "danger")
            return None

    def _propose_assignments(
        num_pods_to_allocate: int,
        hosts_to_try: List[str],
        memory_per_pod: float,
        vendor_code: str,
        start_pod_suggestion: int,
        pods_assigned_in_this_batch: Dict[str, set],
        db_locked_pods: Dict[str, set],
        current_host_capacities_gb: Dict[str, float]
    ) -> Tuple[List[Dict[str, Any]], Optional[str]]:
        # This helper function remains the same as the previous version.
        assignments: List[Dict[str, Any]] = []
        warning_message: Optional[str] = None
        if num_pods_to_allocate <= 0: return [], None
        if memory_per_pod <= 0: return [], "Memory for LabBuild course is 0, cannot allocate."
        unavailable_pods = db_locked_pods.get(vendor_code, set()).union(pods_assigned_in_this_batch.get(vendor_code, set()))
        best_overallocated_contiguous_option: Optional[Tuple[str, int]] = None
        for host_target in hosts_to_try:
            candidate_start_pod = start_pod_suggestion
            attempts = 0
            while attempts < 1000:
                attempts += 1
                potential_range = range(candidate_start_pod, candidate_start_pod + num_pods_to_allocate)
                first_conflict = next((p for p in potential_range if p in unavailable_pods), -1)
                if first_conflict == -1:
                    start_pod, end_pod = potential_range.start, potential_range.stop - 1
                    memory_needed = memory_per_pod + (num_pods_to_allocate - 1) * memory_per_pod * SUBSEQUENT_POD_MEMORY_FACTOR
                    host_capacity = current_host_capacities_gb.get(host_target, 0.0)
                    if host_capacity >= memory_needed:
                        assignments.append({"host": host_target, "start_pod": start_pod, "end_pod": end_pod})
                        for i in range(start_pod, end_pod + 1): pods_assigned_in_this_batch[vendor_code].add(i)
                        current_host_capacities_gb[host_target] -= memory_needed
                        return assignments, None
                    elif best_overallocated_contiguous_option is None:
                        best_overallocated_contiguous_option = (host_target, start_pod)
                    break
                else:
                    candidate_start_pod = first_conflict + 1
        if best_overallocated_contiguous_option:
            host_target, start_pod = best_overallocated_contiguous_option
            end_pod = start_pod + num_pods_to_allocate - 1
            assignments.append({"host": host_target, "start_pod": start_pod, "end_pod": end_pod})
            memory_needed = memory_per_pod + (num_pods_to_allocate - 1) * memory_per_pod * SUBSEQUENT_POD_MEMORY_FACTOR
            warning_message = f"Memory Warning: Host '{host_target}' may be over-allocated."
            for i in range(start_pod, end_pod + 1): pods_assigned_in_this_batch[vendor_code].add(i)
            current_host_capacities_gb[host_target] -= memory_needed
            return assignments, warning_message
        pods_left_to_assign = num_pods_to_allocate
        for host_target in hosts_to_try:
            if pods_left_to_assign <= 0: break
            while pods_left_to_assign > 0:
                search_start_pod = start_pod_suggestion
                while search_start_pod in unavailable_pods: search_start_pod += 1
                potential_block_size = 0
                for i in range(pods_left_to_assign):
                    if (search_start_pod + i) in unavailable_pods: break
                    potential_block_size += 1
                if potential_block_size > 0:
                    start_pod, end_pod = search_start_pod, search_start_pod + potential_block_size - 1
                    assignments.append({"host": host_target, "start_pod": start_pod, "end_pod": end_pod})
                    mem_consumed = memory_per_pod + (potential_block_size - 1) * memory_per_pod * SUBSEQUENT_POD_MEMORY_FACTOR
                    if current_host_capacities_gb.get(host_target, 0.0) < mem_consumed:
                        warn = f"Memory Warning: Host '{host_target}' may be over-allocated by block {start_pod}-{end_pod}."
                        warning_message = (warning_message + " " if warning_message else "") + warn
                    current_host_capacities_gb[host_target] -= mem_consumed
                    pods_left_to_assign -= potential_block_size
                    for i in range(start_pod, end_pod + 1): unavailable_pods.add(i)
                else:
                    break
        if pods_left_to_assign > 0:
            warn_final = f"Could not allocate all pods; {pods_left_to_assign} remain unassigned."
            warning_message = (warning_message + " " if warning_message else "") + warn_final
        return assignments, warning_message.strip() if warning_message else None

    def _process_single_course(course_input, initial_data, batch_state, batch_review_id):
        """
        Processes a single Salesforce course. If it's a Maestro course, it "explodes" it
        into multiple build items. Otherwise, it processes it as a standard course.
        Returns a LIST of (interim_doc, display_doc) tuples.
        """
        sf_code, vendor, sf_course_type = course_input.get('sf_course_code', 'UNKNOWN'), course_input.get('vendor', '').lower(), course_input.get('sf_course_type', 'N/A')
        
        matched_rules = _find_all_matching_rules(initial_data["build_rules"], vendor, sf_code, sf_course_type)
        
        # --- NEW: Maestro Split Build Logic ---
        maestro_rule_config = None
        rule_based_actions = {}
        for rule in matched_rules:
            # Capture the first maestro_split_build action found (due to priority sort)
            if "maestro_split_build" in rule.get("actions", {}) and not maestro_rule_config:
                maestro_rule_config = rule["actions"]["maestro_split_build"]
            # Capture all other highest-priority actions
            for key in ["set_labbuild_course", "host_priority", "allow_spillover", "start_pod_number", "set_max_pods", "override_pods_req"]:
                 if key not in rule_based_actions and key in rule.get("actions", {}):
                     rule_based_actions[key] = rule["actions"][key]

        if maestro_rule_config:
            logger.info(f"Applying Maestro Split Build logic for SF Course '{sf_code}'.")
            processed_items = []
            
            base_interim_doc = {
                "batch_review_id": batch_review_id, "created_at": datetime.datetime.now(pytz.utc), "status": "pending_student_review",
                "sf_course_code": sf_code, "sf_course_type": sf_course_type, "sf_trainer_name": course_input.get('sf_trainer_name', 'N/A'),
                "sf_start_date": _format_date_for_review(course_input.get('sf_start_date'), sf_code), "sf_end_date": _format_date_for_review(course_input.get('sf_end_date'), sf_code),
                "sf_pax_count": int(course_input.get('sf_pax_count', 0)), "vendor": vendor,
                "maestro_split_config_details_rule": maestro_rule_config # Store the rule for later steps
            }

            # Define the parts of the split build
            split_parts = [
                {"name": "Rack 1", "pods": 1, "course": maestro_rule_config.get("rack1_course"), "hosts": [maestro_rule_config.get("rack_host")]},
                {"name": "Rack 2", "pods": 1, "course": maestro_rule_config.get("rack2_course"), "hosts": [maestro_rule_config.get("rack_host")]},
                {"name": "Main Pods", "pods": 2, "course": maestro_rule_config.get("main_course"), "hosts": [h for h in rule_based_actions.get("host_priority", []) if h.lower() != 'hotshot']}
            ]

            for part in split_parts:
                part_doc = base_interim_doc.copy()
                part_doc["_id"] = ObjectId()
                part_doc["final_labbuild_course"] = part["course"]
                part_doc["effective_pods_req"] = part["pods"]
                # part_doc["sf_course_code"] = f"{sf_code}-{part['name'].replace(' ','')}" # Create a unique display code
                
                mem = _get_memory_for_course_local(part["course"], {c['course_name']: c for c in initial_data["course_configs"]})
                part_doc["memory_gb_one_pod"] = mem

                assignments, warn = _propose_assignments(
                    num_pods_to_allocate=part["pods"], hosts_to_try=part["hosts"], memory_per_pod=mem,
                    vendor_code=vendor, start_pod_suggestion=rule_based_actions.get("start_pod_number", 1),
                    pods_assigned_in_this_batch=batch_state["pods_assigned"], db_locked_pods=initial_data["db_locked_pods"],
                    current_host_capacities_gb=batch_state["capacities"]
                )

                display_doc = part_doc.copy()
                display_doc["initial_interactive_sub_rows_student"] = assignments
                display_doc["student_assignment_warning"] = warn
                processed_items.append((part_doc, display_doc))

            return processed_items
            
        # --- Standard (non-Maestro) Build Logic ---
        else:
            final_lb_course, warning = None, None
            available_courses = initial_data["available_lab_courses"].get(vendor, set())
            user_lb_course = course_input.get('labbuild_course')
            if user_lb_course and user_lb_course in available_courses: final_lb_course = user_lb_course
            else:
                if user_lb_course: warning = f"User selection '{user_lb_course}' invalid. "
                rule_course = rule_based_actions.get('set_labbuild_course')
                if rule_course and rule_course in available_courses: final_lb_course = rule_course
                elif rule_course: warning = (warning or "") + f"Rule course '{rule_course}' is invalid."
            
            pods_req_str, pods_req_calc_match = course_input.get('sf_pods_req', '1'), re.match(r"^\s*(\d+)", str(course_input.get('sf_pods_req', '1')))
            eff_pods_req = int(pods_req_calc_match.group(1)) if pods_req_calc_match else 1
            if rule_based_actions.get("override_pods_req") is not None: eff_pods_req = int(rule_based_actions.get("override_pods_req"))
            if rule_based_actions.get("set_max_pods") is not None: eff_pods_req = min(eff_pods_req, int(rule_based_actions.get("set_max_pods")))
            
            mem_per_pod = _get_memory_for_course_local(final_lb_course, {c['course_name']: c for c in initial_data["course_configs"]}) if final_lb_course else 0.0

            hosts_to_try = rule_based_actions.get("host_priority", initial_data["all_available_host_names"])
            if rule_based_actions.get("allow_spillover", True):
                hosts_to_try = list(dict.fromkeys(hosts_to_try + initial_data["all_available_host_names"]))

            assignments, auto_assign_warn = _propose_assignments(
                num_pods_to_allocate=max(0, eff_pods_req), hosts_to_try=hosts_to_try, memory_per_pod=mem_per_pod,
                vendor_code=vendor, start_pod_suggestion=max(1, int(rule_based_actions.get('start_pod_number', 1))),
                pods_assigned_in_this_batch=batch_state["pods_assigned"], db_locked_pods=initial_data["db_locked_pods"],
                current_host_capacities_gb=batch_state["capacities"]
            )
            
            interim_doc = {
                "_id": ObjectId(), "batch_review_id": batch_review_id, "created_at": datetime.datetime.now(pytz.utc), "status": "pending_student_review",
                "sf_course_code": sf_code, "sf_course_type": sf_course_type, "sf_trainer_name": course_input.get('sf_trainer_name', 'N/A'),
                "sf_start_date": _format_date_for_review(course_input.get('sf_start_date'), sf_code), "sf_end_date": _format_date_for_review(course_input.get('sf_end_date'), sf_code),
                "sf_pax_count": int(course_input.get('sf_pax_count', 0)),
                "vendor": vendor, "final_labbuild_course": final_lb_course, "effective_pods_req": max(0, eff_pods_req),
                "memory_gb_one_pod": mem_per_pod, "assignments": []
            }
            if vendor == 'f5':
                next_class_num = batch_state['f5_class_number_cursor']
                while next_class_num in initial_data['db_locked_pods'].get('f5', set()): next_class_num += 1
                interim_doc['f5_class_number'] = next_class_num
                batch_state['f5_class_number_cursor'] = next_class_num + 1

            final_warning = (warning + ". " if warning else "") + (auto_assign_warn or "")
            if not assignments and interim_doc["effective_pods_req"] > 0: final_warning += " Could not auto-assign any pods."
            
            display_doc = interim_doc.copy()
            display_doc["initial_interactive_sub_rows_student"] = assignments
            display_doc["student_assignment_warning"] = final_warning.strip() or None
            return [(interim_doc, display_doc)] # Return as a list for consistency

    def _save_interim_proposals(proposals, batch_review_id):
        if not proposals: return True
        try:
            update_ops = [UpdateOne({"_id": p["_id"]}, {"$set": p}, upsert=True) for p in proposals]
            result = interim_alloc_collection.bulk_write(update_ops)
            logger.info(f"Interim proposals saved to batch '{batch_review_id}': Upserted: {result.upserted_count}")
            return True
        except (PyMongoError, BulkWriteError) as e:
            logger.error(f"Error saving proposals to interim DB: {e}", exc_info=True)
            flash("Error saving initial review session data.", "danger")
            return False

    # --- 2. Main execution flow of the route ---
    if not selected_courses_json:
        flash("No courses selected for review.", "warning")
        return redirect(url_for('main.view_upcoming_courses'))
    try:
        selected_courses_input = json.loads(selected_courses_json)
        if not isinstance(selected_courses_input, list): raise ValueError
    except (json.JSONDecodeError, ValueError):
        flash("Invalid data received for review.", "danger")
        return redirect(url_for('main.view_upcoming_courses'))

    initial_data = _fetch_and_prepare_initial_data()
    if initial_data is None:
        return redirect(url_for('main.view_upcoming_courses'))
    
    batch_review_id = str(ObjectId())
    batch_state = {
        "capacities": initial_data["host_memory_start_of_batch"].copy(),
        "pods_assigned": defaultdict(set),
        "f5_class_number_cursor": initial_data.get("f5_highest_class_number_used", 0) + 1
    }
    
    try:
        if interim_alloc_collection is not None:
            del_res = interim_alloc_collection.delete_many({})
            logger.info(f"Purged {del_res.deleted_count} old interim allocation documents. Starting new batch: {batch_review_id}")
    except PyMongoError as e:
        logger.error(f"Error purging old interim items: {e}", exc_info=True)
        
    db_docs_to_save, display_docs_for_template = [], []
    for course_input in selected_courses_input:
        processed_items = _process_single_course(course_input, initial_data, batch_state, batch_review_id)
        
        for interim_db_doc, display_template_doc in processed_items:
            db_docs_to_save.append(interim_db_doc)
            display_template_doc['interim_doc_id'] = str(interim_db_doc['_id'])
            display_template_doc['final_labbuild_course_student'] = display_template_doc.get('final_labbuild_course')
            display_template_doc['effective_pods_req_student'] = display_template_doc.get('effective_pods_req')
            display_docs_for_template.append(display_template_doc)
            
    if not _save_interim_proposals(db_docs_to_save, batch_review_id):
        return redirect(url_for('main.view_upcoming_courses'))

    return render_template(
        'intermediate_build_review.html',
        selected_courses=display_docs_for_template,
        batch_review_id=batch_review_id,
        current_theme=current_theme,
        all_hosts=initial_data["all_available_host_names"],
        course_configs_for_memory=initial_data["course_configs"],
        subsequent_pod_memory_factor=SUBSEQUENT_POD_MEMORY_FACTOR
    )


@bp.route('/prepare-trainer-pods', methods=['POST'])
def prepare_trainer_pods():
    """
    Receives finalized student assignments, saves them, and then proposes
    assignments for the corresponding trainer pods for a final review.
    This version ensures contiguous allocation of trainer pods for each vendor
    and allows allocation on any available host if priority hosts are full.
    """
    current_theme = request.cookies.get('theme', 'light')
    final_student_plan_json = request.form.get('final_build_plan')
    batch_review_id = request.form.get('batch_review_id')

    # --- 1. Internal Helper Functions ---

    def _update_finalized_student_assignments(batch_id, student_plan_data):
        """
        STEP 1: Persist Student Assignments.
        Updates the interim DB to mark student assignments as 'student_confirmed'
        and returns the full, updated documents for the next step.
        """
        if not student_plan_data:
            return None, "No student plan data provided."
        
        update_ops = []
        doc_ids_to_fetch = []
        for item in student_plan_data:
            doc_id_str = item.get('interim_doc_id')
            if not doc_id_str:
                logger.warning("Missing 'interim_doc_id' in student data.")
                continue
            
            doc_ids_to_fetch.append(ObjectId(doc_id_str))
            # Create a lean payload containing only the user's final decisions
            update_payload = {
                "assignments": item.get("interactive_assignments", []),
                "status": "student_confirmed", 
                "updated_at": datetime.datetime.now(pytz.utc)
            }
            update_ops.append(UpdateOne({"_id": ObjectId(doc_id_str), "batch_review_id": batch_id}, {"$set": update_payload}))

        try:
            if update_ops:
                interim_alloc_collection.bulk_write(update_ops)
            
            confirmed_docs = list(interim_alloc_collection.find({"_id": {"$in": doc_ids_to_fetch}}))
            return confirmed_docs, None
        except (PyMongoError, BulkWriteError, InvalidId) as e:
            logger.error(f"Error updating student assignments in interim DB for batch '{batch_id}': {e}", exc_info=True)
            return None, "Error saving confirmed student assignments."

    def _prepare_data_for_trainer_logic(confirmed_student_courses):
        """
        Aggregates all data needed for the trainer pod allocation logic.
        """
        prep_data = {
            "build_rules": [], "all_hosts_dd": [], "working_host_caps_gb": {},
            "taken_pods_vendor": defaultdict(set), "lb_courses_on_host": defaultdict(set),
            "course_configs_for_mem": []
        }
        try:
            prep_data["build_rules"] = list(build_rules_collection.find().sort("priority", ASCENDING))
            
            hosts_docs = list(host_collection.find({"include_for_build": "true"}))
            prep_data["all_hosts_dd"] = sorted([hd['host_name'] for hd in hosts_docs if 'host_name' in hd])
            initial_host_caps = get_hosts_available_memory_parallel(hosts_docs) if hosts_docs else {}
            prep_data["working_host_caps_gb"] = {h: cap for h, cap in initial_host_caps.items() if cap is not None}
            
            db_cursor = alloc_collection.find({"extend": "true"}, {"courses.vendor": 1, "courses.pod_details": 1})
            for tag_doc in db_cursor:
                for course in tag_doc.get("courses", []):
                    vendor = course.get("vendor", "").lower()
                    if not vendor: continue
                    for pd in course.get("pod_details", []):
                        if pd.get("pod_number") is not None: prep_data["taken_pods_vendor"][vendor].add(pd.get("pod_number"))
                        if vendor == 'f5':
                            if pd.get("class_number") is not None: prep_data["taken_pods_vendor"][vendor].add(pd.get("class_number"))
                            for np in pd.get("pods", []):
                                if np.get("pod_number") is not None: prep_data["taken_pods_vendor"][vendor].add(np.get("pod_number"))
            
            for stud_course in confirmed_student_courses:
                vendor, lb_course = stud_course.get("vendor", "").lower(), stud_course.get("final_labbuild_course")
                mem_one_pod = float(stud_course.get("memory_gb_one_pod", 0.0))
                for asgn in stud_course.get("assignments", []):
                    host = asgn.get("host")
                    if not all([host, vendor, lb_course]): continue
                    prep_data["lb_courses_on_host"][host].add(lb_course)
                    try:
                        s_pod, e_pod = int(asgn.get('start_pod', 0)), int(asgn.get('end_pod', 0))
                        if s_pod > 0 and e_pod >= s_pod:
                            num_pods = e_pod - s_pod + 1
                            for i in range(s_pod, e_pod + 1): prep_data["taken_pods_vendor"][vendor].add(i)
                            mem_deduct = mem_one_pod + ((num_pods - 1) * mem_one_pod * SUBSEQUENT_POD_MEMORY_FACTOR if num_pods > 1 else 0)
                            prep_data["working_host_caps_gb"][host] = max(0, round(prep_data["working_host_caps_gb"].get(host, 0.0) - mem_deduct, 2))
                    except (ValueError, TypeError): pass
            
            prep_data["course_configs_for_mem"] = list(course_config_collection.find({}, {"course_name": 1, "memory_gb_per_pod": 1, "memory": 1, "_id": 0}))
            return prep_data, None
        except (PyMongoError, Exception) as e:
            logger.error(f"Error preparing data for trainer logic: {e}", exc_info=True)
            return None, "Error preparing data for trainer pod assignment."

    def _propose_trainer_pod_for_course(student_course_doc, prep_data, batch_state):
        """
        STEP 3: Propose a single Trainer Pod using a per-vendor cursor and flexible host allocation.
        Determines the assignment for a single trainer pod based on rules and available resources.
        Modifies batch_state in place and returns the DB payload and display data.
        """
        student_sf_code, vendor, sf_type, f5_class_number = student_course_doc.get('sf_course_code'), student_course_doc.get('vendor', '').lower(), student_course_doc.get('sf_course_type', ''), student_course_doc.get('f5_class_number')
        tp_sf_code = f"{student_sf_code}-TP"
        
        relevant_rules = _find_all_matching_rules(prep_data["build_rules"], vendor, student_sf_code, sf_type)
        is_disabled = any(r.get("actions", {}).get("disable_trainer_pod") for r in relevant_rules)

        db_update_payload = {"updated_at": datetime.datetime.now(pytz.utc)}
        display_data = {"interim_doc_id": str(student_course_doc["_id"]), "sf_course_code": tp_sf_code, "vendor": vendor.upper(), "f5_class_number": f5_class_number}

        if is_disabled:
            db_update_payload.update({"trainer_labbuild_course": None, "trainer_assignment": None, "trainer_assignment_warning": "Trainer pod build explicitly disabled by rule."})
            display_data.update({"labbuild_course": "N/A - Disabled", "error_message": db_update_payload["trainer_assignment_warning"]})
        else:
            lb_course, mem_one_pod = student_course_doc.get('final_labbuild_course'), float(student_course_doc.get('memory_gb_one_pod', 0.0))
            tp_rules = _find_all_matching_rules(prep_data["build_rules"], vendor, tp_sf_code, sf_type)
            eff_actions_tp = {}
            for r in tp_rules:
                for k in ["set_labbuild_course", "host_priority", "start_pod_number"]:
                    if k not in eff_actions_tp and k in r.get("actions", {}): eff_actions_tp[k] = r["actions"][k]
            
            if eff_actions_tp.get("set_labbuild_course"): lb_course = eff_actions_tp.get("set_labbuild_course")
            if lb_course != student_course_doc.get('final_labbuild_course'): mem_one_pod = _get_memory_for_course_local(lb_course, {c['course_name']: c for c in prep_data["course_configs_for_mem"]})

            db_update_payload.update({"trainer_labbuild_course": lb_course, "trainer_memory_gb_one_pod": mem_one_pod})
            display_data.update({"labbuild_course": lb_course, "memory_gb_one_pod": mem_one_pod})
            
            host, pod, warning = None, None, None
            if lb_course and mem_one_pod > 0:
                # --- MODIFIED HOST SELECTION LOGIC ---
                priority_hosts_from_rule = eff_actions_tp.get("host_priority")
                all_available_hosts = prep_data.get("all_hosts_dd", [])
                
                if isinstance(priority_hosts_from_rule, list) and priority_hosts_from_rule:
                    # Construct a final list that tries priority hosts first, then spills over to all others.
                    # dict.fromkeys preserves order and removes duplicates.
                    hosts_to_try = list(dict.fromkeys(priority_hosts_from_rule + all_available_hosts))
                    logger.info(f"Trainer pod for '{tp_sf_code}' will try priority hosts first: {priority_hosts_from_rule}, then spill over.")
                else:
                    # If no priority list is defined, simply try all available hosts.
                    hosts_to_try = all_available_hosts
                    logger.info(f"Trainer pod for '{tp_sf_code}' will try all available hosts: {hosts_to_try}")
                # --- END OF MODIFICATION ---

                start_pod_sugg_from_rule = eff_actions_tp.get("start_pod_number", 1)
                cand_pod = batch_state["vendor_pod_cursors"].get(vendor, start_pod_sugg_from_rule)
                
                pod_found = False
                for host_cand in hosts_to_try:
                    search_pod = cand_pod
                    
                    while True: # Keep searching for the next available pod number
                        if search_pod not in batch_state["taken_pods_vendor"].get(vendor, set()):
                            is_first = lb_course not in batch_state["lb_courses_on_host"].get(host_cand, set())
                            mem_needed = round(mem_one_pod * (SUBSEQUENT_POD_MEMORY_FACTOR if not is_first else 1.0), 2)
                            host_cap = batch_state["working_host_caps_gb"].get(host_cand, 0.0)
                            
                            if host_cap >= mem_needed:
                                host, pod = host_cand, search_pod
                                pod_found = True
                                break # Found a spot on this host
                            else:
                                # Not enough memory on this host for any pod of this type, try next host.
                                break 
                        search_pod += 1
                    
                    if pod_found:
                        break # Exit the host loop since we found a spot

                if pod_found:
                    batch_state["taken_pods_vendor"][vendor].add(pod)
                    is_first_on_host = lb_course not in batch_state["lb_courses_on_host"].get(host, set())
                    mem_consumed = round(mem_one_pod * (SUBSEQUENT_POD_MEMORY_FACTOR if not is_first_on_host else 1.0), 2)
                    batch_state["working_host_caps_gb"][host] = max(0, round(batch_state["working_host_caps_gb"].get(host, 0.0) - mem_consumed, 2))
                    batch_state["lb_courses_on_host"][host].add(lb_course)
                    batch_state["vendor_pod_cursors"][vendor] = pod + 1
                    logger.info(f"Assigned trainer pod {pod} to {host}. Next cursor for '{vendor}' is {pod + 1}.")
                else:
                    warning = "No suitable host found with available capacity and a free contiguous pod number."
            else:
                warning = "Cannot assign trainer pod: Missing LabBuild course or memory is zero."
            
            db_update_payload.update({"trainer_assignment": {"host": host, "start_pod": pod, "end_pod": pod} if host and pod is not None else None, "trainer_assignment_warning": warning})
            display_data.update({"assigned_host": host, "start_pod": pod, "end_pod": pod, "error_message": warning})
            
        return db_update_payload, display_data

    def _save_trainer_proposals(trainer_updates, batch_id):
        """STEP 4: Save Trainer Proposals to the Database."""
        if not trainer_updates: return True
        try:
            update_ops = [UpdateOne({"_id": item["_id"], "batch_review_id": batch_id}, {"$set": item["payload"]}) for item in trainer_updates]
            result = interim_alloc_collection.bulk_write(update_ops)
            logger.info(f"Updated {result.modified_count} interim docs with trainer info for batch '{batch_id}'.")
            return True
        except (PyMongoError, BulkWriteError) as e:
            logger.error(f"Error saving trainer proposals to interim DB: {e}", exc_info=True)
            flash("Error saving trainer pod proposals.", "danger")
            return False

    # --- Main execution flow of the route ---
    if not batch_review_id or not final_student_plan_json:
        flash("Session ID or student plan missing.", "danger")
        return redirect(url_for('main.view_upcoming_courses'))
    
    try:
        student_plan_data = json.loads(final_student_plan_json)
        if not isinstance(student_plan_data, list): raise ValueError
    except (json.JSONDecodeError, ValueError):
        flash("Error processing submitted student assignments.", "danger")
        return redirect(url_for('actions.intermediate_build_review', batch_review_id=batch_review_id))

    confirmed_student_courses, err = _update_finalized_student_assignments(batch_review_id, student_plan_data)
    if err:
        flash(err, "danger"); return redirect(url_for('actions.intermediate_build_review', batch_review_id=batch_review_id))

    confirmed_student_courses.sort(key=lambda x: (x.get('vendor', 'zzz'), x.get('sf_start_date', '')))

    prep_data, err = _prepare_data_for_trainer_logic(confirmed_student_courses)
    if err:
        flash(err, "danger"); return redirect(url_for('actions.intermediate_build_review', batch_review_id=batch_review_id))

    batch_trainer_state = {
        "working_host_caps_gb": prep_data["working_host_caps_gb"],
        "taken_pods_vendor": prep_data["taken_pods_vendor"],
        "lb_courses_on_host": prep_data["lb_courses_on_host"],
        "vendor_pod_cursors": defaultdict(lambda: 100)
    }
    
    trainer_updates_for_db = []
    display_items_for_template = []
    
    for student_doc in confirmed_student_courses:
        trainer_payload, trainer_display_data = _propose_trainer_pod_for_course(student_doc, prep_data, batch_trainer_state)
        trainer_updates_for_db.append({"_id": student_doc["_id"], "payload": trainer_payload})
        student_doc.update(trainer_payload)
        display_items_for_template.append(student_doc)

    if not _save_trainer_proposals(trainer_updates_for_db, batch_review_id):
        return redirect(url_for('actions.intermediate_build_review', batch_review_id=batch_review_id))

    return render_template(
        'trainer_pod_review.html',
        courses_and_trainer_pods=display_items_for_template,
        batch_review_id=batch_review_id,
        current_theme=current_theme,
        all_hosts=prep_data["all_hosts_dd"],
        course_configs_for_memory=prep_data["course_configs_for_mem"],
        subsequent_pod_memory_factor=SUBSEQUENT_POD_MEMORY_FACTOR
    )

def _generate_apm_data_for_plan(
        local_final_build_plan_items: List[Dict], # Pass the already fetched interim items
        local_current_apm_entries: Dict[str, Dict[str, Any]],
        local_extended_apm_course_codes: Set[str]
    ) -> Tuple[Dict[str, Dict[str, str]], List[str], List[str]]:
        """
        Generates APM usernames and passwords for the items in the build plan.
        Returns:
            - apm_credentials_map (Dict[apm_auth_code, {"username": str, "password": str}])
            - all_generated_apm_commands (List[str])
            - apm_generation_errors (List[str])
        """
        logger.info(f"--- (Helper) APM Data Generation Started for {len(local_final_build_plan_items)} plan items ---")
        apm_commands_delete_local: List[str] = []
        apm_commands_add_update_local: List[str] = []
        errors_local: List[str] = []
        
        # This will store the final state including allocated usernames and passwords
        # Key: apm_auth_course_code, Value: { vpn_auth_courses, vpn_auth_range, ..., username, password }
        desired_apm_state_with_users: Dict[str, Dict[str, Any]] = {}

        # Step 3a: Build intermediate desired state (details without username yet)
        desired_apm_config_intermediate: Dict[str, Dict[str, Any]] = defaultdict(lambda: {
            "all_pod_numbers": set(), "trainer_name_for_desc": "N/A", "course_type_for_desc": "N/A",
            "labbuild_course_for_version": "N/A", "vendor_short": "xx",
            "is_maestro_overall_for_version": False, "maestro_main_lb_for_version": None
        })

        for item_doc in local_final_build_plan_items:
            original_sf_code = item_doc.get("original_sf_course_code", item_doc.get("sf_course_code"))
            if not original_sf_code: continue

            # Determine if this is for Student APM entry or Trainer APM entry
            # A single interim 'doc' can lead to two APM entries (one for student, one for trainer)
            
            # Process Student Part from 'item_doc'
            if item_doc.get("status") in ["student_confirmed", "trainer_confirmed"]: # trainer_confirmed implies student part is also relevant
                student_assignments = item_doc.get("student_assignments_final", [])
                if student_assignments:
                    apm_auth_code_key_student = original_sf_code
                    student_entry_details = desired_apm_config_intermediate[apm_auth_code_key_student]
                    if student_entry_details["trainer_name_for_desc"] == "N/A": # Initialize if first time for this key
                        student_entry_details["trainer_name_for_desc"] = item_doc.get("sf_trainer_name", "N/A")
                        student_entry_details["course_type_for_desc"] = item_doc.get("sf_course_type", "Course")
                        student_entry_details["vendor_short"] = (item_doc.get("vendor", "xx") or "xx").lower()
                        maestro_config = item_doc.get("maestro_split_config_details_rule")
                        if isinstance(maestro_config, dict) and item_doc.get("vendor") == "cp":
                            student_entry_details["is_maestro_overall_for_version"] = True
                            student_entry_details["maestro_main_lb_for_version"] = maestro_config.get("main_course", item_doc.get("final_labbuild_course_student", "maestro"))
                        else:
                            student_entry_details["labbuild_course_for_version"] = item_doc.get("final_labbuild_course_student", item_doc.get("labbuild_course", "N/A"))
                    for asgn in student_assignments:
                        s, e = asgn.get("start_pod"), asgn.get("end_pod")
                        if s is not None and e is not None: 
                            try: [student_entry_details["all_pod_numbers"].add(p) for p in range(int(s),int(e)+1)]; 
                            except:pass
            
            # Process Trainer Part from 'item_doc'
            if item_doc.get("status") == "trainer_confirmed": # Only if trainer pod is to be built
                trainer_assignments = item_doc.get("trainer_assignment_final", [])
                if trainer_assignments:
                    apm_auth_code_key_trainer = f"{original_sf_code}-TP"
                    trainer_entry_details = desired_apm_config_intermediate[apm_auth_code_key_trainer]
                    if trainer_entry_details["trainer_name_for_desc"] == "N/A":
                        trainer_entry_details["trainer_name_for_desc"] = "Trainer Pods"
                        trainer_entry_details["course_type_for_desc"] = item_doc.get("sf_course_type", "Trainer Setup")
                        trainer_entry_details["labbuild_course_for_version"] = item_doc.get("trainer_labbuild_course", "N/A")
                        trainer_entry_details["vendor_short"] = (item_doc.get("vendor", "xx") or "xx").lower()
                    for asgn in trainer_assignments:
                        s, e = asgn.get("start_pod"), asgn.get("end_pod")
                        if s is not None and e is not None: 
                            try: [trainer_entry_details["all_pod_numbers"].add(p) for p in range(int(s),int(e)+1)]; 
                            except:pass
        
        # Step 3b: Finalize intermediate state (convert pod sets to ranges, etc.)
        for apm_auth_code, details in desired_apm_config_intermediate.items():
            if not details["all_pod_numbers"]: continue
            vpn_range = _create_contiguous_ranges(list(details["all_pod_numbers"]))
            if not vpn_range: continue
            desc = f"{details['trainer_name_for_desc']} - {details['course_type_for_desc']}"
            ver = details["labbuild_course_for_version"]
            if details["is_maestro_overall_for_version"] and details["maestro_main_lb_for_version"]: ver = details["maestro_main_lb_for_version"]
            
            desired_apm_state_with_users[apm_auth_code] = {
                "vpn_auth_courses": desc[:250], "vpn_auth_range": vpn_range,
                "vpn_auth_version": ver, "vpn_auth_expiry_date": "8",
                "password": _generate_random_password(), "vendor_short": details["vendor_short"]
                # Username to be added next
            }
        logger.info(f"(Helper) Prepared {len(desired_apm_state_with_users)} desired APM entries structure.")

        # Step 4 & 5: Username Allocation, Generate Commands
        # ... (This is the same complex logic as in your last working `generate_apm_commands_route`
        #      that iterates `current_apm_entries` and `desired_apm_state_with_users` (which was named desired_apm_state_from_plan before)
        #      to generate `apm_commands_delete_local`, `apm_commands_add_update_local`,
        #      and populates `details["username"]` in `desired_apm_state_with_users`)
        apm_course_code_to_username_to_keep: Dict[str, str] = {} 
        for username, existing_details in local_current_apm_entries.items(): # Use local_
            apm_cc = existing_details.get("vpn_auth_course_code")
            if not apm_cc: apm_commands_delete_local.append(f"course2 del {username}"); continue
            if apm_cc in local_extended_apm_course_codes: # Use local_
                if apm_cc not in apm_course_code_to_username_to_keep: apm_course_code_to_username_to_keep[apm_cc] = username
                if apm_cc in desired_apm_state_with_users: desired_apm_state_with_users[apm_cc]["username"] = username; desired_apm_state_with_users[apm_cc]["is_update"] = True
            elif apm_cc in desired_apm_state_with_users:
                if apm_cc not in apm_course_code_to_username_to_keep: apm_course_code_to_username_to_keep[apm_cc] = username
                new_d = desired_apm_state_with_users[apm_cc]
                if existing_details.get("vpn_auth_range") != new_d["vpn_auth_range"]: apm_commands_add_update_local.append(f'course2 range {username} "{new_d["vpn_auth_range"]}"')
                if existing_details.get("vpn_auth_version") != new_d["vpn_auth_version"]: apm_commands_add_update_local.append(f'course2 version {username} "{new_d["vpn_auth_version"]}"')
                if existing_details.get("vpn_auth_courses") != new_d["vpn_auth_courses"]: apm_commands_add_update_local.append(f'course2 description {username} "{new_d["vpn_auth_courses"]}"')
                if str(existing_details.get("vpn_auth_expiry_date")) != str(new_d["vpn_auth_expiry_date"]): apm_commands_add_update_local.append(f'course2 expiry_date {username} "{new_d["vpn_auth_expiry_date"]}"')
                apm_commands_add_update_local.append(f'course2 password {username} "{new_d["password"]}"')
                desired_apm_state_with_users[apm_cc]["username"] = username; desired_apm_state_with_users[apm_cc]["is_update"] = True
            else: apm_commands_delete_local.append(f"course2 del {username}")
        
        used_x_nums: Dict[str,Set[int]] = defaultdict(set)
        for _, uname_kept in apm_course_code_to_username_to_keep.items():
            m = re.match(r"lab([a-z0-9]+)-(\d+)",uname_kept.lower());
            if m: used_x_nums[m.group(1)].add(int(m.group(2)))

        for apm_auth_c, details_final in desired_apm_state_with_users.items():
            if details_final.get("is_update"): continue
            v_short = details_final["vendor_short"]; uname=None; x=1
            while True:
                if x not in used_x_nums.get(v_short,set()): uname=f"lab{v_short}-{x}"; used_x_nums[v_short].add(x); break
                x+=1
            details_final["username"] = uname # Store assigned username
            apm_commands_add_update_local.append(f'course2 add {uname} "{details_final["password"]}" "{details_final["vpn_auth_range"]}" "{details_final["vpn_auth_version"]}" "{details_final["vpn_auth_courses"]}" "{details_final["vpn_auth_expiry_date"]}" "{apm_auth_c}"')

        all_generated_commands = apm_commands_delete_local + apm_commands_add_update_local
        logger.info(f"(Helper) APM Logic Finished. Generated {len(all_generated_commands)} commands. Errors: {errors_local}")
        
        # Return the map of apm_auth_code to {"username": ..., "password": ...}
        apm_credentials_map_result: Dict[str, Dict[str, str]] = {}
        for apm_auth_code, details_with_user in desired_apm_state_with_users.items():
            if details_with_user.get("username") and details_with_user.get("password"):
                apm_credentials_map_result[apm_auth_code] = {
                    "username": details_with_user["username"],
                    "password": details_with_user["password"]
                }
        
        return all_generated_commands, apm_credentials_map_result, errors_local

@bp.route('/finalize-and-display-build-plan', methods=['POST'])
def finalize_and_display_build_plan():
    """
    Receives final trainer decisions, updates the DB with all final assignments
    (including APM credentials), and displays the full build/teardown plan.
    """
    current_theme = request.cookies.get('theme', 'light')
    batch_review_id = request.form.get('batch_review_id')
    final_trainer_plan_json = request.form.get('final_trainer_assignments_for_review')
    logger.info(f"[finalize_and_display_build_plan] Starting for Batch ID: {batch_review_id}")

    # --- Internal Helper Functions ---

    def _update_finalized_trainer_assignments(batch_id, trainer_assignments_data):
        # ... (This helper is correct and remains unchanged) ...
        if not trainer_assignments_data: return True, "No trainer assignments to update."
        update_ops: List[UpdateOne] = []
        try:
            if interim_alloc_collection is None: return False, "Database service unavailable."
            for trainer_data in trainer_assignments_data:
                doc_id_str = trainer_data.get('interim_doc_id')
                build_trainer = trainer_data.get('build_trainer', False)
                if not doc_id_str: logger.warning("Missing 'interim_doc_id' in trainer data."); continue
                
                payload: Dict[str, Any] = {"updated_at": datetime.datetime.now(pytz.utc)}
                if build_trainer:
                    payload.update({
                        "trainer_labbuild_course": trainer_data.get("labbuild_course"),
                        "trainer_memory_gb_one_pod": trainer_data.get("memory_gb_one_pod"),
                        "trainer_assignment": trainer_data.get("interactive_assignments", []),
                        "trainer_assignment_warning": trainer_data.get("error_message"),
                        "status": "trainer_confirmed"
                    })
                else:
                    payload.update({"trainer_assignment": None, "status": "trainer_skipped_by_user", "trainer_assignment_warning": "Build skipped by user."})
                update_ops.append(UpdateOne({"_id": ObjectId(doc_id_str), "batch_review_id": batch_id}, {"$set": payload}))
            
            if update_ops:
                result = interim_alloc_collection.bulk_write(update_ops)
                logger.info(f"Updated {result.modified_count} trainer assignments in interim (batch '{batch_id}').")
            return True, None
        except Exception as e:
            logger.error(f"Error saving finalized trainer assignments to interim DB: {e}", exc_info=True)
            return False, "Error saving finalized trainer assignments."

    def _fetch_all_plan_data(batch_id):
        # ... (This helper is correct) ...
        finalized_statuses = ["student_confirmed", "trainer_confirmed", "trainer_skipped_by_user", "trainer_disabled_by_rule"]
        try:
            if interim_alloc_collection is None: return [], "Database service unavailable."
            cursor = interim_alloc_collection.find(
                {"batch_review_id": batch_id, "status": {"$in": finalized_statuses}}
            ).sort([("sf_start_date", ASCENDING), ("sf_course_code", ASCENDING)])
            return list(cursor), None
        except PyMongoError as e:
            logger.error(f"Error fetching final plan items for batch '{batch_id}': {e}", exc_info=True)
            return [], "Error fetching review data."
    
    def _generate_apm_data_for_plan(final_plan_items: List[Dict]) -> Tuple[List[str], Dict[str, Any], List[str]]:
        """
        STEP 3: Generate APM Commands.
        Generates APM commands and credentials based on the final build plan.
        """
        logger.info(f"--- APM Data Generation Started for {len(final_plan_items)} plan items ---")
        
        apm_commands_delete, apm_commands_add_update, apm_errors = [], [], []
        desired_apm_state, apm_credentials_map = defaultdict(lambda: {"all_pod_numbers": set()}), {}

        # Fetch current APM state
        try:
            apm_list_url = os.getenv("APM_LIST_URL", "http://connect.rededucation.com:1212/list")
            response = requests.get(apm_list_url, timeout=15); response.raise_for_status()
            current_apm_entries = response.json()
        except Exception as e:
            apm_errors.append(f"Could not fetch current APM state: {e}")
            current_apm_entries = {}
        
        # Fetch extended (keep-alive) tags
        extended_apm_codes = set()
        if alloc_collection is not None:
            try:
                for doc in alloc_collection.find({"extend": "true"}, {"tag": 1, "_id": 0}):
                    if doc.get("tag"): extended_apm_codes.add(doc.get("tag"))
            except PyMongoError: apm_errors.append("Could not fetch extended allocation tags.")

        # --- FIX IS HERE: Ensure 'details' dict is always populated ---
        for item in final_plan_items:
            sf_code = item.get("sf_course_code")
            if not sf_code: continue

            # Student pods
            if item.get("assignments"):
                if "details" not in desired_apm_state[sf_code]:
                    desired_apm_state[sf_code]["details"] = {"trainer": item.get("sf_trainer_name", "N/A"), "type": item.get("sf_course_type", "Course"), "version": item.get("final_labbuild_course", "N/A"), "vendor": item.get("vendor", "xx").lower()}
                for asgn in item.get("assignments", []):
                    s, e = asgn.get("start_pod"), asgn.get("end_pod")
                    if s is not None and e is not None: [desired_apm_state[sf_code]["all_pod_numbers"].add(p) for p in range(int(s), int(e) + 1)]
            
            # Trainer pods
            if item.get("trainer_assignment"):
                key_tp = f"{sf_code}-TP"
                if "details" not in desired_apm_state[key_tp]:
                     desired_apm_state[key_tp]["details"] = {"trainer": "Trainer Pods", "type": item.get("sf_course_type", "Trainer Setup"), "version": item.get("trainer_labbuild_course", "N/A"), "vendor": item.get("vendor", "xx").lower()}
                for asgn_tp in item["trainer_assignment"]:
                    s, e = asgn_tp.get("start_pod"), asgn_tp.get("end_pod")
                    if s is not None and e is not None: [desired_apm_state[key_tp]["all_pod_numbers"].add(p) for p in range(int(s), int(e) + 1)]
        
        # *** FIX IS HERE ***
        # The 'final_desired_state' dictionary was missing the 'version' key.
        final_desired_state = { 
            code: { 
                "vpn_auth_courses": f"{d['details']['trainer']} - {d['details']['type']}"[:250], 
                "vpn_auth_range": _create_contiguous_ranges(list(d["all_pod_numbers"])), 
                "version": d["details"]["version"],  # Ensure 'version' is included
                "password": _generate_random_password(), 
                "vendor_short": d["details"]["vendor"] 
            } for code, d in desired_apm_state.items() if d["all_pod_numbers"] 
        }
        # *** END FIX ***
        
        # Generate commands
        apm_code_to_user = {}
        for user, details in current_apm_entries.items():
            code = details.get("vpn_auth_course_code")
            if not code:
                apm_commands_delete.append(f"course2 del {user}"); continue
            if code in extended_apm_codes:
                apm_code_to_user[code] = user; continue
            if code in final_desired_state:
                apm_code_to_user[code] = user
                new = final_desired_state[code]
                if str(details.get("vpn_auth_range")) != str(new["vpn_auth_range"]): apm_commands_add_update.append(f'course2 range {user} "{new["vpn_auth_range"]}"')
                if str(details.get("vpn_auth_version")) != str(new["version"]): apm_commands_add_update.append(f'course2 version {user} "{new["version"]}"')
                if str(details.get("vpn_auth_courses")) != str(new["vpn_auth_courses"]): apm_commands_add_update.append(f'course2 description {user} "{new["vpn_auth_courses"]}"')
                apm_commands_add_update.append(f'course2 password {user} "{new["password"]}"')
            else:
                apm_commands_delete.append(f"course2 del {user}")
        
        used_x_nums = defaultdict(set)
        for _, uname in apm_code_to_user.items():
            m = re.match(r"lab([a-z0-9]+)-(\d+)", uname.lower()); 
            if m: used_x_nums[m.group(1)].add(int(m.group(2)))

        for code, details in final_desired_state.items():
            if code in apm_code_to_user:
                apm_credentials_map[code] = {"username": apm_code_to_user[code], "password": details["password"]}
                continue
            vendor, x = details["vendor_short"], 1
            while True:
                if x not in used_x_nums[vendor]: new_user = f"lab{vendor}-{x}"; used_x_nums[vendor].add(x); break
                x += 1
            apm_commands_add_update.append(f'course2 add {new_user} "{details["password"]}" "{details["vpn_auth_range"]}" "{details["version"]}" "{details["vpn_auth_courses"]}" "8" "{code}"')
            apm_credentials_map[code] = {"username": new_user, "password": details["password"]}
        
        return apm_commands_delete + apm_commands_add_update, apm_credentials_map, apm_errors


    def _update_db_with_apm(batch_id, apm_creds_map):
        # ... (This helper is correct and remains unchanged) ...
        if not apm_creds_map: return True, "No APM credentials to update."
        update_ops: List[UpdateOne] = []
        for apm_code, creds in apm_creds_map.items():
            if apm_code.endswith("-TP"):
                sf_code = apm_code[:-3]
                payload = {"trainer_apm_username": creds.get("username"), "trainer_apm_password": creds.get("password")}
                update_ops.append(UpdateOne({"batch_review_id": batch_id, "sf_course_code": sf_code}, {"$set": payload}))
            else:
                payload = {"student_apm_username": creds.get("username"), "student_apm_password": creds.get("password")}
                update_ops.append(UpdateOne({"batch_review_id": batch_id, "sf_course_code": apm_code}, {"$set": payload}))
        try:
            if update_ops and interim_alloc_collection is not None: interim_alloc_collection.bulk_write(update_ops)
            return True, None
        except Exception as e:
            logger.error(f"Error updating interim DB with APM credentials: {e}", exc_info=True)
            return False, "Database error while saving APM credentials."

    def _process_plan_for_display(final_plan_docs):
        """
        STEP 6: Prepare Data for Rendering.
        **THIS IS THE CORRECTED VERSION WITH LOGGING**
        """
        processed_items: List[Dict] = []
        logger.info(f"[_process_plan_for_display] Processing {len(final_plan_docs)} final documents for display.")

        for doc in final_plan_docs:
            sf_code = doc.get("sf_course_code")
            vendor = doc.get("vendor")
            logger.debug(f"[_process_plan_for_display] Processing doc for SF Code: {sf_code}")

            # --- Student Part ---
            student_item = {
                "type": "Student Build",
                "sf_course_code": sf_code,
                "original_sf_course_code": sf_code,
                "labbuild_course": doc.get("final_labbuild_course"),
                "sf_course_type": doc.get("sf_course_type"),
                "vendor": vendor,
                "start_date": doc.get("sf_start_date"),
                "end_date": doc.get("sf_end_date"),
                "assignments": doc.get("assignments", []),
                "status_note": doc.get("student_assignment_warning"),
                "sf_trainer_name": doc.get("sf_trainer_name"),
                "f5_class_number": doc.get("f5_class_number"),
                "sf_pax_count": doc.get("sf_pax_count"),
                "effective_pods_req_student": doc.get("effective_pods_req"),
                "memory_gb_one_pod": doc.get("memory_gb_one_pod"),
                "apm_username": doc.get("student_apm_username"),
                "apm_password": doc.get("student_apm_password"),
            }
            processed_items.append(student_item)
            logger.debug(f"[_process_plan_for_display] Added student item: {student_item}")

            # --- Trainer Part ---
            if doc.get("status") in ["trainer_confirmed", "trainer_skipped_by_user", "trainer_disabled_by_rule"]:
                is_skipped = not doc.get("trainer_assignment")
                trainer_sfc_display = sf_code + ("-TP (Skipped)" if is_skipped else "-TP")
                
                trainer_item = {
                    "type": "Trainer Build",
                    "sf_course_code": trainer_sfc_display,
                    "original_sf_course_code": sf_code,
                    "labbuild_course": doc.get("trainer_labbuild_course") or "N/A",
                    "vendor": vendor,
                    "start_date": doc.get("sf_start_date"),
                    "end_date": doc.get("sf_end_date"),
                    "assignments": doc.get("trainer_assignment") or [],
                    "status_note": doc.get("trainer_assignment_warning"),
                    "sf_trainer_name": doc.get("sf_trainer_name"),
                    "f5_class_number": doc.get("f5_class_number"),
                    "apm_username": doc.get("student_apm_username"), # Still needed for context
                    "apm_password": doc.get("student_apm_password"), # Still needed for context
                    "trainer_apm_username": doc.get("trainer_apm_username"), # The crucial field
                    "trainer_apm_password": doc.get("trainer_apm_password")  # The crucial field
                }
                processed_items.append(trainer_item)
                logger.debug(f"[_process_plan_for_display] Added trainer item: {trainer_item}")

        return processed_items
    
    def _sanitize_for_json(data):
        # ... (This helper is correct and remains unchanged) ...
        if isinstance(data, list): return [_sanitize_for_json(i) for i in data]
        if isinstance(data, dict): return {str(k): _sanitize_for_json(v) for k,v in data.items()}
        if isinstance(data, ObjectId): return str(data)
        if isinstance(data, datetime.datetime):
            dt = data.astimezone(pytz.utc) if data.tzinfo else pytz.utc.localize(data)
            return dt.isoformat().replace("+00:00", "Z")
        if isinstance(data, datetime.date): return data.isoformat()
        return data

    # --- Main Route Logic ---
    if not batch_review_id or not final_trainer_plan_json:
        flash("Session ID or final plan missing.", "danger"); return redirect(url_for('main.view_upcoming_courses'))
    
    try:
        trainer_plan_data = json.loads(final_trainer_plan_json)
    except json.JSONDecodeError:
        flash("Error processing final trainer assignments.", "danger"); return redirect(url_for('actions.prepare_trainer_pods', batch_review_id=batch_review_id))

    success, error = _update_finalized_trainer_assignments(batch_review_id, trainer_plan_data)
    if not success:
        flash(error, "danger"); return redirect(url_for('actions.prepare_trainer_pods', batch_review_id=batch_review_id))

    final_plan_docs, error = _fetch_all_plan_data(batch_review_id)
    if error:
        flash(error, "danger"); return redirect(url_for('main.index'))
    logger.info(f"[finalize_and_display_build_plan] Fetched {len(final_plan_docs)} final docs from interim DB.")
    
    apm_commands, apm_creds_map, apm_errors = _generate_apm_data_for_plan(final_plan_docs)
    if apm_errors:
        flash("Note: Errors occurred during APM data generation: " + " | ".join(apm_errors), "info")
    logger.info(f"[finalize_and_display_build_plan] Generated {len(apm_commands)} APM commands and {len(apm_creds_map)} credential pairs.")

    success, error = _update_db_with_apm(batch_review_id, apm_creds_map)
    if not success:
        flash(error, "danger")
        
    final_plan_docs_with_apm, _ = _fetch_all_plan_data(batch_review_id)
    processed_plan_items = _process_plan_for_display(final_plan_docs_with_apm)
    
    sanitized_all_items = _sanitize_for_json(processed_plan_items)
    sanitized_buildable_items = [
        item for item in sanitized_all_items
        if item.get("type") in ["Student Build", "Trainer Build"] and item.get("assignments")
    ]
    logger.info(f"[finalize_and_display_build_plan] Final list of {len(sanitized_all_items)} items being passed to template.")
    logger.debug(f"[finalize_and_display_build_plan] Sanitized data for template: {json.dumps(sanitized_all_items, indent=2)}")

    
    return render_template(
        'final_review_schedule.html',
        all_items_for_review=sanitized_all_items,
        buildable_items_json=json.dumps(sanitized_buildable_items),
        all_review_items_json=json.dumps(sanitized_all_items),
        apm_commands_for_preview=apm_commands,
        batch_review_id=batch_review_id,
        current_theme=current_theme,
    )


@bp.route('/execute-scheduled-builds', methods=['POST'])
def execute_scheduled_builds():
    """
    Receives the confirmed build plan, teardown preference, and scheduling options,
    then schedules the labbuild setup (and optionally teardown) commands.
    """
    confirmed_plan_json = request.form.get('confirmed_build_plan_data')
    schedule_option = request.form.get('schedule_option', 'now')
    batch_review_id = request.form.get('batch_review_id')
    perform_teardown_first = request.form.get('perform_teardown_first') == 'yes'

    logger.info(f"Execute Scheduled Builds for Batch ID '{batch_review_id}': "
                f"Teardown before setup = {perform_teardown_first}, "
                f"Schedule Option = {schedule_option}")

    schedule_start_time_str: Optional[str] = None
    stagger_minutes: int = 30

    if schedule_option == 'specific_time_all':
        schedule_start_time_str = request.form.get('schedule_start_time')
    elif schedule_option == 'staggered':
        schedule_start_time_str = request.form.get('schedule_start_time_staggered')
        try:
            stagger_minutes = int(request.form.get('schedule_stagger_minutes', '30'))
            if stagger_minutes < 1:
                stagger_minutes = 1
        except (ValueError, TypeError):
            stagger_minutes = 30

    if not confirmed_plan_json:
        flash("No confirmed build plan received to schedule.", "danger")
        return redirect(url_for('actions.finalize_and_display_build_plan', batch_review_id=batch_review_id))

    try:
        buildable_items_from_form = json.loads(confirmed_plan_json)
        if not isinstance(buildable_items_from_form, list):
            raise ValueError("Invalid confirmed build plan format.")
    except (json.JSONDecodeError, ValueError) as e:
        flash(f"Error processing confirmed build plan: {e}", "danger")
        return redirect(url_for('actions.finalize_and_display_build_plan', batch_review_id=batch_review_id))

    if scheduler is None or not scheduler.running:
        flash("Scheduler not running. Cannot schedule builds.", "danger")
        return redirect(url_for('main.index'))

    logger.info(f"Received {len(buildable_items_from_form)} buildable items to schedule for "
                f"batch '{batch_review_id}' with option: '{schedule_option}'.")
    
    from apscheduler.triggers.date import DateTrigger
    scheduler_tz_obj = scheduler.timezone
    
    base_run_time_utc: datetime.datetime = datetime.datetime.now(pytz.utc)
    if schedule_option != 'now' and schedule_start_time_str:
        try:
            naive_dt_form = datetime.datetime.fromisoformat(schedule_start_time_str)
            localized_dt = scheduler_tz_obj.localize(naive_dt_form, is_dst=None) if hasattr(scheduler_tz_obj, 'localize') else naive_dt_form.replace(tzinfo=scheduler_tz_obj)
            base_run_time_utc = localized_dt.astimezone(pytz.utc)
        except ValueError as e_date_sched:
            flash(f"Invalid start time format provided. Scheduling jobs for 'now'. Error: {e_date_sched}", "warning")
            base_run_time_utc = datetime.datetime.now(pytz.utc)

    current_operation_block_start_time_utc = base_run_time_utc
    stagger_delta = datetime.timedelta(minutes=stagger_minutes)
    setup_delay_after_teardown = datetime.timedelta(minutes=2)
    
    scheduled_ops_count = 0
    failed_ops_count = 0
    scheduled_op_details_for_interim: List[Dict[str, Any]] = []

    for item_idx, item_to_build in enumerate(buildable_items_from_form):
        item_sf_code = item_to_build.get('original_sf_course_code', item_to_build.get('sf_course_code', f'UNKNOWN_SF_{item_idx}'))
        item_vendor = item_to_build.get('vendor')
        item_lb_course = item_to_build.get('labbuild_course')

        if not item_vendor or not item_lb_course:
            failed_ops_count += 1
            continue
            
        assignments = item_to_build.get("assignments", [])
        if not assignments:
            failed_ops_count += 1
            continue

        for assignment_idx, assignment_block in enumerate(assignments):
            host = assignment_block.get('host')
            start_pod = assignment_block.get('start_pod')
            end_pod = assignment_block.get('end_pod')

            if not host or start_pod is None or end_pod is None:
                failed_ops_count +=1
                continue

            if schedule_option == 'staggered' and (item_idx > 0 or assignment_idx > 0):
                current_operation_block_start_time_utc += stagger_delta
            elif schedule_option == 'now':
                current_operation_block_start_time_utc = datetime.datetime.now(pytz.utc)
            
            is_student_build = item_to_build.get('type') == 'Student Build'
            tag = f"{item_sf_code.replace('/', '_').replace(' ', '_')}"
            if not is_student_build:
                tag += "_TP"
            
            base_args = ['-v', item_vendor, '-g', item_lb_course, '--host', host, '-s', str(start_pod), '-e', str(end_pod), '-t', tag[:45]]
            
            if item_vendor.lower() == 'f5' and item_to_build.get('f5_class_number'):
                 base_args.extend(['-cn', str(item_to_build['f5_class_number'])])

            teardown_job_id = None
            actual_teardown_run_time_utc = current_operation_block_start_time_utc
            if perform_teardown_first:
                args_list_teardown = ['teardown'] + base_args
                job_name_td = f"Teardown_{tag}"
                try:
                    trigger_td = DateTrigger(run_date=actual_teardown_run_time_utc, timezone=pytz.utc)
                    job_td = scheduler.add_job(run_labbuild_task, trigger=trigger_td, args=[args_list_teardown], name=job_name_td, misfire_grace_time=1800, replace_existing=False)
                    teardown_job_id = job_td.id
                    scheduled_op_details_for_interim.append({"operation": "teardown", "job_id": job_td.id, "name": job_name_td, "run_time_utc": actual_teardown_run_time_utc.isoformat(), "args": args_list_teardown})
                    scheduled_ops_count += 1
                except Exception as e_sched_td:
                    logger.error(f"Failed to schedule TEARDOWN job {job_name_td}: {e_sched_td}", exc_info=True)
                    failed_ops_count += 1
            
            args_list_setup = ['setup'] + base_args
            
            if item_to_build.get('start_date'): args_list_setup.extend(['--start-date', item_to_build['start_date']])
            if item_to_build.get('end_date'): args_list_setup.extend(['--end-date', item_to_build['end_date']])
            if item_to_build.get('sf_trainer_name'): args_list_setup.extend(['--trainer-name', item_to_build['sf_trainer_name']])
            
            apm_user = item_to_build.get('apm_username') if is_student_build else item_to_build.get('trainer_apm_username')
            apm_pass = item_to_build.get('apm_password') if is_student_build else item_to_build.get('trainer_apm_password')

            if apm_user: args_list_setup.extend(['--username', apm_user])
            if apm_pass: args_list_setup.extend(['--password', apm_pass])
            
            job_name_setup = f"Setup_{tag}"
            actual_setup_run_time_utc = actual_teardown_run_time_utc + setup_delay_after_teardown if perform_teardown_first else current_operation_block_start_time_utc

            try:
                trigger_setup = DateTrigger(run_date=actual_setup_run_time_utc, timezone=pytz.utc)
                job_setup = scheduler.add_job(run_labbuild_task, trigger=trigger_setup, args=[args_list_setup], name=job_name_setup, misfire_grace_time=7200, replace_existing=False)
                scheduled_op_details_for_interim.append({"operation": "setup", "job_id": job_setup.id, "name": job_name_setup, "run_time_utc": actual_setup_run_time_utc.isoformat(), "args": args_list_setup, "depends_on_job_id": teardown_job_id})
                scheduled_ops_count += 1
            except Exception as e_sched_setup:
                logger.error(f"Failed to schedule SETUP job {job_name_setup}: {e_sched_setup}", exc_info=True)
                failed_ops_count += 1

    final_batch_status = "builds_scheduled_with_errors"
    if scheduled_ops_count > 0 and failed_ops_count == 0:
        final_batch_status = "builds_fully_scheduled"
    elif scheduled_ops_count == 0 and failed_ops_count > 0:
        final_batch_status = "build_scheduling_all_failed"
    elif scheduled_ops_count == 0 and failed_ops_count == 0:
         final_batch_status = "no_builds_to_schedule"
    
    if interim_alloc_collection and batch_review_id:
        try:
            update_filter = {"batch_review_id": batch_review_id, "status": {"$in": ["student_confirmed", "trainer_confirmed", "trainer_disabled_by_rule", "trainer_skipped_by_user"]}}
            update_doc = {"$set": {"status": final_batch_status, "scheduled_operations": scheduled_op_details_for_interim, "scheduled_at_utc": datetime.datetime.now(pytz.utc)}}
            interim_alloc_collection.update_many(update_filter, update_doc)
        except PyMongoError as e_upd_interim:
            logger.error(f"Error updating final interim status for batch '{batch_review_id}': {e_upd_interim}", exc_info=True)

    flash(f"Attempted to schedule operations. Scheduled: {scheduled_ops_count}. Failures: {failed_ops_count}.", 'success' if failed_ops_count == 0 and scheduled_ops_count > 0 else 'warning' if scheduled_ops_count > 0 else 'danger')
    return redirect(url_for('main.index'))

def _dispatch_bulk_labbuild_tasks(
        tasks_to_run: List[Dict[str, Any]], # Each dict: {'args': list, 'description': str, 'db_update_info': tuple (optional)}
        action_name: str,
        operation_if_power_toggle: Optional[str] = None):
    """
    Helper to run a list of labbuild tasks sequentially in a background thread.
    Optionally updates power state in DB after each power toggle task.
    """
    if not tasks_to_run:
        logger.info(f"No tasks to dispatch for bulk {action_name}.")
        return

    def worker():
        total_tasks = len(tasks_to_run)
        logger.info(f"Starting bulk {action_name} worker for {total_tasks} tasks.")
        success_count = 0
        fail_count = 0

        for i, task_info in enumerate(tasks_to_run):
            args = task_info['args']
            description = task_info['description']
            db_update_info = task_info.get('db_update_info') # For power toggle

            logger.info(f"Executing task {i+1}/{total_tasks} for bulk {action_name}: {description} - CMD: {' '.join(args)}")
            try:
                # run_labbuild_task itself handles subprocess.run and logging
                # We assume it's synchronous for this sequential helper,
                # or we'd need a more complex Future-based approach here too.
                # For now, assuming run_labbuild_task blocks until completion.
                run_labbuild_task(args) # This runs the labbuild.py script
                
                # If it's a power toggle and task appeared successful (no exception from run_labbuild_task),
                # update the DB. A more robust way would be for run_labbuild_task to return success/failure.
                if action_name == "power_toggle" and db_update_info and operation_if_power_toggle:
                    tag, course, host, item_num, item_type, class_num_ctx = db_update_info
                    new_power_state_bool = (operation_if_power_toggle == 'start')
                    update_power_state_in_db(tag, course, host, item_num, item_type, new_power_state_bool, class_num_ctx)
                
                success_count += 1
            except Exception as e:
                logger.error(f"Error in bulk {action_name} task {description}: {e}", exc_info=True)
                fail_count += 1
            # Add a small delay between tasks if desired, e.g., time.sleep(1)
        
        logger.info(f"Bulk {action_name} worker finished. Success: {success_count}, Failed: {fail_count}.")
        # Flash messages from a background thread is tricky.
        # Usually, you'd use SSE or another mechanism to update UI.
        # For now, main route will flash a general submission message.

    thread = threading.Thread(target=worker, daemon=True)
    thread.start()
    flash(f"Submitted {len(tasks_to_run)} tasks for bulk {action_name}. Check logs for progress.", "info")


@bp.route('/bulk-toggle-power', methods=['POST'])
def bulk_toggle_power():
    selected_items_json = request.form.get('selected_items_json')
    operation = request.args.get('operation') # 'start' or 'stop'

    if not selected_items_json or not operation or operation not in ['start', 'stop']:
        flash("Invalid bulk power toggle request.", "danger")
        return redirect(url_for('main.view_allocations', **request.form)) # Pass back filters

    try:
        selected_items = json.loads(selected_items_json)
        if not isinstance(selected_items, list):
            raise ValueError("Selected items data is not a list.")
    except (json.JSONDecodeError, ValueError) as e:
        flash(f"Error processing selected items: {e}", "danger")
        return redirect(url_for('main.view_allocations', **request.form))

    tasks_to_run = []
    for item in selected_items:
        # Validate item structure (basic)
        if not all(k in item for k in ['vendor', 'course', 'host', 'item_type', 'item_number', 'tag']):
            logger.warning(f"Skipping invalid item in bulk power toggle: {item}")
            continue

        args = [
            'manage',
            '-v', item['vendor'],
            '-g', item['course'],
            '--host', item['host'],
            '-t', item['tag'],
            '-o', operation
        ]
        item_num_str = str(item['item_number'])
        if item['item_type'] == 'f5_class':
            args.extend(['-cn', item_num_str])
        elif item['item_type'] == 'pod':
            args.extend(['-s', item_num_str, '-e', item_num_str])
            if item.get('vendor', '').lower() == 'f5' and item.get('class_number'):
                args.extend(['-cn', str(item['class_number'])])
        else:
            logger.warning(f"Unknown item_type '{item['item_type']}' for bulk power toggle. Skipping item: {item.get('tag')}/{item.get('course')}/{item_num_str}")
            continue
        
        description = f"Power {operation} for {item['item_type']} {item_num_str} (Tag: {item['tag']}, Course: {item['course']})"
        db_update_info = (
            item['tag'], item['course'], item['host'], item['item_number'],
            item['item_type'], item.get('class_number') 
        )
        tasks_to_run.append({'args': args, 'description': description, 'db_update_info': db_update_info})

    if tasks_to_run:
        _dispatch_bulk_labbuild_tasks(tasks_to_run, "power_toggle", operation_if_power_toggle=operation)
    else:
        flash("No valid items found to process for bulk power toggle.", "info")
    
    # Preserve filters from the form submission for redirection
    preserved_filters = {k: v for k, v in request.form.items() if k.startswith('filter_') or k in ['page', 'per_page']}
    return redirect(url_for('main.view_allocations', **preserved_filters))


@bp.route('/bulk-teardown-items', methods=['POST'])
def bulk_teardown_items():
    selected_items_json = request.form.get('selected_items_json')
    if not selected_items_json:
        flash("No items selected for bulk teardown.", "warning")
        return redirect(url_for('main.view_allocations', **request.form))

    try:
        selected_items = json.loads(selected_items_json)
        if not isinstance(selected_items, list): raise ValueError
    except:
        flash("Error processing selected items for teardown.", "danger")
        return redirect(url_for('main.view_allocations', **request.form))

    # --- Grouping and Range Merging Logic ---
    tasks_to_run = []
    # Group items by a tuple key: (vendor, course, host, tag, class_number_for_f5)
    grouped_items = defaultdict(lambda: {'pods': set(), 'f5_classes': set()})

    for item in selected_items:
        try:
            if not all(k in item for k in ['vendor', 'course', 'host', 'item_type', 'item_number', 'tag']):
                logger.warning(f"Skipping invalid item in bulk teardown: {item}")
                continue
            
            # **FIX**: Ensure item_number is converted to an integer immediately.
            item_number_int = int(item['item_number'])

            class_context = int(item.get('class_number')) if item.get('vendor', '').lower() == 'f5' and item.get('class_number') else None
            group_key = (item['vendor'], item['course'], item['host'], item['tag'], class_context)

            if item['item_type'] == 'f5_class':
                grouped_items[group_key]['f5_classes'].add(item_number_int)
            elif item['item_type'] == 'pod':
                grouped_items[group_key]['pods'].add(item_number_int)
        except (ValueError, TypeError) as e:
             logger.warning(f"Skipping item due to data conversion error: {e}. Item: {item}")
             continue
    
    logger.info(f"Bulk Teardown: Grouped selected items into {len(grouped_items)} potential jobs.")

    # Process each group to create labbuild commands
    for group_key, items in grouped_items.items():
        vendor, course, host, tag, class_num = group_key

        # 1. Process F5 Classes (each is a separate job)
        for f5_class_to_tear in sorted(list(items['f5_classes'])):
            args = ['teardown', '-v', vendor, '-g', course, '--host', host, '-t', tag, '-cn', str(f5_class_to_tear)]
            desc = f"Teardown F5 Class {f5_class_to_tear} (Tag: {tag})"
            tasks_to_run.append({'args': args, 'description': desc})

        # 2. Process Pods (merge contiguous ranges)
        if items['pods']:
            sorted_pods = sorted(list(items['pods']))
            
            if not sorted_pods: continue

            # **FIX**: Ensure start_range and end_range are always integers.
            start_range = sorted_pods[0]
            end_range = sorted_pods[0]

            for i in range(1, len(sorted_pods)):
                # Now this comparison is safe (int vs int)
                if sorted_pods[i] != end_range + 1:
                    # Create job for the completed range
                    args = ['teardown', '-v', vendor, '-g', course, '--host', host, '-t', tag, '-s', str(start_range), '-e', str(end_range)]
                    if class_num is not None: args.extend(['-cn', str(class_num)])
                    desc = f"Teardown Pods {start_range}-{end_range} (Tag: {tag})"
                    tasks_to_run.append({'args': args, 'description': desc})
                    
                    # Start a new range (with an integer)
                    start_range = sorted_pods[i]
                    end_range = sorted_pods[i]
                else:
                    # Extend the current range
                    end_range = sorted_pods[i]
            
            # Add the last range after the loop finishes
            args = ['teardown', '-v', vendor, '-g', course, '--host', host, '-t', tag, '-s', str(start_range), '-e', str(end_range)]
            if class_num is not None: args.extend(['-cn', str(class_num)])
            desc = f"Teardown Pods {start_range}-{end_range} (Tag: {tag})"
            tasks_to_run.append({'args': args, 'description': desc})

    # --- Dispatch the generated tasks ---
    if tasks_to_run:
        _dispatch_bulk_labbuild_tasks(tasks_to_run, "infrastructure_teardown")
    else:
        flash("No valid items found to process for bulk infrastructure teardown.", "info")

    preserved_filters = {k: v for k, v in request.form.items() if k.startswith('filter_') or k in ['page', 'per_page']}
    return redirect(url_for('main.view_allocations', **preserved_filters))


@bp.route('/bulk-db-delete-items', methods=['POST'])
def bulk_db_delete_items():
    selected_items_json = request.form.get('selected_items_json')
    if not selected_items_json:
        flash("No items selected for bulk DB deletion.", "warning")
        return redirect(url_for('main.view_allocations', **request.form))

    try:
        selected_items = json.loads(selected_items_json)
        if not isinstance(selected_items, list): raise ValueError
    except:
        flash("Error processing selected items for DB deletion.", "danger")
        return redirect(url_for('main.view_allocations', **request.form))

    deleted_count = 0
    failed_count = 0
    processed_tags_courses = set() # To avoid redundant messages for same course in a tag

    for item in selected_items:
        try:
            tag = item.get('tag')
            course_name = item.get('course')
            item_type = item.get('item_type')
            # item_number holds pod_number for type 'pod' or class_number for type 'f5_class'
            item_number = int(item['item_number']) if item.get('item_number') else None
            
            # For F5 pods, class_number context is in item['class_number']
            # For F5 class itself, item_number is class_number and item['class_number'] is also class_number
            class_number_context = int(item['class_number']) if item.get('class_number') else None

            pod_to_delete = None
            class_to_delete = None

            if item_type == 'f5_class':
                class_to_delete = item_number
            elif item_type == 'pod':
                pod_to_delete = item_number
                if item.get('vendor', '').lower() == 'f5': # F5 pod needs class context
                    class_to_delete = class_number_context 
                    if class_to_delete is None: # Should not happen if data is consistent
                        logger.warning(f"F5 pod {pod_to_delete} missing class context for DB delete. Item: {item}")
                        # This might delete all pods with this number if class_to_delete is None in delete_from_database
                        # For safety, skip if F5 pod and class context is missing from the selected item data
                        failed_count +=1
                        continue
            
            if not tag or not course_name:
                logger.warning(f"Skipping item for DB delete due to missing tag/course: {item}")
                failed_count += 1
                continue

            # Call delete_from_database for each distinct item
            # delete_from_database handles if pod_number or class_number is None
            if delete_from_database(tag, course_name=course_name, pod_number=pod_to_delete, class_number=class_to_delete):
                deleted_count += 1
                log_msg_key = (tag, course_name, pod_to_delete, class_to_delete)
                if log_msg_key not in processed_tags_courses:
                    logger.info(f"DB entry deleted for: Tag {tag}, Course {course_name}, Pod {pod_to_delete}, Class {class_to_delete}")
                    processed_tags_courses.add(log_msg_key)
            else:
                failed_count += 1
                logger.error(f"Failed DB delete for: Tag {tag}, Course {course_name}, Pod {pod_to_delete}, Class {class_to_delete}")

        except Exception as e:
            failed_count += 1
            logger.error(f"Error during bulk DB delete for item {item}: {e}", exc_info=True)

    if deleted_count > 0:
        flash(f"Successfully deleted {deleted_count} DB entries.", "success")
    if failed_count > 0:
        flash(f"Failed to delete {failed_count} DB entries. Check logs.", "danger")
    if deleted_count == 0 and failed_count == 0:
        flash("No valid items found to process for DB deletion.", "info")
        
    preserved_filters = {k: v for k, v in request.form.items() if k.startswith('filter_') or k in ['page', 'per_page']}
    return redirect(url_for('main.view_allocations', **preserved_filters))

def set_cell_style(cell, bold: bool = False, italic: bool = False, underline: str = None,
                   strike: bool = False, font_name: str = 'Calibri', font_size: int = 11,
                   font_color: str = None,  # Hex ARGB color string e.g., "FF00FF00" for green
                   alignment_horizontal: str = None, # e.g., 'general', 'left', 'right', 'center', 'justify'
                   alignment_vertical: str = None,   # e.g., 'top', 'center', 'bottom', 'justify', 'distributed'
                   wrap_text: bool = None,
                   shrink_to_fit: bool = None,
                   indent: int = 0,
                   text_rotation: int = 0,
                   border_style: str = None,    # e.g., 'thin', 'medium', 'thick', 'double', 'dotted', 'dashed'
                   border_color: str = "000000", # Default black
                   fill_color: str = None,      # Hex ARGB color string e.g., "FFFFFF00" for yellow
                   number_format: str = None    # e.g., 'General', '0.00', 'mm-dd-yy'
                   ):
    """
    Applies comprehensive styling to an openpyxl cell.
    Creates new style objects as openpyxl styles are immutable.
    """
    # --- Font ---
    # Start with defaults or existing font properties
    current_font = cell.font
    new_font_kwargs = {
        'name': current_font.name if current_font.name is not None else font_name,
        'sz': current_font.sz if current_font.sz is not None else font_size,
        'bold': current_font.bold if bold is None else bold,
        'italic': current_font.italic if italic is None else italic,
        'underline': current_font.underline if underline is None else underline,
        'strike': current_font.strike if strike is None else strike,
        'color': current_font.color if font_color is None else (Color(font_color) if isinstance(font_color, str) else font_color)
        # Copy other font properties if needed like scheme, family, charset, etc.
        # For simplicity, we'll stick to common ones here.
    }
    # Only create a new Font object if there's a reason to (a change or explicit setting)
    # This check can be more granular if performance is critical for millions of cells.
    cell.font = Font(**new_font_kwargs)

    # --- Alignment ---
    if (alignment_horizontal is not None or
            alignment_vertical is not None or
            wrap_text is not None or
            shrink_to_fit is not None or
            indent != 0 or # Default indent is 0, so any non-zero is a change
            text_rotation != 0): # Default rotation is 0

        existing_alignment = cell.alignment
        new_align_kwargs = {}

        # Preserve existing alignment properties if not overridden
        if existing_alignment:
            new_align_kwargs.update({
                'horizontal': existing_alignment.horizontal,
                'vertical': existing_alignment.vertical,
                'textRotation': existing_alignment.textRotation, # openpyxl uses textRotation (int)
                'wrapText': existing_alignment.wrapText,
                'shrinkToFit': existing_alignment.shrinkToFit,
                'indent': existing_alignment.indent,
                'relativeIndent': existing_alignment.relativeIndent,
                'justifyLastLine': existing_alignment.justifyLastLine,
                'readingOrder': existing_alignment.readingOrder,
            })

        # Apply new/overridden values
        if alignment_horizontal is not None:
            new_align_kwargs['horizontal'] = alignment_horizontal
        if alignment_vertical is not None:
            new_align_kwargs['vertical'] = alignment_vertical
        if wrap_text is not None:
            new_align_kwargs['wrapText'] = wrap_text
        if shrink_to_fit is not None:
            new_align_kwargs['shrinkToFit'] = shrink_to_fit
        if indent != 0: # Only set if different from default
            new_align_kwargs['indent'] = indent
        if text_rotation != 0: # Only set if different from default
            new_align_kwargs['textRotation'] = text_rotation
        
        cell.alignment = Alignment(**new_align_kwargs)

    # --- Border ---
    if border_style:
        side = Side(border_style=border_style, color=border_color)
        cell.border = Border(left=side, right=side, top=side, bottom=side)

    # --- Fill ---
    if fill_color:
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

    # --- Number Format ---
    if number_format:
        cell.number_format = number_format

def get_host_shortname(hostname_fqdn_or_short: str) -> str:
    """Returns a standardized short host name (ni, cl, ul, un, ho, k2) or original if not matched."""
    if not hostname_fqdn_or_short:
        return "N/A"
    
    hn_lower = hostname_fqdn_or_short.lower()
    
    if "nightbird" in hn_lower or hn_lower == "ni": return "Ni"
    if "cliffjumper" in hn_lower or hn_lower == "cl": return "Cl"
    if "ultramagnus" in hn_lower or hn_lower == "ul": return "Ul"
    if "unicron" in hn_lower or hn_lower == "un": return "Un"
    if "hotshot" in hn_lower or hn_lower == "ho": return "Ho"
    if "k2" in hn_lower : return "K2" # Assuming K2 is another short name

    # Fallback to first two letters if it's an FQDN and not recognized
    if '.' in hn_lower:
        short = hn_lower.split('.')[0][:2]
        if short == "ni": return "Ni"
        # Add other FQDN prefix mappings if needed
    
    return hostname_fqdn_or_short # Return original if no specific mapping

@bp.route('/export-allocations-excel')
def export_allocations_excel():
    logger.info("Exporting allocations to Excel (New Format)...")
    if alloc_collection is None or course_config_collection is None or host_collection is None:
        flash("Database collections unavailable for export.", "danger")
        return redirect(url_for('main.view_allocations'))

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Allocations"

        # --- 1. RAM Summary Header (as per image) ---
        ram_summary_data_from_image = [
            ["RAM Summary", None, None, None], # Merged later
            [None, "Allocated RAM", "CPU", "HDD"],
            ["Nightbird", 2000, 2000, 2000],
            ["Hotshot", 2000, 2000, 2000],
            ["Cliffjumper", 1200, 1200, 1200],
            ["Unicron", 1000, 1000, 1000],
            ["Ultramagnus", 1000, 1000, 1000],
            ["K2", 1000, 1000, 1000] # Added K2 based on common hosts
        ]
        yellow_fill = "FFFF00"
        for r_idx, r_data in enumerate(ram_summary_data_from_image, start=1):
            for c_idx, value in enumerate(r_data, start=1):
                if value is not None:
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    is_header_row = r_idx <= 2
                    set_cell_style(cell, bold=is_header_row, fill_color=yellow_fill, alignment_horizontal="center" if is_header_row or c_idx > 1 else "left")
        ws.merge_cells('A1:D1')
        current_row = len(ram_summary_data_from_image) + 2 # Start main table after some space

        # --- Main Table Headers (from image) ---
        main_headers = [
            "Course Job Code", "Start Date", "Last Day", "Location", "Trainer Name", "Course Name",
            "Start/End Pod", None, "Username", "Password", "Class", "Students", "Vendor Pods", "Group",
            "Version", "Course Version", "RAM", "Virtual Hosts", "Ni", "Cl", "Ul", "Un", "Ho", "K2" # Added K2 here too
        ] # Col H is for the "->" arrow

        header_fill_color = "ADD8E6" # Light Blue
        thin_black_border = Border(left=Side(style='thin', color="000000"), 
                                 right=Side(style='thin', color="000000"), 
                                 top=Side(style='thin', color="000000"), 
                                 bottom=Side(style='thin', color="000000"))

        # --- Data Fetching and Processing ---
        all_alloc_docs = list(alloc_collection.find({}))
        all_course_configs_list = list(course_config_collection.find({}))
        course_configs_dict = {cc['course_name']: cc for cc in all_course_configs_list}
        
        # Get host locations mapping if available
        host_locations = {h.get('host_name'): h.get('location', 'Virtual') for h in host_collection.find({}, {"host_name":1, "location":1})}


        processed_data_by_vendor = defaultdict(list)

        for tag_doc in all_alloc_docs:
            tag_name = tag_doc.get("tag", "UNKNOWN_TAG")
            for course_alloc in tag_doc.get("courses", []):
                vendor_code = course_alloc.get("vendor", "UN").upper()
                labbuild_course_name = course_alloc.get("course_name")
                course_config = course_configs_dict.get(labbuild_course_name, {})

                for pod_detail in course_alloc.get("pod_details", []):
                    is_trainer_pod = "-TP" in tag_name.upper()
                    excel_row_data = {
                        "Course Job Code": tag_name if not is_trainer_pod else tag_name.replace("-TP",""), # Base code for TP
                        "Start Date": course_alloc.get("start_date", "N/A") if not is_trainer_pod else "",
                        "Last Day": course_alloc.get("end_date", "N/A") if not is_trainer_pod else "", # Assuming end_date is last day
                        "Location": host_locations.get(pod_detail.get("host"), "Virtual"),
                        "Trainer Name": course_alloc.get("trainer_name", "N/A") if not is_trainer_pod else "",
                        "Course Name": labbuild_course_name if not is_trainer_pod else "Trainer pods",
                        "Username": course_config.get("student_user", "labcp-X/labpa-X") if not is_trainer_pod else \
                                    (course_config.get("trainer_user", "labcp-X") if "-TP" in tag_name.upper() else ""),
                        "Password": course_config.get("student_password", "password") if not is_trainer_pod else \
                                    (course_config.get("trainer_password", "password") if "-TP" in tag_name.upper() else ""),
                        "Class": "", "Students": "", "Vendor Pods": "", # To be filled
                        "Group": course_config.get("group", ""), # From courseconfig
                        "Version": labbuild_course_name, # Main LabBuild course name
                        "Course Version": course_config.get("course_version", course_config.get("build#", "")), # Specific build/version
                        "RAM": 0, # To be calculated
                        "Virtual Hosts": "", # Representative VM
                        "Ni": 0, "Cl": 0, "Ul": 0, "Un": 0, "Ho": 0, "K2": 0, # Host RAM allocations
                        "_host_shortname": get_host_shortname(pod_detail.get("host")), # Internal helper
                        "_original_host_fqdn": pod_detail.get("host")
                    }

                    # Determine representative Virtual Host (e.g., first component)
                    components_list = course_config.get("components", [])
                    if not components_list and course_config.get("groups"): # F5 structure
                        for grp in course_config.get("groups",[]): components_list.extend(grp.get("component",[]))
                    
                    if components_list:
                        first_comp = components_list[0]
                        base_vm_name = first_comp.get("base_vm", first_comp.get("component_name", "N/A"))
                        excel_row_data["Virtual Hosts"] = base_vm_name


                    # Pods, Class, Students calculation
                    pod_num = pod_detail.get("pod_number")
                    class_num = pod_detail.get("class_number")
                    num_student_pods_in_range = 0

                    if vendor_code == "F5" and class_num is not None:
                        excel_row_data["Class"] = class_num
                        f5_student_pods = pod_detail.get("pods", [])
                        if isinstance(f5_student_pods, list):
                            excel_row_data["Students"] = len(f5_student_pods)
                            excel_row_data["Vendor Pods"] = ", ".join(sorted([str(p.get("pod_number")) for p in f5_student_pods if p.get("pod_number") is not None]))
                            num_student_pods_in_range = len(f5_student_pods)
                        # For F5, "Start/End Pod" in Excel might refer to the class number itself if no student pods, or range of student pods
                        excel_row_data["Start/End Pod"] = str(class_num) if not f5_student_pods else f"{min(p['pod_number'] for p in f5_student_pods)}-{max(p['pod_number'] for p in f5_student_pods)}"

                    elif pod_num is not None: # Non-F5 or F5 student pod without class context (unlikely from your data)
                        # The image for PA/CP shows Start/End Pod as a range.
                        # This pod_detail from currentallocation is often a single pod_number from build.
                        # We need to infer the original range if this is a student pod.
                        # For simplicity here, if it's a single pod, Start/End will be the same.
                        # Your "-TP" logic might override this.
                        excel_row_data["Start/End Pod"] = f"{pod_num} -> {pod_num}" # Default if range not obvious from this single entry
                        excel_row_data["Students"] = 1 # Assume 1 student per basic pod_detail entry unless grouped
                        excel_row_data["Vendor Pods"] = str(pod_num)
                        num_student_pods_in_range = 1
                    
                    # Override Start/End Pod for Trainer Pods (often a different range)
                    if is_trainer_pod:
                        # Trainer pods often have a specific range in their tag or pre-defined
                        # This part needs logic to correctly parse trainer pod ranges from tag or config
                        # Example: If tag "COURSE-TP-101-102"
                        tp_range_match = re.search(r'-(\d+)-(\d+)$', tag_name)
                        if tp_range_match:
                            excel_row_data["Start/End Pod"] = f"{tp_range_match.group(1)} -> {tp_range_match.group(2)}"
                            excel_row_data["Students"] = int(tp_range_match.group(2)) - int(tp_range_match.group(1)) + 1
                        else: # Fallback if TP range not in tag
                            excel_row_data["Start/End Pod"] = f"{pod_num} -> {pod_num}" if pod_num else "N/A"
                            excel_row_data["Students"] = 1 if pod_num else 0
                        excel_row_data["Username"] = course_config.get("trainer_user", excel_row_data["Username"])
                        excel_row_data["Password"] = course_config.get("trainer_password", excel_row_data["Password"])


                    # Calculate RAM for this row based on number of "student" pods and host
                    # This is a complex part based on your image.
                    # The RAM per host (Ni, Cl, etc.) is shown per *course line item*.
                    # It's the total RAM for the pods in THAT line item on THAT host.
                    # If a course like CP Maestro is split (e.g. 2 pods on Unicron, 1 on Hotshot),
                    # it appears as multiple lines in your Excel for the same "Course Job Code".
                    
                    ram_per_pod_from_config = float(course_config.get("memory_gb_per_pod", course_config.get("memory", 0)))
                    if isinstance(ram_per_pod_from_config, str) : # handle if memory is string
                        try: ram_per_pod_from_config = float(ram_per_pod_from_config)
                        except ValueError: ram_per_pod_from_config = 0.0
                    
                    calculated_ram_for_row = ram_per_pod_from_config * num_student_pods_in_range if num_student_pods_in_range > 0 else ram_per_pod_from_config # if 0 students, maybe it's a single entity ram
                    if is_trainer_pod: # Trainer pods might have different RAM or count as 1 entity
                        calculated_ram_for_row = ram_per_pod_from_config # Assuming trainer pod RAM is per entity
                        
                    excel_row_data["RAM"] = calculated_ram_for_row
                    
                    host_short_key = excel_row_data["_host_shortname"]
                    if host_short_key in ["Ni", "Cl", "Ul", "Un", "Ho", "K2"]:
                        excel_row_data[host_short_key] = calculated_ram_for_row
                    
                    processed_data_by_vendor[vendor_code].append(excel_row_data)

        # --- Write to Excel ---
        # Define vendor order for sheet
        vendor_display_order = ["PA", "CP", "NU", "F5", "AV", "PR", "OT"] # Add others as needed

        for vendor_code_ordered in vendor_display_order:
            if vendor_code_ordered not in processed_data_by_vendor:
                continue
            
            vendor_specific_data = processed_data_by_vendor[vendor_code_ordered]
            if not vendor_specific_data:
                continue

            # Vendor Section Title (e.g., "PA Courses")
            title_cell = ws.cell(row=current_row, column=1, value=f"{vendor_code_ordered} Courses")
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(main_headers))
            set_cell_style(title_cell, bold=True, fill_color=header_fill_color, alignment_horizontal="left", font_size=12)
            current_row += 1

            # Write Main Headers for this vendor section
            for col_idx, header_text in enumerate(main_headers, start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=header_text if header_text is not None else "")
                set_cell_style(cell, bold=True, fill_color=header_fill_color, border_style="thin", alignment_horizontal="center", wrap_text=True, font_size=10)
            current_row += 1
            
            vendor_total_pods_count = 0

            for item_data in vendor_specific_data:
                col_offset = 0
                for header_key in main_headers: # Iterate based on header order
                    if header_key is None: # For the "->" arrow column
                        ws.cell(row=current_row, column=col_offset + 1, value="->")
                        set_cell_style(ws.cell(row=current_row, column=col_offset + 1), border_style="thin", alignment_horizontal="center")
                    elif header_key == "Start/End Pod": # Special handling for combined field
                         cell_val = item_data.get(header_key, "")
                         ws.cell(row=current_row, column=col_offset + 1, value=str(cell_val).split(" -> ")[0] if " -> " in str(cell_val) else str(cell_val) ) # Start
                         set_cell_style(ws.cell(row=current_row, column=col_offset + 1), border_style="thin", alignment_horizontal="center")
                         # Arrow in next column (which was None in headers)
                         ws.cell(row=current_row, column=col_offset + 2, value="->")
                         set_cell_style(ws.cell(row=current_row, column=col_offset + 2), border_style="thin", alignment_horizontal="center")
                         # End pod value in column after arrow
                         ws.cell(row=current_row, column=col_offset + 3, value=str(cell_val).split(" -> ")[1] if " -> " in str(cell_val) else "")
                         set_cell_style(ws.cell(row=current_row, column=col_offset + 3), border_style="thin", alignment_horizontal="center")
                         col_offset += 1 # Account for the extra column used by "->" and end pod
                    else:
                        cell_val = item_data.get(header_key, "")
                        ws.cell(row=current_row, column=col_offset + 1, value=cell_val)
                        align_h = "center" if header_key in ["Class", "Students", "Vendor Pods", "RAM", "Ni","Cl","Ul","Un","Ho","K2"] else "left"
                        set_cell_style(ws.cell(row=current_row, column=col_offset + 1), border_style="thin", alignment_horizontal=align_h, alignment_vertical="top", wrap_text=True)
                    col_offset += 1
                current_row += 1
                
                # Accumulate pods for vendor total (using "Students" count for now, needs refinement)
                try: vendor_total_pods_count += int(item_data.get("Students", 0))
                except ValueError: pass


            # Total Pods for Vendor
            if vendor_specific_data: # Only add total if there was data
                total_pods_label_cell = ws.cell(row=current_row, column=main_headers.index("Students") + 1 -1, value="Total Pods") # Col before "Students"
                total_pods_value_cell = ws.cell(row=current_row, column=main_headers.index("Students") + 1, value=vendor_total_pods_count)
                
                set_cell_style(total_pods_label_cell, bold=True, fill_color=header_fill_color, border_style="thin", alignment_horizontal="right")
                set_cell_style(total_pods_value_cell, bold=True, fill_color=header_fill_color, border_style="thin", alignment_horizontal="center")
                
                # Merge cells for "Total Pods" label across preceding columns
                label_col_start_merge = 1
                label_col_end_merge = main_headers.index("Students") + 1 - 2 # Up to column before label
                if label_col_end_merge >= label_col_start_merge:
                     ws.merge_cells(start_row=current_row, start_column=label_col_start_merge, end_row=current_row, end_column=label_col_end_merge)
                     for c_merge in range(label_col_start_merge, label_col_end_merge + 1):
                        set_cell_style(ws.cell(row=current_row, column=c_merge), fill_color=header_fill_color, border_style="thin")
                
                # Style remaining cells in total row
                ram_col_idx = main_headers.index("RAM") +1
                host_ram_cols_start_idx = main_headers.index("Ni") + 1
                
                # Sum RAM for hosts in this vendor section
                host_ram_sums = defaultdict(float)
                for item_d in vendor_specific_data:
                    for host_s_key in ["Ni", "Cl", "Ul", "Un", "Ho", "K2"]:
                        try: host_ram_sums[host_s_key] += float(item_d.get(host_s_key, 0))
                        except ValueError: pass
                
                # Write summed host RAM
                ws.cell(row=current_row, column=ram_col_idx, value=sum(host_ram_sums.values())) # Total RAM column
                set_cell_style(ws.cell(row=current_row, column=ram_col_idx), bold=True, fill_color=header_fill_color, border_style="thin", alignment_horizontal="center")


                for i_host_key, host_s_key_val in enumerate(["Ni", "Cl", "Ul", "Un", "Ho", "K2"]):
                    cell = ws.cell(row=current_row, column=host_ram_cols_start_idx + i_host_key, value=host_ram_sums[host_s_key_val])
                    set_cell_style(cell, bold=True, fill_color=header_fill_color, border_style="thin", alignment_horizontal="center")

                # Style any remaining cells in the total row (e.g. Virtual Hosts)
                for c_empty_total in range(main_headers.index("Vendor Pods") + 2, ram_col_idx): # Cols between Vendor Pods and RAM
                    set_cell_style(ws.cell(row=current_row, column=c_empty_total), fill_color=header_fill_color, border_style="thin")
                for c_empty_total_after in range(host_ram_cols_start_idx + len(["Ni", "Cl", "Ul", "Un", "Ho", "K2"]), len(main_headers) + 1): # After last host RAM col
                    set_cell_style(ws.cell(row=current_row, column=c_empty_total_after), fill_color=header_fill_color, border_style="thin")


                current_row += 1
            current_row += 2 # Extra space between vendor blocks

        # --- Auto-size columns (basic fit) ---
        # This is a basic auto-size, might not be perfect for merged cells or very long content.
        # Set specific widths for better control if needed.
        column_widths_map = {
            'A': 20, 'B': 12, 'C': 12, 'D': 15, 'E': 22, 'F': 50, # Course Job Code to Course Name
            'G': 12, 'H': 3, 'I': 12, # Start/End Pod, Arrow, End Pod, Username
            'J': 12, 'K': 7, 'L': 7, 'M': 15, # Password, Class, Students, Vendor Pods
            'N': 25, 'O': 25, 'P': 25, # Group, Version, Course Version
            'Q': 7, 'R': 30, # RAM, Virtual Hosts
            'S': 4, 'T': 4, 'U': 4, 'V': 4, 'W': 4, 'X': 4 # Ni, Cl, Ul, Un, Ho, K2
        }
        for col_letter, width in column_widths_map.items():
            ws.column_dimensions[col_letter].width = width
        # Example for specific columns from your image:
        # ws.column_dimensions['A'].width = 20 # Course Job Code
        # ws.column_dimensions['F'].width = 60 # Course Name
        # ... and so on for all 24 columns ...

        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        logger.info("Excel export generated successfully (new format).")
        return send_file(
            excel_buffer,
            as_attachment=True,
            download_name="labbuild_allocations_detailed.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        logger.error(f"Error generating detailed Excel export: {e}", exc_info=True)
        flash("Error generating detailed Excel export. Please check server logs.", "danger")
        return redirect(url_for('main.view_allocations'))


def _generate_random_password(length=8) -> str:
    """Generates a random numeric password of specified length."""
    return "".join(random.choice(string.digits) for _ in range(length))

def _create_contiguous_ranges(pod_numbers: List[Union[int,str]]) -> str:
    """
    Converts a list of pod numbers (can be int or string)
    into a comma-separated string of contiguous ranges.
    Example: [1, 2, 3, '5', 6, 8] -> "1-3,5-6,8"
    """
    if not pod_numbers:
        return ""
    
    processed_pod_numbers: List[int] = []
    for p in pod_numbers:
        try:
            processed_pod_numbers.append(int(p))
        except (ValueError, TypeError):
            logger.warning(f"Could not convert pod number '{p}' to int in _create_contiguous_ranges. Skipping.")
            continue
    
    if not processed_pod_numbers:
        return ""
    
    pod_numbers_sorted_unique = sorted(list(set(processed_pod_numbers)))
    if not pod_numbers_sorted_unique: # Should be redundant if processed_pod_numbers is checked
        return ""

    ranges = []
    start_range = pod_numbers_sorted_unique[0]
    end_range = pod_numbers_sorted_unique[0]

    for i in range(1, len(pod_numbers_sorted_unique)):
        if pod_numbers_sorted_unique[i] == end_range + 1:
            end_range = pod_numbers_sorted_unique[i]
        else:
            if start_range == end_range:
                ranges.append(str(start_range))
            else:
                ranges.append(f"{start_range}-{end_range}")
            start_range = pod_numbers_sorted_unique[i]
            end_range = pod_numbers_sorted_unique[i]
    
    # Add the last range
    if start_range == end_range:
        ranges.append(str(start_range))
    else:
        ranges.append(f"{start_range}-{end_range}")
        
    return ",".join(ranges)


@bp.route('/generate-apm-commands', methods=['POST'])
def generate_apm_commands_route():
    batch_review_id_req = request.json.get('batch_review_id') # For context/logging
    logger.info(f"--- APM Command Generation Route Called (Batch ID context: {batch_review_id_req}) ---")

    # Fetch necessary current data for the APM logic helper
    current_apm_entries_route: Dict[str, Dict[str, Any]] = {}
    try:
        apm_list_url_route = os.getenv("APM_LIST_URL", "http://connect.rededucation.com:1212/list")
        response_route = requests.get(apm_list_url_route, timeout=15); response_route.raise_for_status()
        current_apm_entries_route = response_route.json()
    except Exception as e_route: 
        logger.error(f"APM Route: Error fetching current APM entries: {e_route}")
        return jsonify({"commands": ["# ERROR: Could not fetch current APM entries from server."], "message": f"Error fetching current APM: {e_route}", "error": True}), 500

    extended_apm_codes_route: Set[str] = set()
    try: # Fetch extended codes
        if alloc_collection is not None:
            cursor_route = alloc_collection.find({"extend": "true"}, {"tag": 1, "courses.apm_vpn_auth_course_code": 1, "_id": 0})
            for doc_route in cursor_route:
                found_route = False;
                for ca_route in doc_route.get("courses",[]):
                    if ca_route.get("apm_vpn_auth_course_code"): extended_apm_codes_route.add(ca_route["apm_vpn_auth_course_code"]); found_route=True
                if not found_route and doc_route.get("tag"): extended_apm_codes_route.add(doc_route.get("tag"))
    except PyMongoError as e_ext_route: logger.error(f"APM Route: Error fetching extended: {e_ext_route}")

    plan_items_for_route_apm: List[Dict] = []
    if interim_alloc_collection is not None:
        try:
            # Fetch all confirmed items, as batch_id might be lost if user reloads final review page
            # Or, if JS always sends batch_id, use it:
            query_filter_route = {"status": {"$in": ["student_confirmed", "trainer_confirmed"]}}
            if batch_review_id_req: # If JS sends it, use it to scope.
                 query_filter_route["batch_review_id"] = batch_review_id_req
            plan_cursor_route = interim_alloc_collection.find(query_filter_route)
            plan_items_for_route_apm = list(plan_cursor_route)
        except PyMongoError as e_plan_route:
            logger.error(f"APM Route: Error fetching plan items: {e_plan_route}")
            return jsonify({"commands": ["# ERROR: Could not fetch build plan."], "message": f"DB error: {e_plan_route}", "error": True}), 500
    
    # Call the refactored APM logic
    all_commands, _, apm_errors = _generate_apm_data_for_plan(
        plan_items_for_route_apm,
        current_apm_entries_route,
        extended_apm_codes_route
    )
    
    num_actual_adds = sum(1 for cmd in all_commands if " add " in cmd)
    num_actual_deletes = sum(1 for cmd in all_commands if " del " in cmd)
    num_actual_updates = len(all_commands) - num_actual_adds - num_actual_deletes
    final_message = f"Generated {len(all_commands)} APM commands ({num_actual_deletes} deletions, {num_actual_adds} adds, {num_actual_updates} field updates)."

    if apm_errors:
        final_message = "Errors occurred during APM generation: " + " | ".join(apm_errors) + ". " + final_message
        return jsonify({"commands": ["# Errors occurred. Review messages and generated commands.", *apm_errors, *all_commands], "message": final_message, "error": True}), 200
    if not all_commands: return jsonify({"commands": ["# No APM changes identified."], "message": "No APM changes identified."})
    
    logger.info(f"APM Route: Command Generation Successful. {final_message}")
    return jsonify({"commands": all_commands, "message": final_message})

@bp.route('/allocations/update-summary', methods=['POST'])
def update_allocation_summary():
    """
    API endpoint to handle in-line edits of an allocation group's summary.
    """
    if not request.is_json:
        return jsonify({"success": False, "error": "Invalid request format, JSON expected."}), 415

    data = request.json
    logger.info(f"Received allocation summary update request: {data}")

    # --- 1. Extract and Validate Identifiers ---
    tag = data.get('tag')
    # The group is identified by its first course name for the update query
    course_name = data.get('course_name')

    if not tag or not course_name:
        return jsonify({"success": False, "error": "Missing tag or course_name identifier."}), 400

    # --- 2. Build the MongoDB Update Payload ---
    update_payload = {}
    # Create a mapping from the keys received from JS to the DB field names
    field_map = {
        'start_date': 'courses.$.start_date',
        'end_date': 'courses.$.end_date',
        'trainer_name': 'courses.$.trainer_name',
        'apm_username': 'courses.$.apm_username',
        'apm_password': 'courses.$.apm_password'
    }

    for key, db_field in field_map.items():
        if key in data: # Check if the key exists in the request
            update_payload[db_field] = data[key]

    if not update_payload:
        return jsonify({"success": True, "message": "No changes detected."})

    # --- 3. Perform the Database Update ---
    if alloc_collection is None:
        return jsonify({"success": False, "error": "Database service unavailable."}), 503

    try:
        result = alloc_collection.update_one(
            {"tag": tag, "courses.course_name": course_name},
            {"$set": update_payload}
        )

        if result.matched_count == 0:
            logger.warning(f"No allocation found for tag '{tag}' and course '{course_name}' to update.")
            return jsonify({"success": False, "error": "Allocation not found. It may have been recently deleted."}), 404
        
        if result.modified_count == 0:
            logger.info(f"Allocation for tag '{tag}' matched but no fields were changed.")
            return jsonify({"success": True, "message": "No changes made."})

        logger.info(f"Successfully updated allocation summary for tag '{tag}', course '{course_name}'.")
        return jsonify({"success": True, "message": "Allocation updated successfully."})

    except PyMongoError as e:
        logger.error(f"DB error updating allocation summary for tag '{tag}': {e}", exc_info=True)
        return jsonify({"success": False, "error": f"Database error: {e}"}), 500
    except Exception as e:
        logger.error(f"Unexpected error updating allocation summary for tag '{tag}': {e}", exc_info=True)
        return jsonify({"success": False, "error": "An unexpected server error occurred."}), 500
    
@bp.route('/bulk-teardown-tags', methods=['POST'])
def bulk_teardown_tags():
    """
    Handles bulk teardown requests for multiple tags from the notifications page.
    """
    tags_to_teardown = request.form.getlist('tags_to_teardown')
    if not tags_to_teardown:
        flash("No tags were selected for teardown.", "warning")
        return redirect(url_for('main.all_notifications'))

    logger.info(f"Initiating BULK teardown for {len(tags_to_teardown)} tags: {', '.join(tags_to_teardown)}")
    
    tasks_to_run = []
    
    if alloc_collection is None:
        flash("Database unavailable. Cannot process teardown.", "danger")
        return redirect(url_for('main.all_notifications'))
        
    try:
        # Fetch all details for the selected tags in a single query
        tag_docs_cursor = alloc_collection.find({"tag": {"$in": tags_to_teardown}})
        
        for tag_doc in tag_docs_cursor:
            tag = tag_doc.get("tag")
            for course in tag_doc.get("courses", []):
                course_name = course.get("course_name")
                vendor = course.get("vendor")
                if not all([course_name, vendor]): continue

                # Group pod details by host to create efficient teardown commands
                items_by_host = defaultdict(lambda: {"pods": set(), "classes": set()})
                for pd in course.get("pod_details", []):
                    host = pd.get("host")
                    if not host: continue
                    if vendor.lower() == 'f5' and pd.get("class_number") is not None:
                        items_by_host[host]["classes"].add(pd.get("class_number"))
                    elif pd.get("pod_number") is not None:
                        items_by_host[host]["pods"].add(pd.get("pod_number"))
                
                # Create teardown tasks
                for host, items in items_by_host.items():
                    for class_num in sorted(list(items["classes"])):
                        args = ['teardown', '-v', vendor, '-g', course_name, '--host', host, '-t', tag, '-cn', str(class_num)]
                        tasks_to_run.append({'args': args, 'description': f"Tag '{tag}', Class {class_num}"})
                    if items["pods"]:
                        # This could be further optimized to merge ranges, but individual calls are safer
                        for pod_num in sorted(list(items["pods"])):
                            args = ['teardown', '-v', vendor, '-g', course_name, '--host', host, '-t', tag, '-s', str(pod_num), '-e', str(pod_num)]
                            tasks_to_run.append({'args': args, 'description': f"Tag '{tag}', Pod {pod_num}"})

    except Exception as e:
        logger.error(f"Error preparing bulk teardown tasks: {e}", exc_info=True)
        flash("An error occurred while preparing the teardown jobs.", "danger")
        return redirect(url_for('main.all_notifications'))

    if tasks_to_run:
        _dispatch_bulk_labbuild_tasks(tasks_to_run, "bulk tag teardown")
    else:
        flash("No valid items found within the selected tags to tear down.", "info")

    return redirect(url_for('main.all_notifications'))

@bp.route('/export-current-lab-report')
def export_current_lab_report():
    """
    Dummy endpoint for the 'Current Lab Report' export functionality.
    """
    logger.info("DUMMY ACTION: 'Export Current Lab Report' was triggered.")
    flash("Feature not yet implemented: The 'Current Lab Report' export is under development.", "info")
    return redirect(url_for('main.view_allocations'))

# --- NEW: Dummy Route for Trainer Pod Allocation ---
@bp.route('/export-trainer-pod-allocation')
def export_trainer_pod_allocation():
    """
    Dummy endpoint for the 'Trainer Pod Allocation' export functionality.
    """
    logger.info("DUMMY ACTION: 'Trainer Pod Allocation' export was triggered.")
    flash("Feature not yet implemented: The 'Trainer Pod Allocation' export is under development.", "info")
    return redirect(url_for('main.view_allocations'))