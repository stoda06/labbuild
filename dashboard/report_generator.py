# labbuild/dashboard/report_generator.py

import io
import logging
import requests
import pymongo
from flask import current_app
from datetime import datetime
from collections import defaultdict
from typing import List, Dict, Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


from constants import (
    INTERIM_ALLOCATION_COLLECTION,
    ALLOCATION_COLLECTION,
    HOST_COLLECTION,
    COURSE_CONFIG_COLLECTION # Assuming 'courseconfig' is defined in constants
)

from excelreport_config import (
    ExcelStyle, LOG_LEVEL_GENERATE_EXCEL,
    AU_HOST_NAMES, US_HOST_NAMES,
    AVAILABLE_RAM_GB, SUMMARY_ENV_ORDER,
    RAM_SUMMARY_START_COL, EXCEL_GROUP_ORDER, EXCEL_COLUMN_WIDTHS
)


# --- Logging Setup ---
logger = logging.getLogger(__name__)

# ==============================================================================
# ALL THE HELPER AND DATA PROCESSING FUNCTIONS
# ==============================================================================

def calculate_ram_summary(all_allocations: List[Dict], host_map: Dict) -> Dict[str, float]:
    allocated_ram_by_env = {env_key: 0 for env_key in host_map.keys()}
    for allocation in all_allocations:
        ram = convert_to_numeric(allocation.get("ram"))
        if not ram or ram <= 0: continue
        virtual_hosts_str = allocation.get("virtual_hosts", "").lower() if allocation.get("virtual_hosts") else ""
        if not virtual_hosts_str: continue
        for env_key, host_prefix in host_map.items():
            if host_prefix.lower() in virtual_hosts_str:
                allocated_ram_by_env[env_key] += ram
    return allocated_ram_by_env

def convert_to_numeric(val):
    if val is None: return None
    if isinstance(val, (int, float)): return val
    if isinstance(val, str):
        val = val.strip()
        if not val: return None
        try: return int(val)
        except ValueError:
            try: return float(val)
            except ValueError: return val
    return val

def apply_style(cell, trainer=False, use_green_fill=False, is_summary=False):
    if cell is None: return
    cell.border = ExcelStyle.THIN_BORDER.value
    cell.alignment = ExcelStyle.CENTER_ALIGNMENT.value
    if trainer: cell.fill = ExcelStyle.LIGHT_BLUE_FILL.value
    elif use_green_fill: cell.fill = ExcelStyle.GREEN_FILL.value
    elif is_summary: cell.fill = ExcelStyle.LIGHT_BLUE_FILL.value

def write_cell(sheet, row, col, value, trainer=False, use_green_fill=False, number_format=None, is_summary=False):
    cell = sheet.cell(row=row, column=col)
    cell.value = convert_to_numeric(value) if not (isinstance(value, str) and value.startswith("=")) else value
    apply_style(cell, trainer, use_green_fill, is_summary)
    if number_format: cell.number_format = number_format

def write_group_title(sheet, row, title):
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=34)
    cell = sheet.cell(row=row, column=1, value=title)
    cell.font = Font(bold=True, size=14)
    return row + 1

def write_merged_header(sheet, row, col, header_text):
    sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 2)
    for c in range(col, col + 3):
        cell = sheet.cell(row=row, column=c)
        cell.font = Font(bold=True)
        cell.border = ExcelStyle.THIN_BORDER.value
        cell.fill = ExcelStyle.LIGHT_BLUE_FILL.value
        cell.alignment = ExcelStyle.CENTER_ALIGNMENT.value
    sheet.cell(row=row, column=col, value=header_text)

def write_summary_section(sheet, row_offset, allocated_ram_by_env: Dict[str, float]):
    start_col = RAM_SUMMARY_START_COL
    env_keys = SUMMARY_ENV_ORDER
    headers = ["RAM Summary", "Total", *env_keys]
    for i, header in enumerate(headers):
        col = start_col + i
        cell = sheet.cell(row=row_offset, column=col, value=header)
        cell.font = Font(bold=True)
        apply_style(cell, is_summary=True)
    available_ram_row, allocated_ram_row = row_offset + 1, row_offset + 2
    allocated_pct_row, remaining_ram_row = row_offset + 3, row_offset + 4
    sheet.cell(row=available_ram_row, column=start_col, value="Available RAM (GB)").font = Font(bold=True)
    sheet.cell(row=allocated_ram_row, column=start_col, value="Allocated RAM (GB)").font = Font(bold=True)
    sheet.cell(row=allocated_pct_row, column=start_col, value="Allocated RAM (%)").font = Font(bold=True)
    sheet.cell(row=remaining_ram_row, column=start_col, value="Remaining RAM (GB)").font = Font(bold=True)
    total_col_letter = get_column_letter(start_col + 1)
    for i, env_key in enumerate(env_keys):
        col, col_letter = start_col + 2 + i, get_column_letter(start_col + 2 + i)
        available_ram = AVAILABLE_RAM_GB.get(env_key, 0)
        write_cell(sheet, available_ram_row, col, available_ram, is_summary=False, number_format='0')
        allocated_ram = allocated_ram_by_env.get(env_key, 0)
        write_cell(sheet, allocated_ram_row, col, allocated_ram, is_summary=False, number_format='0.0')
        formula_pct = f"=IF({col_letter}{available_ram_row}>0, {col_letter}{allocated_ram_row}/{col_letter}{available_ram_row}, 0)"
        write_cell(sheet, allocated_pct_row, col, formula_pct, is_summary=False, number_format="0%")
        formula_rem = f"={col_letter}{available_ram_row}-{col_letter}{allocated_ram_row}"
        write_cell(sheet, remaining_ram_row, col, formula_rem, is_summary=False, number_format='0.0')
    first_env, last_env = get_column_letter(start_col + 2), get_column_letter(start_col + 1 + len(env_keys))
    write_cell(sheet, available_ram_row, start_col + 1, f"=SUM({first_env}{available_ram_row}:{last_env}{available_ram_row})", number_format='0')
    write_cell(sheet, allocated_ram_row, start_col + 1, f"=SUM({first_env}{allocated_ram_row}:{last_env}{allocated_ram_row})", number_format='0.0')
    write_cell(sheet, allocated_pct_row, start_col + 1, f"=IF({total_col_letter}{available_ram_row}>0, {total_col_letter}{allocated_ram_row}/{total_col_letter}{available_ram_row}, 0)", number_format="0%")
    write_cell(sheet, remaining_ram_row, start_col + 1, f"={total_col_letter}{available_ram_row}-{total_col_letter}{allocated_ram_row}", number_format='0.0')
    for r in [allocated_ram_row, allocated_pct_row, remaining_ram_row]:
        for c in range(start_col + 1, start_col + 2 + len(env_keys)):
            cell = sheet.cell(row=r, column=c)
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            cell.font = Font(bold=True, color="FF0000")

def _write_data_row(sheet, row, entry, headers, header_keys, header_pos, is_trainer, use_green_fill, host_map: Dict):
    trainer_flag, green_flag = (True, False) if is_trainer else (False, use_green_fill)
    virtual_hosts_str = entry.get("virtual_hosts", "").lower() if entry.get("virtual_hosts") else ""
    col = 1
    for h in header_keys:
        if h == "Start/End Pod":
            pod = entry.get("start_end_pod", "")
            parts = pod.replace("→", "->").replace("–", "-").split("-")
            left, right = (parts[0].strip(), parts[1].strip()) if len(parts) > 1 else (parts[0].strip() if parts else "", "")
            right = right if right else left
            write_cell(sheet, row, col, left, trainer=trainer_flag, use_green_fill=green_flag); col += 1
            write_cell(sheet, row, col, "->", trainer=trainer_flag, use_green_fill=green_flag); col += 1
            write_cell(sheet, row, col, right, trainer=trainer_flag, use_green_fill=green_flag); col += 1
        elif h in host_map:
            host_name_for_this_column = host_map[h].lower()
            if host_name_for_this_column in virtual_hosts_str:
                ram_col, pods_col = get_column_letter(header_pos["RAM"]), get_column_letter(header_pos["Vendor Pods"])
                formula = f"=IF({pods_col}{row}<=1, {ram_col}{row}, {ram_col}{row} + ({pods_col}{row}-1)*{ram_col}{row}/2)"
                write_cell(sheet, row, col, formula, trainer_flag, green_flag, number_format='0.0')
            else:
                write_cell(sheet, row, col, 0, trainer_flag, green_flag, number_format='0.0')
            col += 1
        else:
            data_key = headers.get(h)
            val_to_write = entry.get(data_key) if data_key else ""
            num_format = '0.0' if h == "RAM" else None
            write_cell(sheet, row, col, val_to_write, trainer_flag, green_flag, number_format=num_format)
            col += 1

def write_group_summary_boxes(sheet, start_row, header_pos, group_pod_total, group_host_ram_totals, group_name, group_end_col):
    vendor_col = header_pos.get("Vendor Pods")
    has_pod_summary, has_ram_summary = vendor_col and group_pod_total > 0, sum(group_host_ram_totals.values()) > 0
    if not (has_pod_summary or has_ram_summary): return start_row
    label_start, label_end = start_row, start_row + 1
    value_start, value_end = start_row + 2, start_row + 3
    summary_end = value_end
    thin, no_side = Side(style='thin'), Side(style=None)
    border_top, border_bottom = Border(left=thin, right=thin, top=thin, bottom=no_side), Border(left=thin, right=thin, top=no_side, bottom=thin)
    for r in range(start_row, summary_end + 1):
        for c in range(1, group_end_col + 1):
            cell = sheet.cell(row=r, column=c)
            cell.alignment = ExcelStyle.CENTER_ALIGNMENT.value
            cell.border = border_top if r in [label_start, value_start] else border_bottom
    if has_pod_summary:
        merge_start, merge_end = vendor_col - 1, vendor_col + 1
        sheet.merge_cells(start_row=label_start, start_column=merge_start, end_row=label_end, end_column=merge_end)
        label_cell = sheet.cell(row=label_start, column=merge_start, value="Total Pods")
        label_cell.font, label_cell.alignment, label_cell.fill = Font(bold=True, size=16), Alignment(horizontal='center', vertical='center'), ExcelStyle.LIGHT_BLUE_FILL.value
        sheet.merge_cells(start_row=value_start, start_column=merge_start, end_row=value_end, end_column=merge_end)
        value_cell = sheet.cell(row=value_start, column=merge_start, value=group_pod_total)
        value_cell.font, value_cell.alignment, value_cell.fill = Font(bold=True, size=14), Alignment(horizontal='center', vertical='center'), ExcelStyle.LIGHT_BLUE_FILL.value
    if has_ram_summary:
        for host_key, host_ram in group_host_ram_totals.items():
            if host_col := header_pos.get(host_key):
                cell_top, cell_bottom = sheet.cell(row=value_start, column=host_col), sheet.cell(row=value_end, column=host_col)
                cell_bottom.value, cell_bottom.number_format, cell_bottom.font = host_ram if host_ram > 0 else 0, '0', Font(bold=True, color="FFFF00")
                cell_top.fill, cell_bottom.fill = ExcelStyle.LIGHT_BLUE_FILL.value, ExcelStyle.LIGHT_BLUE_FILL.value
    return start_row + 4

def apply_outer_border(sheet, start_row, end_row, start_col, end_col):
    medium, no_side = Side(style='medium'), Side(style=None)
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            cell = sheet.cell(row=r, column=c)
            b = cell.border
            is_top, is_bottom, is_left, is_right = (r == start_row), (r == end_row), (c == start_col), (c == end_col)
            if is_top or is_bottom or is_left or is_right:
                cell.border = Border(left=medium if is_left else b.left, right=medium if is_right else b.right, top=no_side if is_top else b.top, bottom=medium if is_bottom else b.bottom)

def determine_us_au_location(virtual_hosts_str: str) -> str:
    if not virtual_hosts_str: return ""
    hosts_lower = virtual_hosts_str.lower()
    if any(au in hosts_lower for au in AU_HOST_NAMES): return "AU"
    if any(us in hosts_lower for us in US_HOST_NAMES): return "US"
    return ""

def format_date(date_str):
    if not date_str: return ""
    try: return datetime.strptime(date_str, '%Y-%m-%d').strftime('%A, %d/%m')
    except Exception: return date_str

def get_vendor_prefix_map():
    return {gn.split(' ')[0].lower(): gn for gn in ExcelStyle.DEFAULT_COURSE_GROUPS.value.keys()}

def find_location_from_code(course_code: str, location_map: dict) -> str:
    if not course_code: return ""
    for loc_code in sorted(location_map.keys(), key=len, reverse=True):
        if loc_code in course_code: return location_map[loc_code]
    return ""

def _find_field(doc, keys):
    course_details = doc.get('courses', [{}])[0]
    for key in keys:
        if doc.get(key) is not None: return doc.get(key)
        if course_details.get(key) is not None: return course_details.get(key)
    return None

# ==============================================================================
# CREDENTIAL HANDLING
# ==============================================================================
def fetch_apm_credentials():
    try:
        url = "http://connect:1212/list"
        response = requests.get(url, timeout=70)
        response.raise_for_status()
        return response.json()
    except requests.RequestException as e:
        logger.error(f"Failed to fetch APM credentials: {e}")
        return {}

def build_apm_lookup(apm_data):
    lookup_by_code = {}
    if not isinstance(apm_data, dict):
        logger.warning("APM credential data is not valid. Cannot build lookup.")
        return {}
    for username, entry in apm_data.items():
        course_code = entry.get("vpn_auth_course_code")
        if course_code and course_code.lower() != "trainer":
            lookup_by_code[course_code] = {
                "username": username,
                "password": entry.get("vpn_auth_class")
            }
    return lookup_by_code

# ==============================================================================
# DATA UNPACKING FUNCTIONS
# ==============================================================================
def unpack_current_allocations(documents, vendor_map, apm_lookup_by_code, location_map, host_to_vcenter_map, ram_lookup_map):
    unpacked_data = {'standard': [], 'extended': [], 'trainer': []}
    grouped = defaultdict(lambda: {'pod_numbers': set(), 'hosts': set(), 'doc': None})
    for doc in documents:
        if not (key := doc.get('tag')): continue
        if not grouped[key]['doc']: grouped[key]['doc'] = doc
        for c in doc.get('courses', []):
            for pd in c.get('pod_details', []):
                for p in pd.get("pods", [pd]):
                    if p.get("pod_number") is not None: grouped[key]['pod_numbers'].add(int(p.get("pod_number")))
                    if p.get("host"): grouped[key]['hosts'].add(p.get("host"))

    for course_code, val in grouped.items():
        if not val['pod_numbers']: continue
        doc = val['doc']

        # <<< START OF FIX for RAM >>>
        numeric_ram = None
        # Attempt 1: Get RAM from the allocation document itself
        raw_ram_from_alloc = _find_field(doc, ['ram', 'memory_gb_one_pod'])
        if raw_ram_from_alloc is not None:
            numeric_ram = convert_to_numeric(raw_ram_from_alloc)

        # Attempt 2 (Fallback): If RAM not found, use the courseconfig lookup map
        if not isinstance(numeric_ram, (int, float)) or numeric_ram <= 0:
            course_version = _find_field(doc, ['course_name', 'version'])
            if course_version and course_version in ram_lookup_map:
                ram_from_config = ram_lookup_map.get(course_version)
                numeric_ram = convert_to_numeric(ram_from_config)
                logger.debug(f"Used fallback RAM '{numeric_ram}' for course '{course_version}'")
        # <<< END OF FIX for RAM >>>

        lookup_key = course_code.upper().removesuffix('-TP')
        creds_from_apm = apm_lookup_by_code.get(lookup_key, {})
        final_username = _find_field(doc, ['apm_username']) or creds_from_apm.get('username')
        final_password = _find_field(doc, ['apm_password']) or creds_from_apm.get('password')
        
        pod_nums, pod_count = sorted(val['pod_numbers']), len(val['pod_numbers'])
        pod_range = f"{pod_nums[0]}-{pod_nums[-1]}" if pod_count > 1 else str(pod_nums[0])
        course_hosts = sorted(val['hosts'])
        
        if course_code.upper().endswith("-TP"):
            pod_type, target_list = 'trainer', unpacked_data['trainer']
        elif doc.get('extend') == 'true':
            pod_type, target_list = 'extended', unpacked_data['extended']
        else:
            pod_type, target_list = 'standard', unpacked_data['standard']

        base_pod_data = {
            'course_code': course_code, 'us_au_location': determine_us_au_location(", ".join(course_hosts)),
            'pod_type': pod_type, 'start_end_pod': pod_range, 'location': find_location_from_code(course_code, location_map),
            'course_start_date': format_date(_find_field(doc, ['start_date', 'sf_start_date'])),
            'last_day': format_date(_find_field(doc, ['end_date', 'sf_end_date'])),
            'username': final_username, 'password': final_password,
            'trainer_name': _find_field(doc, ['trainer_name']),
            'ram': numeric_ram,  # Use the cleaned/fallback numeric RAM value
            'vendor_pods': pod_count, 'students': pod_count,
            'course_name': _find_field(doc, ['course_name', 'version']) or 'N/A',
            'version': _find_field(doc, ['course_name', 'version']) or 'N/A',
            'course_version': _find_field(doc, ['course_name', 'version']) or 'N/A',
        }
        if not course_hosts:
            pod = base_pod_data.copy()
            pod['virtual_hosts'], pod['vcenter_name'] = None, ""
            target_list.append(pod)
        else:
            for host in course_hosts:
                pod = base_pod_data.copy()
                pod['virtual_hosts'], pod['vcenter_name'] = host, host_to_vcenter_map.get(host, "")
                target_list.append(pod)
    return unpacked_data['standard'], unpacked_data['extended'], unpacked_data['trainer']

# ==============================================================================
# EXCEL GENERATION & MAIN ORCHESTRATION
# ==============================================================================
# labbuild/dashboard/report_generator.py

# ... (all other functions remain the same) ...

# ==============================================================================
# EXCEL GENERATION & MAIN ORCHESTRATION
# ==============================================================================
def generate_excel_in_memory(course_allocations: List[Dict], trainer_pods: List[Dict], extended_pods: List[Dict], host_map: Dict) -> io.BytesIO:
    wb = Workbook()
    sheet = wb.active

    sheet.title = "Labbuild"
    all_data = course_allocations + trainer_pods + extended_pods
    allocated_ram_data = calculate_ram_summary(all_data, host_map)
    expanded_data = []
    for entry in all_data:
        hosts_val = entry.get("virtual_hosts")
        hosts = hosts_val if isinstance(hosts_val, str) else ""
        host_list = [h.strip() for h in hosts.split(",") if h.strip()]
        if len(host_list) <= 1: expanded_data.append(entry)
        else:
            for host in host_list:
                new_entry = entry.copy()
                new_entry["virtual_hosts"] = host
                expanded_data.append(new_entry)
    all_data = expanded_data
    grouped = {g: [] for g in ExcelStyle.DEFAULT_COURSE_GROUPS.value}
    for entry in all_data:
        for gname, fn in ExcelStyle.DEFAULT_COURSE_GROUPS.value.items():
            if fn(entry.get("course_code", "")):
                grouped[gname].append(entry)
                break
    group_order, current_row = EXCEL_GROUP_ORDER, 12
    for group_name in group_order:
        if not (records := grouped.get(group_name, [])): continue
        group_start_row = current_row
        logger.info(f"Writing group: {group_name} with {len(records)} records")
        current_row = write_group_title(sheet, current_row, group_name)
        headers = ExcelStyle.DEFAULT_HEADER_COURSE_MAPPING.value[group_name]
        header_keys, col_idx, header_pos, group_end_col = list(headers), 1, {}, 0
        for h in header_keys:
            header_pos[h], width = col_idx, 3 if h == "Start/End Pod" else 1
            col_idx += width
            group_end_col += width
        col = 1
        for h in header_keys:
            if h == "Start/End Pod": write_merged_header(sheet, current_row, col, h); col += 3
            else:
                cell = sheet.cell(row=current_row, column=col, value=h)
                cell.font = Font(bold=True)
                apply_style(cell, is_summary=True)
                col += 1
        current_row += 1
        group_host_ram_totals, group_pod_total = {k: 0 for k in host_map}, 0
        
        # Combine all records for simpler summary logic
        all_records_in_group = [e for e in records]

        # Function to calculate RAM based on pods
        def get_calculated_ram(entry):
            base_ram = convert_to_numeric(entry.get("ram")) or 0
            if base_ram <= 0:
                return 0
            pods = convert_to_numeric(entry.get("vendor_pods")) or 1
            if pods <= 1:
                return base_ram
            else:
                return base_ram + (pods - 1) * base_ram / 2

        # <<< START OF FIX >>>
        # Calculate totals using the corrected logic first
        for entry in all_records_in_group:
            group_pod_total += convert_to_numeric(entry.get("vendor_pods")) or 0
            calculated_ram_for_entry = get_calculated_ram(entry)
            if calculated_ram_for_entry > 0:
                 for host_key, host_name in host_map.items():
                    if host_name.lower() in (entry.get("virtual_hosts", "") or "").lower():
                        group_host_ram_totals[host_key] += calculated_ram_for_entry
        # <<< END OF FIX >>>
        
        # Now write the rows
        non_trainer_pods = [e for e in records if e.get("pod_type") != "trainer"]
        trainer_pods_in_group = [e for e in records if e.get("pod_type") == "trainer"]
        
        for entry in non_trainer_pods:
            is_extended = entry.get("pod_type") == "extended"
            _write_data_row(sheet, current_row, entry, headers, header_keys, header_pos, False, is_extended, host_map)
            current_row += 1
        
        if non_trainer_pods and trainer_pods_in_group:
            for col in range(1, group_end_col + 1):
                sheet.cell(row=current_row, column=col).border = ExcelStyle.THIN_BORDER.value
            current_row += 1
            
        for entry in trainer_pods_in_group:
            _write_data_row(sheet, current_row, entry, headers, header_keys, header_pos, True, False, host_map)
            current_row += 1

        summary_start_row = current_row
        current_row = write_group_summary_boxes(sheet, summary_start_row, header_pos, group_pod_total, group_host_ram_totals, group_name, group_end_col)
        end_row = current_row - 1 if summary_start_row < current_row else summary_start_row - 1
        apply_outer_border(sheet, group_start_row, end_row, 1, group_end_col)
        current_row += 1
        
    write_summary_section(sheet, 2, allocated_ram_data)
    for col_letter, width in EXCEL_COLUMN_WIDTHS.items():
        sheet.column_dimensions[col_letter].width = width
    in_memory_fp = io.BytesIO()
    wb.save(in_memory_fp)
    in_memory_fp.seek(0)
    return in_memory_fp

def get_full_report_data(db):
    try:
        if db is None: raise ConnectionError("A valid database connection was not provided.")
        logger.info("Fetching data using provided DB connection...")
        # <<< START OF FIX for RAM >>>
        # Fetch allocations, configs, and other metadata
        current_docs = list(db[ALLOCATION_COLLECTION].find({}))
        course_configs = list(db[COURSE_CONFIG_COLLECTION].find({}, {"course_name": 1, "memory": 1}))
        locations_data, host_docs = list(db["locations"].find({})), list(db[HOST_COLLECTION].find({}))

        # Create the RAM lookup map from course configs
        ram_lookup_map = {
            doc['course_name']: doc['memory']
            for doc in course_configs if 'course_name' in doc and 'memory' in doc
        }
        logger.info(f"Built RAM lookup map with {len(ram_lookup_map)} entries from courseconfig.")
        # <<< END OF FIX for RAM >>>

        host_map_for_summary = {doc['host_shortcode'].capitalize(): doc['host_shortcode'] for doc in host_docs if 'host_shortcode' in doc and doc.get('include_for_build') == 'true'}
        logger.info(f"Dynamically built HOST_MAP for summary: {host_map_for_summary}")
        
        host_to_vcenter_map = {}
        for doc in host_docs:
            if doc.get('host_name') and doc.get('vcenter'):
                vcenter_fqdn = doc['vcenter']
                short_vcenter_name = vcenter_fqdn.split('.')[0] if isinstance(vcenter_fqdn, str) else ""
                host_to_vcenter_map[doc['host_name']] = short_vcenter_name

        location_map = {loc['code']: loc['name'] for loc in locations_data if 'code' in loc and 'name' in loc}
        logger.info("Successfully fetched data from MongoDB.")
        
        vendor_map = get_vendor_prefix_map()
        apm_data = fetch_apm_credentials()
        apm_lookup_by_code = build_apm_lookup(apm_data)
        
        # Pass the new ram_lookup_map to the unpacking function
        course_allocs, extended_pods, all_trainer_pods = unpack_current_allocations(current_docs, vendor_map, apm_lookup_by_code, location_map, host_to_vcenter_map, ram_lookup_map)
        
        logger.info(f"Processed: {len(course_allocs)} Standard | {len(extended_pods)} Extended | {len(all_trainer_pods)} Trainer pods")
        return course_allocs, all_trainer_pods, extended_pods, host_map_for_summary
    except Exception as e:
        logger.error(f"FATAL: An error occurred during data fetching: {e}", exc_info=True)
        raise e