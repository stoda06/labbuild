from enum import Enum
from openpyxl.styles import PatternFill, Border, Side, Alignment
import logging

# ==============================================================================
# EXCEL REPORT: GENERAL & LAYOUT CONSTANTS
# ==============================================================================
LOG_LEVEL_GENERATE_EXCEL = logging.DEBUG

HOST_MAP = {
    "Ni": "nightbird", "Cl": "cliffjumper", "Ul": "ultramagnus",
    "Un": "unicron", "Ho": "hotshot", "Tr": "Trainer"
}
AU_HOST_NAMES = ["nightbird", "cliffjumper", "ultramagnus", "unicron"]
US_HOST_NAMES = ["hotshot"]

AVAILABLE_RAM_GB = {
    "Ni": 1800, "Cl": 500, "Ul": 2000, "Un": 2000, "Ho": 2000, "Tr": 1000,
}

SUMMARY_ENV_ORDER = ["Ni", "Cl", "Ul", "Un", "Ho", "Tr"]
RAM_SUMMARY_START_COL = 19

EXCEL_GROUP_ORDER = [
    "PA Courses", "CP Courses", "NU Courses",
    "F5 Courses", "AV Courses", "PR Courses"
]

EXCEL_COLUMN_WIDTHS = {
    'A': 22, 'B': 8, 'C': 14, 'D': 14, 'E': 14, 'F': 24, 'G': 28, 'H': 6,
    'I': 4, 'J': 6, 'K': 14, 'L': 14, 'M': 8, 'N': 8, 'O': 10, 'P': 6,
    'Q': 10, 'R': 18, 'S': 15, 'T': 12, 'U': 8, 'V': 8, 'W': 8, 'X': 8,
    'Y': 8, 'Z': 8, 'AA': 8
}

# ==============================================================================
# EXCEL REPORT: HEADER & GROUP DEFINITIONS
# ==============================================================================

DEFAULT_PA_HEADER = {
    "Course Code": 'course_code', "US/AU": 'us_au_location', "Start Date": 'course_start_date',
    "Last Day": 'last_day', "Location": 'location', "Trainer Name": 'trainer_name',
    "Course Name": 'course_name', "Start/End Pod": 'start_end_pod',
    "Username": 'username', "Password": 'password', "Class": 'class_number',
    "Students": 'students', "Vendor Pods": 'vendor_pods', "Group": 'course_type',
    "Version": 'version', "Course Version": 'course_version',
    "RAM": 'ram', "Virtual Hosts": 'virtual_hosts',
    "Ni": '', "Cl": '', "Ul": '', "Un": '', "Ho": '', "Tr": ''
}

DEFAULT_CP_HEADER = {
    "Course Code": 'course_code', "US/AU": 'us_au_location', "Start Date": 'course_start_date',
    "Last Day": 'last_day', "Location": 'location', "Trainer Name": 'trainer_name',
    "Course Name": 'course_name', "Start/End Pod": 'start_end_pod',
    "Username": 'username', "Password": 'password', "Class": 'class_number',
    "Students": 'students', "Vendor Pods": 'vendor_pods', "Group": 'course_type',
    "Version": 'version', "vcenter": 'vcenter_name', "RAM": 'ram', "Virtual Hosts": 'virtual_hosts',
    "Ni": '', "Cl": '', "Ul": '', "Un": '', "Ho": '', "Tr": ''
}

DEFAULT_NU_HEADER = DEFAULT_CP_HEADER.copy()

DEFAULT_F5_HEADER = {
    "Course Code": 'course_code', "US/AU": 'us_au_location', "Start Date": 'course_start_date',
    "Last Day": 'last_day', "Location": 'location', "Trainer Name": 'trainer_name',
    "Course Name": 'course_name', "Start/End Pod": 'start_end_pod',
    "Username": 'username', "Password": 'password', "Class": 'class_number',
    "Students": 'students', "Vendor Pods": 'vendor_pods', "Group": 'course_type',
    "Version": 'version', "buildf5?": 'buildf5', "RAM": 'ram', "Virtual Hosts": 'virtual_hosts',
    "Ni": '', "Cl": '', "Ul": '', "Un": '', "Ho": '', "Tr": ''
}

DEFAULT_GENERIC_HEADER = DEFAULT_CP_HEADER.copy()
DEFAULT_HEADER = DEFAULT_GENERIC_HEADER

DEFAULT_COURSE_GROUPS = {
    "PA Courses": lambda code: str(code).startswith("PA"),
    "CP Courses": lambda code: str(code).startswith("CP"),
    "NU Courses": lambda code: str(code).startswith("NU"),
    "F5 Courses": lambda code: str(code).startswith("F5"),
    "AV Courses": lambda code: str(code).startswith("AV"),
    "PR Courses": lambda code: str(code).startswith("PR")
}

# ==============================================================================
# EXCEL REPORT: STYLING ENUM
# ==============================================================================
class ExcelStyle(Enum):
    SHEET_TITLE = "Sheet1"
    FONT_SIZE_BOLD = 14
    THIN_BORDER = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )
    MEDIUM_OUTER_BORDER = Border(
        left=Side(style='medium'),
        right=Side(style='medium'),
        top=Side(style='medium'),
        bottom=Side(style='medium')
    )
    CENTER_ALIGNMENT = Alignment(horizontal='center', vertical='center')
    LIGHT_BLUE_FILL = PatternFill(start_color="FF00B0F0", end_color="FF00B0F0", fill_type="solid")
    GREEN_FILL = PatternFill(start_color="FF70AD47", end_color="FF70AD47", fill_type="solid")
    DEFAULT_COURSE_GROUPS = DEFAULT_COURSE_GROUPS
    DEFAULT_HEADER_COURSE_MAPPING = {
        "AV Courses": DEFAULT_GENERIC_HEADER,
        "PA Courses": DEFAULT_PA_HEADER,
        "CP Courses": DEFAULT_CP_HEADER,
        "NU Courses": DEFAULT_NU_HEADER,
        "F5 Courses": DEFAULT_F5_HEADER,
        "PR Courses": DEFAULT_GENERIC_HEADER,
    }

# ==============================================================================
# TRAINER POD REPORT CONSTANTS
# ==============================================================================
REPORT_SECTIONS = [
    "F5 COURSE", "CHECK POINT", "PALO ALTO", "NUTANIX", "AV COURSES", "PR COURSE"
]

VENDOR_GROUP_MAP = {
    "F5": "F5 COURSE", "CP": "CHECK POINT", "PA": "PALO ALTO", "PANGF": "PALO ALTO",
    "PCNSA": "PALO ALTO", "NU": "NUTANIX", "AV": "AV COURSES", "PR": "PR COURSE"
}

TRAINER_SHEET_HEADERS = {
    "Course Name": "course_name", "Pod Number": "pod_number", "Username": "username",
    "Password": "password", "Version": "version", "RAM": "ram", "Class": "class",
    "Host": "host", "vCenter": "vcenter", "Taken By": "taken_by",
    "Don't Delete Until US Courses Complete": "notes"
}
