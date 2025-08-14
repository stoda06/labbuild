import io
import logging
import requests
from datetime import datetime
from typing import List, Dict
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# --- Setup basic logging to see the output ---
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# --- Constants ---
ALLOCATION_COLLECTION = "currentallocation"
HOST_COLLECTION = "host"
INTERIM_ALLOCATION_COLLECTION = "interimallocation"
COURSE_CONFIG_COLLECTION = "courseconfig"  ### NEW: Added constant for the course config collection

class ExcelStyle:
    CENTER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)
    THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    CYAN_HEADER_FILL = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")
    BLACK_BOLD_FONT = Font(bold=True, color="000000")

TRAINER_SHEET_HEADERS = {
    'Course Name': 'course_name',
    'Pod Number': 'pod_number',
    'Username': 'username',
    'Password': 'password',
    'Version': 'version',
    'RAM': 'ram',
    'Class': 'class',
    'Host': 'host',
    'vCenter': 'vcenter',
    'Taken By': 'taken_by',
    "Don't Delete Until US Courses Complete": 'notes'
}

REPORT_SECTIONS = [
    "F5 COURSE",
    "CHECK POINT",
    "PALO ALTO",
    "Other Vendors" # Fallback section
]

VENDOR_GROUP_MAP = {
    "PA": "PALO ALTO",
    "F5": "F5 COURSE",
    "CP": "CHECK POINT",
}

logger = logging.getLogger(__name__)

# ==============================================================================
# SPECIALIZED DATA FETCHING HELPERS
# ==============================================================================

def _create_host_to_vcenter_map(db: object) -> Dict[str, str]:
    """Helper to create a mapping from host names to their vCenter."""
    host_map = {}
    try:
        collection = db[HOST_COLLECTION]
        for host_doc in collection.find({}):
            host_name = host_doc.get("host_name")
            vcenter = host_doc.get("vcenter")
            if host_name and vcenter:
                clean_host_name = host_name.strip().lower()
                host_map[clean_host_name] = vcenter
    except Exception as e:
        logger.error(f"Could not create host-to-vcenter map: {e}", exc_info=True)
    logger.info(f"Host-to-vCenter Map created. Total entries: {len(host_map)}")
    return host_map

### NEW: Helper function to create the RAM lookup map for fallbacks ###
def _create_ram_lookup_map(db: object) -> Dict[str, int]:
    """
    Creates a fallback map from course version names to their standard RAM values
    by reading from the course_config collection.
    """
    ram_map = {}
    try:
        collection = db[COURSE_CONFIG_COLLECTION]
        # Fetch only the fields we need
        course_configs = list(collection.find({}, {"course_name": 1, "memory": 1}))
        for doc in course_configs:
            # Ensure both fields exist and are valid before adding to the map
            if doc.get('course_name') and doc.get('memory'):
                ram_map[doc['course_name']] = doc['memory']
        logger.info(f"Built RAM lookup map with {len(ram_map)} entries for fallbacks.")
    except Exception as e:
        logger.error(f"Could not create RAM lookup map from '{COURSE_CONFIG_COLLECTION}': {e}", exc_info=True)
    return ram_map

def _fetch_credentials_from_course2() -> Dict[str, Dict[str, str]]:
    """Fetches credentials and stores username, password, and class name."""
    credential_map = {}
    url = 'http://connect:1212/list'
    try:
        logger.info(f"Fetching credentials from course2 service at {url}")
        response = requests.get(url, timeout=30)
        response.raise_for_status()
        api_data = response.json()

        for username, details in api_data.items():
            if isinstance(details, dict) and details.get('vpn_auth_course_code') == 'Trainer':
                course_version = details.get('vpn_auth_version')
                password_value = details.get('vpn_auth_class')
                class_name = details.get('vpn_auth_class')

                if course_version and username and password_value:
                    lookup_key = course_version.strip()
                    credential_map[lookup_key] = {
                        "username": username,
                        "password": password_value,
                        "class": class_name
                    }
        logger.info(f"Successfully loaded {len(credential_map)} TRAINER credentials from course2.")
    except requests.exceptions.RequestException as e:
        logger.error(f"FATAL: Could not connect to course2 service. Credentials will be blank. Error: {e}")
    except (ValueError, requests.exceptions.JSONDecodeError):
        logger.error(f"FATAL: Could not parse JSON response from course2 service.")
    return credential_map

### MODIFIED: Added `ram_lookup_map` as a parameter ###
def _fetch_from_current_allocations(db: object, host_map: Dict, credential_map: Dict, ram_lookup_map: Dict) -> List[Dict]:
    """Data Source 1: Fetches trainer pods from 'currentallocation'."""
    processed_data = []
    collection = db[ALLOCATION_COLLECTION]
    
    query = {
        "$or": [
            {"tag": "untagged"},
            {"tag": {"$regex": "Trainer-Pods|[-_]TP", "$options": "i"}}
        ]
    }
    logger.info(f"Fetching CURRENT week data from '{ALLOCATION_COLLECTION}'...")

    try:
        matching_docs = list(collection.find(query))
        logger.info(f"Found {len(matching_docs)} documents matching the trainer pod query.")
        if not matching_docs:
            logger.warning(f"The collection '{ALLOCATION_COLLECTION}' was found, but no documents matched the query. Check if tags like 'CCTE-Trainer-Pods' exist.")
            return []
    except Exception as e:
        logger.error(f"Error executing query on '{ALLOCATION_COLLECTION}': {e}", exc_info=True)
        return []

    for doc in matching_docs:
        for course_item in doc.get('courses', []):
            if not isinstance(course_item, dict): continue
            
            course_name_from_db = course_item.get('course_name', 'N/A')
            creds = credential_map.get(course_name_from_db.strip(), {})
            vendor_code = course_item.get('vendor', '').upper()
            vendor_group_name = VENDOR_GROUP_MAP.get(vendor_code, 'Other Vendors')
            class_value = creds.get('class', '')

            for pod_detail in course_item.get('pod_details', []):
                if not isinstance(pod_detail, dict): continue

                ### MODIFIED: Replaced the simple .get() with the new resilient logic ###
                # Step 1: Check for specific RAM keys in the pod_detail or course_item.
                ram_value = pod_detail.get('memory_gb_one_pod') or pod_detail.get('ram')
                if not ram_value:
                    ram_value = course_item.get('memory_gb_one_pod') or course_item.get('ram')

                # Step 2: If still not found, use the fallback lookup map.
                if not ram_value:
                    course_version_key = course_name_from_db.strip()
                    ram_value = ram_lookup_map.get(course_version_key, 'N/A') # Use 'N/A' as the final default.
                
                host_name_from_alloc = pod_detail.get('host', 'N/A')
                short_host_name = host_name_from_alloc.strip().split('.')[0]
                lookup_key = short_host_name.lower()
                full_vcenter = host_map.get(lookup_key, 'N/A')
                
                processed_data.append({
                    'course_name': course_name_from_db,
                    'pod_number': pod_detail.get('pod_number'),
                    'username': creds.get('username', ''),
                    'password': creds.get('password', ''),
                    'version': course_name_from_db,
                    'ram': ram_value,  # Use the resiliently-fetched RAM value
                    'class': class_value,
                    'host': host_name_from_alloc, 
                    'vcenter': full_vcenter.split('.')[0],
                    'taken_by': '', 'notes': '', 'vendor': vendor_group_name
                })
    return processed_data

### MODIFIED: Added `ram_lookup_map` as a parameter ###
def _fetch_from_interim_allocations(db: object, host_map: Dict, credential_map: Dict, ram_lookup_map: Dict) -> List[Dict]:
    """Data Source 2: Fetches trainer pods from 'interimallocation'."""
    processed_data = []
    collection = db[INTERIM_ALLOCATION_COLLECTION]
    
    query = {
        "$or": [
            {"sf_trainer_name": "Trainer Pods"},
            {"sf_course_type": {"$regex": "Trainer Pod", "$options": "i"}},
            {"sf_course_code": {"$regex": "Trainer Pod", "$options": "i"}}
        ]
    }
    logger.info(f"Fetching NEXT week data from '{INTERIM_ALLOCATION_COLLECTION}'...")

    all_docs = list(collection.find(query))
    logger.info(f"Found {len(all_docs)} documents matching the interim trainer pod query.")

    for doc in all_docs:
        if not doc.get('assignments'):
            host = doc.get('host')
            start_pod = doc.get('start_pod')
            if host and start_pod is not None:
                doc['assignments'] = [{
                    'host': host, 'start_pod': start_pod, 'end_pod': doc.get('end_pod', start_pod)
                }]

    for doc in all_docs:
        username, password = '', ''
        doc_username = doc.get('student_apm_username') or doc.get('username')
        doc_password = doc.get('student_apm_password') or doc.get('password')

        if doc_username and doc_password:
            username, password = doc_username, doc_password
        else:
            course_version_key_from_db = doc.get('final_labbuild_course') or doc.get('labbuild_course')
            credential_lookup_key = course_version_key_from_db.strip() if course_version_key_from_db else None
            
            if credential_lookup_key:
                creds = credential_map.get(credential_lookup_key, {})
                username, password = creds.get('username', ''), creds.get('password', '')
                if not username: logger.warning(f"Creds fallback FAILED for interim doc ID {doc.get('_id')} with key '{credential_lookup_key}'")
            else:
                 logger.warning(f"Direct creds missing and no course key in interim doc ID {doc.get('_id')}")

        course_version_key_from_db = doc.get('final_labbuild_course') or doc.get('labbuild_course')
        vendor_code = doc.get('vendor_shortcode', '').upper()
        vendor_group_name = VENDOR_GROUP_MAP.get(vendor_code, 'Other Vendors')
        class_value = ''
        
        ### MODIFIED: Replaced the simple .get() with the new resilient logic ###
        # Step 1: Check for specific RAM keys at the top level of the document.
        ram_per_pod = doc.get('memory_gb_one_pod') or doc.get('ram')

        # Step 2: If not found, use the fallback lookup map with the course version.
        if not ram_per_pod:
            course_version_key = (course_version_key_from_db or "").strip()
            if course_version_key:
                ram_per_pod = ram_lookup_map.get(course_version_key, 'N/A')
            else:
                ram_per_pod = 'N/A' # Final fallback if no version key exists
        
        for assignment in doc.get('assignments', []):
            try:
                start, end = assignment.get('start_pod'), assignment.get('end_pod')
                if start is not None and end is not None:
                    for pod_num in range(int(start), int(end) + 1):
                        host_name_from_alloc = assignment.get('host', 'N/A')
                        short_host_name = host_name_from_alloc.strip().split('.')[0]
                        lookup_key = short_host_name.lower()
                        full_vcenter = host_map.get(lookup_key, 'N/A')
                        
                        processed_data.append({
                            'course_name': doc.get('sf_course_type', 'N/A'),
                            'pod_number': pod_num,
                            'username': username,
                            'password': password,
                            'version': course_version_key_from_db,
                            'ram': ram_per_pod, # Use the resiliently-fetched RAM value
                            'class': class_value,
                            'host': host_name_from_alloc, 
                            'vcenter': full_vcenter.split('.')[0],
                            'taken_by': '', 
                            'notes': doc.get('trainer_assignment_warning', ''),
                            'vendor': vendor_group_name
                        })
            except (ValueError, TypeError) as e:
                logger.warning(f"Skipping malformed assignment in doc '{doc.get('_id')}': {e}")
                continue
    return processed_data

# ==============================================================================
# MAIN PUBLIC FUNCTION WITH DECISION LOGIC
# ==============================================================================

### MODIFIED: Orchestrator now creates and passes the ram_lookup_map ###
def fetch_trainer_pod_data(db: object) -> List[Dict]:
    """Decides data source, fetches and processes trainer pod data."""
    today_weekday = datetime.today().weekday()
    
    # Create all necessary lookup maps once
    host_map = _create_host_to_vcenter_map(db)
    credential_map = _fetch_credentials_from_course2()
    ram_lookup_map = _create_ram_lookup_map(db) # NEW: Create the RAM fallback map
    
    if today_weekday < 2:
        logger.info("Day is before Wednesday. Using CURRENT week's data source ('currentallocation').")
        # Pass the ram_lookup_map to the function
        return _fetch_from_current_allocations(db, host_map, credential_map, ram_lookup_map)
    else:
        logger.info("Day is on or after Wednesday. Using NEXT week's data source ('interimallocation').")
        # Pass the ram_lookup_map to the function
        return _fetch_from_interim_allocations(db, host_map, credential_map, ram_lookup_map)

# ==============================================================================
# EXCEL REPORT GENERATION LOGIC (No changes needed here)
# ==============================================================================

def _apply_style(cell, fill=None, font=None, alignment=None, border=None):
    """Helper to apply multiple styles to a cell."""
    if fill: cell.fill = fill
    if font: cell.font = font
    if alignment: cell.alignment = alignment
    if border: cell.border = border

def create_trainer_report_in_memory(trainer_pods: List[Dict]) -> io.BytesIO:
    """Generates the trainer pod allocation report in memory."""
    if not trainer_pods:
        logger.warning("No trainer pod data was provided to the report generator. Returning an empty report.")
        wb = Workbook()
        sheet = wb.active
        sheet['A1'] = "No Trainer Pod Data Found"
        in_memory_fp = io.BytesIO()
        wb.save(in_memory_fp)
        in_memory_fp.seek(0)
        return in_memory_fp

    wb = Workbook()
    sheet = wb.active
    sheet.title = "Trainer Pod Allocation"

    num_columns = len(TRAINER_SHEET_HEADERS)
    
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_columns)
    title_cell = sheet["A1"]
    title_cell.value = "TRAINER POD ALLOCATION"
    _apply_style(title_cell, font=Font(bold=True, size=16), alignment=ExcelStyle.CENTER_ALIGNMENT)
    sheet.row_dimensions[1].height = 20

    grouped_by_vendor = defaultdict(list)
    for pod in trainer_pods:
        vendor_name = pod.get("vendor", "Other Vendors")
        grouped_by_vendor[vendor_name].append(pod)

    current_row = 3
    for section_name in REPORT_SECTIONS:
        records_for_this_section = grouped_by_vendor.get(section_name, [])
        if records_for_this_section:
            sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=num_columns)
            cell = sheet.cell(row=current_row, column=1, value=section_name.upper())
            _apply_style(cell, fill=ExcelStyle.CYAN_HEADER_FILL, font=ExcelStyle.BLACK_BOLD_FONT)
            current_row += 1

            headers = list(TRAINER_SHEET_HEADERS.keys())
            for col_idx, header_text in enumerate(headers, 1):
                cell = sheet.cell(row=current_row, column=col_idx, value=header_text)
                _apply_style(cell, fill=ExcelStyle.CYAN_HEADER_FILL, font=ExcelStyle.BLACK_BOLD_FONT,
                             alignment=ExcelStyle.CENTER_ALIGNMENT, border=ExcelStyle.THIN_BORDER)
            current_row += 1

            def sort_key(record):
                course = record.get('course_name', '')
                try: pod_num = int(record.get('pod_number', 0))
                except (ValueError, TypeError): pod_num = 0
                return (course, pod_num)

            for record in sorted(records_for_this_section, key=sort_key):
                for col_idx, header in enumerate(headers, 1):
                    data_key = TRAINER_SHEET_HEADERS[header]
                    value = record.get(data_key)
                    cell = sheet.cell(row=current_row, column=col_idx, value=value)
                    _apply_style(cell, border=ExcelStyle.THIN_BORDER, alignment=ExcelStyle.CENTER_ALIGNMENT)
                current_row += 1
            
            current_row += 1

    header_to_col_letter = {header: get_column_letter(i) for i, header in enumerate(TRAINER_SHEET_HEADERS.keys(), 1)}
    
    sheet.column_dimensions[header_to_col_letter['Course Name']].width = 35
    sheet.column_dimensions[header_to_col_letter['Pod Number']].width = 12
    sheet.column_dimensions[header_to_col_letter['Username']].width = 15
    sheet.column_dimensions[header_to_col_letter['Password']].width = 15
    sheet.column_dimensions[header_to_col_letter['Version']].width = 25
    sheet.column_dimensions[header_to_col_letter['RAM']].width = 8
    sheet.column_dimensions[header_to_col_letter['Class']].width = 10
    sheet.column_dimensions[header_to_col_letter['Host']].width = 25
    sheet.column_dimensions[header_to_col_letter['vCenter']].width = 20
    sheet.column_dimensions[header_to_col_letter["Don't Delete Until US Courses Complete"]].width = 40

    in_memory_fp = io.BytesIO()
    wb.save(in_memory_fp)
    in_memory_fp.seek(0)
    
    logger.info("Successfully generated trainer pod report in memory.")
    return in_memory_fp