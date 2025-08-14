# labbuild/dashboard/upcoming_report_generator.py
# This file generates the "Upcoming Lab Report" which uses data from the
# interimallocation collection and also includes extended courses from the
# currentallocation collection.

import io
import logging
import re
import requests
import pymongo
from flask import current_app
from datetime import datetime
from collections import defaultdict
from typing import List, Dict, Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


# Note: We now import all relevant collection constants
from constants import (
    INTERIM_ALLOCATION_COLLECTION,
    ALLOCATION_COLLECTION,
    HOST_COLLECTION,
    COURSE_CONFIG_COLLECTION
)

from excelreport_config import (
    ExcelStyle,
    AU_HOST_NAMES, US_HOST_NAMES,
    AVAILABLE_RAM_GB, SUMMARY_ENV_ORDER,
    RAM_SUMMARY_START_COL, EXCEL_GROUP_ORDER, EXCEL_COLUMN_WIDTHS
)

# --- Logging Setup ---
logger = logging.getLogger(__name__)

# ==============================================================================
# ALL THE HELPER AND DATA PROCESSING FUNCTIONS
# ==============================================================================

def calculate_ram_summary(all_allocations: List[Dict], host_map: Dict) -> Dict[str, float]:
    allocated_ram_by_env = {env_key: 0 for env_key in host_map.keys()}
    for allocation in all_allocations:
        ram = convert_to_numeric(allocation.get("ram"))
        if not ram or ram <= 0:
            continue
        virtual_hosts_str = allocation.get("virtual_hosts", "")
        virtual_hosts_list = [h.strip().lower() for h in virtual_hosts_str.split(",") if h.strip()]
        if not virtual_hosts_list:
            continue
        for env_key, host_prefix in host_map.items():
            host_prefix_lower = host_prefix.lower().strip()
            if any(host_prefix_lower == vh for vh in virtual_hosts_list):
                total_pods_for_course = convert_to_numeric(allocation.get("vendor_pods")) or 1
                if total_pods_for_course <= 1:
                    allocated_ram_by_env[env_key] += ram
                else:
                    allocated_ram_by_env[env_key] += ram + (total_pods_for_course - 1) * ram / 2
    return allocated_ram_by_env



def convert_to_numeric(val):
    if val is None: return None
    if isinstance(val, (int, float)): return val
    if isinstance(val, str):
        val = val.strip()
        if not val: return None
        try: return int(val)
        except ValueError:
            try: return float(val)
            except ValueError: return val
    return val

# In upcoming_report_generator.py

def apply_style(cell, trainer=False, use_green_fill=False, is_summary=False, is_header=False):
    if cell is None: return
    cell.border = ExcelStyle.THIN_BORDER.value
    cell.alignment = ExcelStyle.CENTER_ALIGNMENT.value
    if is_header:
        cell.fill = ExcelStyle.HEADER_FILL.value
    elif trainer:
        cell.fill = ExcelStyle.LIGHT_BLUE_FILL.value
    elif use_green_fill:
        cell.fill = ExcelStyle.GREEN_FILL.value
    elif is_summary:
        # This is for the top RAM summary section, which can remain blue
        cell.fill = ExcelStyle.LIGHT_BLUE_FILL.value

def write_cell(sheet, row, col, value, trainer=False, use_green_fill=False, number_format=None, is_summary=False):
    cell = sheet.cell(row=row, column=col)
    cell.value = convert_to_numeric(value) if not (isinstance(value, str) and value.startswith("=")) else value
    apply_style(cell, trainer, use_green_fill, is_summary)
    if number_format: cell.number_format = number_format

def write_group_title(sheet, row, title):
    cell = sheet.cell(row=row, column=1, value=title)
    cell.font = Font(bold=True, size=14)
    return row + 1

# In upcoming_report_generator.py

def write_merged_header(sheet, row, col, header_text):
    sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 2)
    for c in range(col, col + 3):
        cell = sheet.cell(row=row, column=c)
        cell.font = Font(bold=True)
        apply_style(cell, is_header=True) # Use the new flag
    sheet.cell(row=row, column=col, value=header_text)

# labbuild/dashboard/upcoming_report_generator.py

def write_summary_section(sheet, row_offset, total_ram_by_host: Dict[str, float]):
    start_col = RAM_SUMMARY_START_COL
    env_keys = SUMMARY_ENV_ORDER
    headers = ["RAM Summary", "Total", *env_keys]
    for i, header in enumerate(headers):
        col = start_col + i
        cell = sheet.cell(row=row_offset, column=col, value=header)
        cell.font = Font(bold=True)
        apply_style(cell, is_summary=True)
        
    available_ram_row, allocated_ram_row = row_offset + 1, row_offset + 2
    allocated_pct_row, remaining_ram_row = row_offset + 3, row_offset + 4
    
    sheet.cell(row=available_ram_row, column=start_col, value="Available RAM (GB)").font = Font(bold=True)
    sheet.cell(row=allocated_ram_row, column=start_col, value="Allocated RAM (GB)").font = Font(bold=True)
    sheet.cell(row=allocated_pct_row, column=start_col, value="Allocated RAM (%)").font = Font(bold=True)
    sheet.cell(row=remaining_ram_row, column=start_col, value="Remaining RAM (GB)").font = Font(bold=True)
    
    total_col_letter = get_column_letter(start_col + 1)

    for i, env_key in enumerate(env_keys):
        col, col_letter = start_col + 2 + i, get_column_letter(start_col + 2 + i)
        available_ram = AVAILABLE_RAM_GB.get(env_key, 0)
        write_cell(sheet, available_ram_row, col, available_ram, is_summary=False, number_format='0')
        
        allocated_ram_value = total_ram_by_host.get(env_key, 0)
        write_cell(sheet, allocated_ram_row, col, allocated_ram_value, is_summary=False, number_format='0.0')
        
        formula_pct = f"=IF({col_letter}{available_ram_row}>0, {col_letter}{allocated_ram_row}/{col_letter}{available_ram_row}, 0)"
        write_cell(sheet, allocated_pct_row, col, formula_pct, is_summary=False, number_format="0%")
        
        formula_rem = f"={col_letter}{available_ram_row}-{col_letter}{allocated_ram_row}"
        write_cell(sheet, remaining_ram_row, col, formula_rem, is_summary=False, number_format='0.0')
        
    first_env, last_env = get_column_letter(start_col + 2), get_column_letter(start_col + 1 + len(env_keys))
    write_cell(sheet, available_ram_row, start_col + 1, f"=SUM({first_env}{available_ram_row}:{last_env}{available_ram_row})", number_format='0')
    write_cell(sheet, allocated_ram_row, start_col + 1, f"=SUM({first_env}{allocated_ram_row}:{last_env}{allocated_ram_row})", number_format='0.0')
    write_cell(sheet, allocated_pct_row, start_col + 1, f"=IF({total_col_letter}{available_ram_row}>0, {total_col_letter}{allocated_ram_row}/{total_col_letter}{available_ram_row}, 0)", number_format="0%")
    write_cell(sheet, remaining_ram_row, start_col + 1, f"={total_col_letter}{available_ram_row}-{total_col_letter}{allocated_ram_row}", number_format='0.0')
    
    for r in [allocated_ram_row, allocated_pct_row, remaining_ram_row]:
        for c in range(start_col + 1, start_col + 2 + len(env_keys)):
            cell = sheet.cell(row=r, column=c)
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            cell.font = Font(bold=True, color="FF0000")

    # ✅ Add clean outer border around entire RAM summary block
    end_row = remaining_ram_row
    end_col = start_col + 1 + len(env_keys)

    apply_outer_border(
        sheet,
        start_row=row_offset,
        end_row=end_row,
        start_col=start_col,
        end_col=end_col
    )

def write_overview_summary(sheet, trainer_pods, standard_pods, extended_pods):
    start_row = 1
    col = 1

    trainer_pods_count = sum(convert_to_numeric(p.get('vendor_pods') or 0) for p in trainer_pods)
    extended_pods_count = sum(convert_to_numeric(p.get('vendor_pods') or 0) for p in extended_pods)
    standard_pods_count = sum(convert_to_numeric(p.get('vendor_pods') or 0) for p in standard_pods)

    total_pods = trainer_pods_count + extended_pods_count + standard_pods_count

    total_ram = 0
    for pod in trainer_pods + standard_pods + extended_pods:
        ram = convert_to_numeric(pod.get('ram')) or 0
        pods = convert_to_numeric(pod.get('vendor_pods')) or 1
        if pods <= 1:
            total_ram += ram
        else:
            total_ram += ram + (pods - 1) * ram / 2

    total_available_ram = sum(AVAILABLE_RAM_GB.values())
    total_ram_pct = round((total_ram / total_available_ram) * 100) if total_available_ram else 0

    summary_data = [
        ("Trainer Pods (Blue)", trainer_pods_count, ExcelStyle.LIGHT_BLUE_FILL.value),
        ("Extended Pods (Green)", extended_pods_count, ExcelStyle.GREEN_FILL.value),
        ("Standard Pods", standard_pods_count, None),
        ("Total Pods", total_pods, None),
        ("Total RAM (%)", f"{total_ram_pct}%", None)
    ]

    row = start_row
    section_title_cell = sheet.cell(row=row, column=col, value="Overview Summary")
    section_title_cell.font = Font(bold=True, size=14)
    section_title_cell.alignment = Alignment(horizontal='left')
    sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+1)
    row += 1

    for label, value, fill in summary_data:
        label_cell = sheet.cell(row=row, column=col, value=label)
        label_cell.alignment = Alignment(horizontal='left')
        label_cell.font = Font(bold=True)

        value_cell = sheet.cell(row=row, column=col + 1, value=value)
        value_cell.alignment = Alignment(horizontal='center')

        if fill:
            label_cell.fill = fill
            value_cell.fill = fill

        # ✅ MODIFIED color logic for "Total RAM (%)" based on new thresholds
        if label == "Total RAM (%)":
            percent_value = int(str(value).replace("%", "").strip())
            if percent_value > 90:
                color = "FF0000"  # Red
            elif percent_value >= 80:
                color = "FFFF00"  # Yellow
            else:
                color = "92D050"  # Green

            value_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        row += 1

    end_row = row - 1
    thin = Side(style='thin')
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for r in range(start_row, end_row + 1):
        for c in range(col, col + 2):
            cell = sheet.cell(row=r, column=c)
            cell.border = border

def _write_data_row(sheet, row, entry, headers, header_keys, header_pos, is_trainer, use_green_fill, host_map: Dict):
    trainer_flag, green_flag = (True, False) if is_trainer else (False, use_green_fill)
    virtual_hosts_str = entry.get("virtual_hosts", "")
    virtual_hosts_list = [h.strip().lower() for h in virtual_hosts_str.split(",") if h.strip()]
    col = 1

    for h in header_keys:
        if h == "Start/End Pod":
            pod = entry.get("start_end_pod", "")
            parts = pod.replace("→", "->").replace("–", "-").split("-")
            left, right = (parts[0].strip(), parts[1].strip()) if len(parts) > 1 else (parts[0].strip() if parts else "", "")
            right = right if right else left
            write_cell(sheet, row, col, left, trainer=trainer_flag, use_green_fill=green_flag); col += 1
            write_cell(sheet, row, col, "->", trainer=trainer_flag, use_green_fill=green_flag); col += 1
            write_cell(sheet, row, col, right, trainer=trainer_flag, use_green_fill=green_flag); col += 1
        elif h in host_map:
            host_name_for_this_column = host_map[h].lower().strip()
            if any(host_name_for_this_column == vh for vh in virtual_hosts_list):
                ram_col, pods_col = get_column_letter(header_pos["RAM"]), get_column_letter(header_pos["Vendor Pods"])
                formula = f"=IF({pods_col}{row}<=1, {ram_col}{row}, {ram_col}{row} + ({pods_col}{row}-1)*{ram_col}{row}/2)"
                write_cell(sheet, row, col, formula, trainer_flag, green_flag, number_format='0.0')
            else:
                write_cell(sheet, row, col, 0, trainer_flag, green_flag, number_format='0.0')
            col += 1
        else:
            data_key = headers.get(h)
            val_to_write = entry.get(data_key) if data_key else ""
            num_format = '0.0' if h == "RAM" else None
            write_cell(sheet, row, col, val_to_write, trainer_flag, green_flag, number_format=num_format)
            col += 1

# upcoming_report_generator.py

def write_group_summary_boxes(sheet, start_row, header_pos, group_pod_total_or_formula, group_host_ram_totals, group_name, group_end_col, data_start_row, data_end_row):
    vendor_col = header_pos.get("Vendor Pods")
    students_col = header_pos.get("Students")
    has_pod_summary = vendor_col and students_col and (group_pod_total_or_formula is not None)
    has_ram_summary = sum(group_host_ram_totals.values()) > 0
    if not (has_pod_summary or has_ram_summary): return start_row
    
    label_start, label_end = start_row, start_row + 1
    value_start, value_end = start_row + 2, start_row + 3
    summary_end = value_end
    thin, no_side = Side(style='thin'), Side(style=None)
    border_top, border_bottom = Border(left=thin, right=thin, top=thin, bottom=no_side), Border(left=thin, right=thin, top=no_side, bottom=thin)
    for r in range(start_row, summary_end + 1):
        for c in range(1, group_end_col + 1):
            cell = sheet.cell(row=r, column=c)
            cell.alignment = ExcelStyle.CENTER_ALIGNMENT.value
            cell.border = border_top if r in [label_start, value_start] else border_bottom
    
    if has_pod_summary:
        merge_start = students_col
        merge_end = vendor_col
        
        sheet.merge_cells(start_row=label_start, start_column=merge_start, end_row=label_end, end_column=merge_end)
        label_cell = sheet.cell(row=label_start, column=merge_start, value="Total Pods")
        label_cell.font = Font(bold=True, size=20)
        label_cell.alignment = Alignment(horizontal='center', vertical='center')
        label_cell.fill = ExcelStyle.LIGHT_BLUE_FILL.value  # <-- MODIFIED

        sheet.merge_cells(start_row=value_start, start_column=merge_start, end_row=value_end, end_column=merge_end)
        
        value_cell = sheet.cell(row=value_start, column=merge_start, value=group_pod_total_or_formula)
        value_cell.font = Font(bold=True, size=20)
        value_cell.alignment = Alignment(horizontal='center', vertical='center')
        value_cell.fill = ExcelStyle.LIGHT_BLUE_FILL.value   # <-- MODIFIED

    if has_ram_summary:
        for host_key, host_ram in group_host_ram_totals.items():
            if host_col := header_pos.get(host_key):
                cell_top, cell_bottom = sheet.cell(row=value_start, column=host_col), sheet.cell(row=value_end, column=host_col)
                
                # ✅ Create a SUM formula for the host's RAM column for the current group
                host_col_letter = get_column_letter(host_col)
                cell_bottom.value = f"=SUM({host_col_letter}{data_start_row}:{host_col_letter}{data_end_row})" if data_start_row <= data_end_row else 0.0
                
                cell_bottom.number_format = '0.0'
                cell_bottom.font = Font(bold=True, color="FFFF00")
                cell_top.fill = ExcelStyle.LIGHT_BLUE_FILL.value
                cell_bottom.fill = ExcelStyle.LIGHT_BLUE_FILL.value
    return start_row + 4

def apply_outer_border(sheet, start_row, end_row, start_col, end_col):
    thin_side = Side(style='thin')
    medium_side = Side(style='medium')

    def set_border_side(cell, top=None, left=None, right=None, bottom=None):
        border = cell.border.copy()
        if top: border.top = top
        if left: border.left = left
        if right: border.right = right
        if bottom: border.bottom = bottom
        cell.border = border

    for col in range(start_col, end_col + 1):
        set_border_side(sheet.cell(row=start_row, column=col), top=thin_side)
        set_border_side(sheet.cell(row=end_row, column=col), bottom=medium_side)
    
    for row in range(start_row, end_row + 1):
        set_border_side(sheet.cell(row=row, column=start_col), left=medium_side)
        set_border_side(sheet.cell(row=row, column=end_col), right=medium_side)

def determine_us_au_location(virtual_hosts_str: str) -> str:
    if not virtual_hosts_str: return ""
    hosts_lower = virtual_hosts_str.lower()
    if any(au in hosts_lower for au in AU_HOST_NAMES): return "AU"
    if any(us in hosts_lower for us in US_HOST_NAMES): return "US"
    return ""

def format_date(date_str):
    if not date_str: return ""
    try: return datetime.strptime(date_str, '%Y-%m-%d').strftime('%A, %d/%m')
    except Exception: return date_str

def get_vendor_prefix_map():
    return {gn.split(' ')[0].lower(): gn for gn in ExcelStyle.DEFAULT_COURSE_GROUPS.value.keys()}

def find_location_from_code(course_code: str, location_map: dict) -> str:
    if not course_code: return ""
    for loc_code in sorted(location_map.keys(), key=len, reverse=True):
        if loc_code in course_code: return location_map[loc_code]
    return ""

def _find_value_across_docs(docs: list, key_priority_list: list):
    for key in key_priority_list:
        for doc in docs:
            value = doc.get(key)
            if value is not None and value != '':
                return value
    return None

# ==============================================================================
# DATA UNPACKING FUNCTIONS
# ==============================================================================
def extract_apm_command(apm_cmds):
    if apm_cmds and isinstance(apm_cmds, list):
        line = apm_cmds[0]
        parts = re.findall(r'"([^"]*)"', line)
        if len(parts) >= 4:
            return parts[-4]
    return ""
    
# In labbuild/dashboard/upcoming_report_generator.py
# Final version - handles both flat (your sample) and nested 'assignments' structures.

# In labbuild/dashboard/upcoming_report_generator.py
# Final version - fixes grouping for consolidated trainer pods.

def unpack_interim_allocations(documents, vendor_map, location_map, ram_lookup_map, host_vcenter_map):
    logger.info("Starting unpack_interim_allocations process...")

    # Data Normalization Step (handles flat vs. nested structure)
    for doc in documents:
        if not doc.get('assignments'):
            host = doc.get('host')
            start_pod = doc.get('start_pod')
            if host and start_pod is not None:
                doc['assignments'] = [{
                    'host': host,
                    'start_pod': start_pod,
                    'end_pod': doc.get('end_pod', start_pod)
                }]
                logger.debug(f"Normalized flat allocation for doc with tag '{doc.get('tag')}' into assignments array.")

    standard_docs_raw = []
    trainer_docs_raw = []

    for doc in documents:
        course_type = str(doc.get('sf_course_type', '')).lower()
        course_code = str(doc.get('sf_course_code', '')).lower()
        if 'trainer pod' in course_type or 'trainer pod' in course_code:
            trainer_docs_raw.append(doc)
        else:
            standard_docs_raw.append(doc)
            
    logger.info(f"Separated documents: {len(standard_docs_raw)} standard, {len(trainer_docs_raw)} trainer.")

    processed_trainer_pods = []
    for doc in trainer_docs_raw:
        ram = doc.get('memory_gb_one_pod')
        course_version = doc.get('final_labbuild_course') or doc.get('labbuild_course')
        if not ram and course_version in ram_lookup_map:
            ram = ram_lookup_map.get(course_version)

        username = doc.get('student_apm_username') or doc.get('apm_username') or doc.get('username')
        password = doc.get('student_apm_password') or doc.get('apm_password') or doc.get('password')
        
        assignments = doc.get('assignments', [])
        hosts = sorted(set(a.get('host') for a in assignments if a.get('host')))
        host_str = ", ".join(hosts)
        vcenters = sorted(list(set(host_vcenter_map.get(h.lower(), '') for h in hosts if h)))
        vcenter_str = ", ".join(filter(None, vcenters))

        pod_ranges = []
        pod_count = 0
        for a in assignments:
            start, end = a.get('start_pod'), a.get('end_pod', a.get('start_pod'))
            if start is not None:
                pod_ranges.append(str(start) if end is None or start == end else f"{start}-{end}")
                pod_count += (int(end) - int(start) + 1)

        # ======================================================================
        # MODIFIED LOGIC: Intelligently determine the grouping code.
        # This is the key fix.
        # ======================================================================
        related_courses = doc.get('related_student_courses', [])
        sf_code = doc.get('sf_course_code', '')
        grouping_code = ''

        if related_courses:
            grouping_code = related_courses[0]
        elif sf_code.lower().startswith("consolidated for:"):
            # If it's a consolidated code, parse it to get a real course code for grouping.
            # E.g., "Consolidated for: PA330VEU250820, ..." -> "PA330VEU250820"
            code_part = sf_code[17:].strip() # Get everything after "Consolidated for:"
            first_code = code_part.split(',')[0].strip()
            grouping_code = first_code
        else:
            # Fallback to the original code if it's not a consolidated string.
            grouping_code = sf_code
        # ======================================================================

        trainer_pod_entry = {
            'course_code': grouping_code,  # Use the newly determined grouping_code
            'location': "",
            'us_au_location': determine_us_au_location(host_str),
            'course_start_date': format_date(doc.get('sf_start_date') or doc.get('start_date')),
            'last_day': format_date(doc.get('sf_end_date') or doc.get('end_date')),
            'trainer_name': doc.get('sf_trainer_name') or doc.get('trainer_name') or "Trainer",
            'course_name': doc.get('sf_course_type', ''),
            'start_end_pod': ", ".join(sorted(pod_ranges)),
            'username': username,
            'password': password,
            'class_number': doc.get('f5_class_number'),
            'students': len(related_courses),
            'vendor_pods': doc.get('effective_pods_req') or pod_count or 0,
            'ram': ram,
            'virtual_hosts': host_str,
            'vcenter': vcenter_str,
            'pod_type': 'trainer',
            'version': course_version,
            'course_version': course_version,
            'apm_command_value': extract_apm_command(doc.get('apm_commands', [])) or course_version,
        }
        processed_trainer_pods.append(trainer_pod_entry)

    # ... (The rest of the function remains the same) ...

    grouped_courses = defaultdict(lambda: {'docs': [], 'assignments': []})
    for doc in standard_docs_raw:
        if code := doc.get('sf_course_code'):
            grouped_courses[code]['docs'].append(doc)
            new_assignments = doc.get('assignments', [])
            existing_assignments = grouped_courses[code]['assignments']
            for a in new_assignments:
                if a not in existing_assignments:
                    existing_assignments.append(a)

    processed_standard_courses = []
    logger.info(f"Processing {len(grouped_courses)} unique standard course codes.")
    for code, data in grouped_courses.items():
        base_doc = data['docs'][0]
        all_docs = data['docs']

        final_username = _find_value_across_docs(all_docs, ['student_apm_username', 'apm_username', 'username'])
        final_password = _find_value_across_docs(all_docs, ['student_apm_password', 'apm_password', 'password'])
        final_ram = _find_value_across_docs(all_docs, ['memory_gb_one_pod', 'ram'])
        
        course_version = base_doc.get('final_labbuild_course') or base_doc.get('labbuild_course')
        if not final_ram and course_version in ram_lookup_map:
            final_ram = ram_lookup_map.get(course_version)

        hosts_from_assignments = sorted(set(a.get('host') for a in data['assignments'] if a.get('host')))
        final_host_str = ", ".join(hosts_from_assignments) or _find_value_across_docs(all_docs, ['virtual_hosts']) or ""

        final_hosts_list = [h.strip() for h in final_host_str.split(',') if h.strip()]
        vcenters = sorted(list(set(host_vcenter_map.get(h.lower(), '') for h in final_hosts_list if h)))
        final_vcenter_str = ", ".join(filter(None, vcenters))

        pod_ranges = []
        pod_count = 0
        for a in data['assignments']:
            start, end = a.get('start_pod'), a.get('end_pod', a.get('start_pod'))
            if start is not None:
                pod_ranges.append(str(start) if end is None or start == end else f"{start}-{end}")
                pod_count += (int(end) - int(start) + 1)
        
        us_au_loc = determine_us_au_location(final_host_str)

        course = {
            'course_code': code,
            'location': find_location_from_code(code, location_map),
            'us_au_location': us_au_loc,
            'course_start_date': format_date(base_doc.get('sf_start_date') or base_doc.get('start_date')),
            'last_day': format_date(base_doc.get('sf_end_date') or base_doc.get('end_date')),
            'trainer_name': base_doc.get('sf_trainer_name') or base_doc.get('trainer_name'),
            'course_name': base_doc.get('sf_course_type'),
            'start_end_pod': ", ".join(sorted(pod_ranges)),
            'username': final_username,
            'password': final_password,
            'class_number': base_doc.get('f5_class_number'),
            'students': base_doc.get('sf_pax_count', 0),
            'vendor_pods': base_doc.get('effective_pods_req') or pod_count or 0,
            'ram': final_ram,
            'virtual_hosts': final_host_str,
            'vcenter': final_vcenter_str,
            'pod_type': 'default',
            'version': course_version,
            'course_version': course_version,
            'apm_command_value': extract_apm_command(base_doc.get('apm_commands', [])) or course_version,
        }
        processed_standard_courses.append(course)

    logger.info(f"Finished processing. Returning {len(processed_standard_courses)} standard courses and {len(processed_trainer_pods)} trainer pods.")
    return processed_standard_courses, processed_trainer_pods



def unpack_extended_allocations(documents: List[Dict], location_map: Dict, ram_lookup_map: Dict, host_vcenter_map: Dict) -> List[Dict]:
    extended_courses = []
    logger.info(f"Unpacking {len(documents)} extended allocations from currentallocation.")
    for doc in documents:
        course_details = doc.get('courses', [{}])[0]
        if not course_details:
            logger.warning(f"Skipping extended allocation doc with _id {doc.get('_id')} due to missing 'courses' data.")
            continue
        pod_details_list = doc.get('pod_details') or course_details.get('pod_details', [])
        pod_numbers = sorted([p.get('pod_number') for p in pod_details_list if p.get('pod_number') is not None])
        hosts = sorted(list(set(p.get('host') for p in pod_details_list if p.get('host'))))
        vcenters = sorted(list(set(host_vcenter_map.get(h.lower(), '') for h in hosts if h)))
        final_vcenter_str = ", ".join(filter(None, vcenters))
        start_end_pod = ""
        if pod_numbers:
            start_end_pod = f"{pod_numbers[0]}-{pod_numbers[-1]}" if len(pod_numbers) > 1 else str(pod_numbers[0])
            
        course_version = course_details.get('course_name')
        ram_value = course_details.get('memory_gb_one_pod') or course_details.get('ram')
        if not ram_value:
            if course_version in ram_lookup_map:
                ram_value = ram_lookup_map.get(course_version)
        course_code = doc.get('tag', '')
        course = {
            'course_code': course_code,
            'location': find_location_from_code(course_code, location_map),
            'us_au_location': determine_us_au_location(", ".join(hosts)),
            'course_start_date': format_date(course_details.get('start_date')),
            'last_day': format_date(course_details.get('end_date')),
            'trainer_name': course_details.get('trainer_name'),
            'course_name': course_version,
            'start_end_pod': start_end_pod,
            'username': course_details.get('apm_username'),
            'password': course_details.get('apm_password'),
            'class_number': None,
            'students': len(pod_numbers),
            'vendor_pods': len(pod_numbers),
            'ram': ram_value,
            'virtual_hosts': ", ".join(hosts),
            'vcenter': final_vcenter_str,
            'pod_type': 'extended',
            'version': course_version,
            'course_version': course_version,
            # ✅ MODIFIED: Fallback to course version if APM command is not found
            'apm_command_value': extract_apm_command(doc.get('apm_commands', [])) or course_version,
        }
        extended_courses.append(course)
    return extended_courses


# ==============================================================================
# EXCEL GENERATION & MAIN ORCHESTRATION
# ==============================================================================
# labbuild/dashboard/upcoming_report_generator.py

# labbuild/dashboard/upcoming_report_generator.py

def generate_excel_in_memory(course_allocations: List[Dict], trainer_pods: List[Dict], extended_pods: List[Dict], host_map: Dict) -> io.BytesIO:
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Upcoming Labs"

    all_data = course_allocations + trainer_pods + extended_pods

    grouped = {g: [] for g in ExcelStyle.DEFAULT_COURSE_GROUPS.value}
    for entry in all_data:
        for gname, fn in ExcelStyle.DEFAULT_COURSE_GROUPS.value.items():
            if fn(entry.get("course_code", "")):
                grouped[gname].append(entry)
                break

    total_ram_by_host = {k: 0 for k in host_map}

    group_order, current_row = EXCEL_GROUP_ORDER, 12
    for group_name in group_order:
        records = grouped.get(group_name, [])
        if not records:
            continue

        current_row += 1
        group_start_row = current_row

        logger.info(f"Writing group: {group_name} with {len(records)} records")
        current_row = write_group_title(sheet, current_row, group_name)
        
        headers = ExcelStyle.DEFAULT_HEADER_COURSE_MAPPING.value[group_name].copy()

        if "Course Version" in headers:
            new_headers = {}
            for header, data_key in headers.items():
                if header == "Course Version":
                    new_headers["vCenter"] = "vcenter"
                else:
                    new_headers[header] = data_key
            headers = new_headers
        
        if "Group" in headers:
            headers["Group"] = "apm_command_value"

        header_keys, col_idx, header_pos, group_end_col = list(headers), 1, {}, 0
        for h in header_keys:
            header_pos[h] = col_idx
            width = 3 if h == "Start/End Pod" else 1
            col_idx += width
            group_end_col += width

        col = 1
        for h in header_keys:
            if h == "Start/End Pod":
                write_merged_header(sheet, current_row, col, h)
                col += 3
            else:
                header_text = "APM Commands" if h == "Group" else h
                cell = sheet.cell(row=current_row, column=col, value=header_text)
                cell.font = Font(bold=True)
                apply_style(cell, is_header=True) 
                col += 1
        
        current_row += 1
        data_start_row = current_row

        group_host_ram_totals = {k: 0 for k in host_map}

        records.sort(key=lambda x: 1 if x.get("pod_type") == "trainer" else 0)

        trainer_started = False
        for idx, entry in enumerate(records):
            is_trainer = entry.get("pod_type") == "trainer"
            is_extended = entry.get("pod_type") == "extended"

            if is_trainer and not trainer_started:
                if idx > 0 and any(r.get("pod_type") == "extended" for r in records[:idx]):
                    current_row += 1
                trainer_started = True

            _write_data_row(
                sheet, current_row, entry,
                headers, header_keys, header_pos,
                is_trainer, is_extended, host_map
            )

            if (entry_ram := convert_to_numeric(entry.get("ram")) or 0) > 0:
                for host_key, host_name in host_map.items():
                    if host_name.lower() in (entry.get("virtual_hosts", "") or "").lower():
                        total_pods_for_entry = convert_to_numeric(entry.get("vendor_pods")) or 1
                        if total_pods_for_entry <= 1:
                            group_host_ram_totals[host_key] += entry_ram
                        else:
                            group_host_ram_totals[host_key] += entry_ram + (total_pods_for_entry - 1) * entry_ram / 2
            
            current_row += 1

        data_end_row = current_row - 1

        for host_key, ram in group_host_ram_totals.items():
            total_ram_by_host[host_key] += ram

        pod_total_formula = 0
        if (students_col_num := header_pos.get("Students")) and data_start_row <= data_end_row:
            students_col_letter = get_column_letter(students_col_num)
            pod_total_formula = f"=SUM({students_col_letter}{data_start_row}:{students_col_letter}{data_end_row})"

        summary_start_row = current_row
        current_row = write_group_summary_boxes(
            sheet, summary_start_row, header_pos,
            pod_total_formula, group_host_ram_totals,
            group_name, group_end_col, data_start_row, data_end_row
        )
        end_row = current_row - 1 if summary_start_row < current_row else summary_start_row - 1

        apply_outer_border(sheet, group_start_row + 1, end_row, 1, group_end_col)

        current_row += 1

    # Write overview and RAM summary sections
    write_overview_summary(sheet, trainer_pods, course_allocations, extended_pods)
    write_summary_section(sheet, 2, total_ram_by_host)

    # ✅ MODIFIED: Patch "Allocated RAM (GB)" row in RAM Summary section with new universal formula
    allocated_ram_row = 4  # Adjust if your summary section starts elsewhere
    courses_data_start_row = 15  # Adjust if your first courses section starts at another row

    for i, env_key in enumerate(SUMMARY_ENV_ORDER):
        col_idx = RAM_SUMMARY_START_COL + 2 + i
        col_letter = get_column_letter(col_idx)
        
        # Apply the new formula universally to all columns in the summary
        formula = f"=SUM({col_letter}{courses_data_start_row}:{col_letter}1002)/2"
        
        cell = sheet.cell(row=allocated_ram_row, column=col_idx)
        cell.value = formula
        cell.number_format = '0.0'

    for col_letter, width in EXCEL_COLUMN_WIDTHS.items():
        sheet.column_dimensions[col_letter].width = width

    in_memory_fp = io.BytesIO()
    wb.save(in_memory_fp)
    in_memory_fp.seek(0)
    return in_memory_fp

def get_upcoming_report_data(db):
    try:
        if db is None:
            raise ConnectionError("A valid database connection was not provided.")

        # --- Data Fetching ---
        logger.info("Fetching data for current/staged labs from INTERIM_ALLOCATION_COLLECTION...")
        interim_docs = list(db[INTERIM_ALLOCATION_COLLECTION].find({}))

        logger.info("Fetching extended courses from ALLOCATION_COLLECTION...")
        extended_docs = list(db[ALLOCATION_COLLECTION].find({"extend": "true"}))
        logger.info(f"Found {len(extended_docs)} allocations marked for extension.")

        # ✅ FIXED: Fetch future-dated, non-extended courses from the main allocation collection to capture all upcoming labs.
        today_str = datetime.now().strftime('%Y-%m-%d')
        logger.info(f"Fetching upcoming labs from ALLOCATION_COLLECTION starting from {today_str}...")
        upcoming_query = {
            "courses.0.start_date": {"$gte": today_str},
            "extend": {"$ne": "true"}
        }
        upcoming_docs_from_alloc = list(db[ALLOCATION_COLLECTION].find(upcoming_query))
        logger.info(f"Found {len(upcoming_docs_from_alloc)} upcoming non-extended allocations.")

        # --- Lookup Data ---
        logger.info("Fetching course configs for RAM fallback...")
        course_configs = list(db[COURSE_CONFIG_COLLECTION].find({}, {"course_name": 1, "memory": 1}))
        ram_lookup_map = {doc['course_name']: doc['memory'] for doc in course_configs if 'course_name' in doc and 'memory' in doc}
        logger.info(f"Built RAM lookup map with {len(ram_lookup_map)} entries.")

        locations_data = list(db["locations"].find({}))
        location_map = {loc['code']: loc['name'] for loc in locations_data if 'code' in loc and 'name' in loc}

        host_docs = list(db[HOST_COLLECTION].find({}))
        host_vcenter_map = {}
        logger.info("Building robust host-to-vCenter lookup map...")
        for doc in host_docs:
            vcenter_fqdn = doc.get('vcenter') or doc.get('vcentre')
            if not (vcenter_fqdn and isinstance(vcenter_fqdn, str)):
                continue
            short_vcenter = vcenter_fqdn.split('.')[0]
            for key in ['host_name', 'fqdn', 'host_shortcode']:
                if identifier := doc.get(key):
                    if isinstance(identifier, str):
                        host_vcenter_map[identifier.lower()] = short_vcenter
        logger.info(f"Built vCenter map with {len(host_vcenter_map)} keys.")

        vendor_map = get_vendor_prefix_map()
        host_map_for_summary = {
            doc['host_shortcode'].capitalize(): doc['host_name']
            for doc in host_docs
            if 'host_shortcode' in doc and 'host_name' in doc and str(doc.get('include_for_build')).lower() == 'true'
        }

        # --- Data Unpacking & Merging ---

        # Unpack from interim collection first (these are often high-priority or currently staging)
        course_allocs, trainer_pods = unpack_interim_allocations(
            interim_docs, vendor_map, location_map, ram_lookup_map, host_vcenter_map
        )

        # Unpack the newly fetched upcoming courses from the main allocation collection
        upcoming_allocs = unpack_extended_allocations(
            upcoming_docs_from_alloc, location_map, ram_lookup_map, host_vcenter_map
        )

        # Unpack extended courses as before
        extended_pods = unpack_extended_allocations(
            extended_docs, location_map, ram_lookup_map, host_vcenter_map
        )

        # ✅ FIXED: Merge and de-duplicate, giving interim-sourced data priority to prevent double-counting.
        interim_course_codes = {c['course_code'] for c in course_allocs}
        unique_upcoming_allocs = [
            alloc for alloc in upcoming_allocs
            if alloc.get('course_code') not in interim_course_codes
        ]
        course_allocs.extend(unique_upcoming_allocs)

        logger.info(f"Report Data Summary: {len(course_allocs)} Standard/Upcoming Courses | {len(trainer_pods)} Trainer Pods | {len(extended_pods)} Extended Pods")
        return course_allocs, trainer_pods, extended_pods, host_map_for_summary

    except Exception as e:
        logger.error(f"FATAL: An error occurred during upcoming report data fetching: {e}", exc_info=True)
        raise e